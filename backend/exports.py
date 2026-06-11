"""Excel + PDF export helpers."""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak


CATEGORIES = {
    "contacts": {
        "title": "Contacts",
        "headers": ["Name", "Company", "Phone", "Email", "Address", "City", "State", "ZIP", "Billing Address", "Billing City", "Billing State", "Billing ZIP"],
        "keys": ["contact_name", "company_name", "phone", "email", "address", "city", "state", "zip_code", "billing_address", "billing_city", "billing_state", "billing_zip"],
    },
    "properties": {
        "title": "Properties",
        "headers": ["Property", "Address", "Line 2", "City", "State", "ZIP", "On-Site Contact", "Phone", "Notes"],
        "keys": ["property_name", "property_address", "property_address_line2", "property_city", "property_state", "property_zip", "property_contact_name", "property_contact_phone", "notes"],
    },
    "projects": {
        "title": "Projects",
        "headers": ["Title", "Type", "Status", "Lead Source", "Referral", "Project Type", "Current Roof", "Proposed Roof", "Option A", "Option B", "Option C", "Chosen Amount", "Chosen Date", "Date Sent", "Materials", "Labor", "Subcontractor", "Other", "Profit"],
        "keys": ["title", "deal_type", "status", "lead_source", "referral_source", "project_type", "current_roof_type", "proposed_roof_type", "proposal_option_1", "proposal_option_2", "proposal_option_3", "chosen_amount", "chosen_date", "date_sent", "materials_cost", "labor_cost", "subcontractor_cost", "other_expenses", "_profit"],
    },
    "vendors": {
        "title": "Vendors",
        "headers": ["Name", "Category", "Phone", "Email", "TIN/EIN", "Address", "City", "State", "ZIP"],
        "keys": ["name", "category", "phone", "email", "tin_ein", "address", "city", "state", "zip_code"],
    },
    "subcontractors": {
        "title": "Subcontractors",
        "headers": ["Name", "Category", "Phone", "Email", "TIN/EIN", "Address", "City", "State", "ZIP"],
        "keys": ["name", "category", "phone", "email", "tin_ein", "address", "city", "state", "zip_code"],
    },
}


def _row_values(rec: dict, keys):
    out = []
    for k in keys:
        if k == "_profit":
            costs = float(rec.get("materials_cost", 0) or 0) + float(rec.get("labor_cost", 0) or 0) + float(rec.get("subcontractor_cost", 0) or 0) + float(rec.get("other_expenses", 0) or 0)
            out.append(round(float(rec.get("chosen_amount", 0) or 0) - costs, 2))
        else:
            v = rec.get(k, "")
            out.append(v if v is not None else "")
    return out


def to_excel(sections: list) -> bytes:
    """sections: list of (category_key, records). Returns xlsx bytes."""
    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    for key, records in sections:
        cfg = CATEGORIES[key]
        ws = wb.create_sheet(cfg["title"])
        ws.append(cfg["headers"])
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
        for rec in records:
            ws.append(_row_values(rec, cfg["keys"]))
        # widen columns
        for col_idx, hdr in enumerate(cfg["headers"], 1):
            max_len = len(str(hdr))
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=1, values_only=True):
                for v in cell:
                    max_len = max(max_len, min(len(str(v) if v is not None else ""), 40))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 2
        ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_pdf(sections: list) -> bytes:
    """sections: list of (category_key, records). Returns pdf bytes."""
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter), leftMargin=0.4 * inch, rightMargin=0.4 * inch, topMargin=0.5 * inch, bottomMargin=0.5 * inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=18, textColor=colors.HexColor("#0A0A0A"))
    eyebrow = ParagraphStyle("eyebrow", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=8, textColor=colors.HexColor("#1D4ED8"), leading=10, letterSpacing=1)
    body = ParagraphStyle("body", parent=styles["Normal"], fontName="Helvetica", fontSize=8, textColor=colors.HexColor("#27272A"))

    story = []
    for idx, (key, records) in enumerate(sections):
        cfg = CATEGORIES[key]
        story.append(Paragraph("SEALTECH CRM EXPORT", eyebrow))
        story.append(Paragraph(cfg["title"], title_style))
        story.append(Spacer(1, 0.15 * inch))
        # Table
        headers = cfg["headers"]
        data = [headers]
        for rec in records:
            row = [Paragraph(str(v if v not in (None, "") else "—"), body) for v in _row_values(rec, cfg["keys"])]
            data.append(row)
        if len(data) == 1:
            story.append(Paragraph("No records.", body))
        else:
            tbl = Table(data, repeatRows=1)
            tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("TOPPADDING", (0, 0), (-1, 0), 8),
                ("ALIGN", (0, 0), (-1, 0), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E4E4E7")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAFA")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(tbl)
        if idx < len(sections) - 1:
            story.append(PageBreak())
    doc.build(story)
    return buf.getvalue()
