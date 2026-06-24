from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import logging
import uuid
import asyncio
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Literal, Dict, Any

import bcrypt
import jwt
import csv
import io
from io import BytesIO
import smtplib
import secrets
import string
import re
from openpyxl import load_workbook
from fastapi import FastAPI, APIRouter, Depends, HTTPException, Request, status, UploadFile, File, Form, Query, Header, Body
from fastapi.responses import Response, StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr, ConfigDict

from storage import init_storage, put_object, get_object, APP_NAME
from exports import to_excel, to_pdf, CATEGORIES as EXPORT_CATEGORIES
from spec_sheet import build_spec_sheet
from books import make_router as make_books_router, seed_default_entities
import gl
import coi_reminders
import project_photos
import trash
import assessment as assessment_module


# ----- DB -----
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]


# ----- App -----
app = FastAPI(title="Roofing CRM API")
api_router = APIRouter(prefix="/api")

JWT_ALGORITHM = "HS256"
ACCESS_TOKEN_MINUTES = 60 * 24  # 24h for convenience


# ----- Helpers -----
def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def now_iso() -> str:
    return now_utc().isoformat()


def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))


def get_jwt_secret() -> str:
    return os.environ["JWT_SECRET"]


def create_access_token(user_id: str, email: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "exp": now_utc() + timedelta(minutes=ACCESS_TOKEN_MINUTES),
        "type": "access",
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)


async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"id": payload["sub"]})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        user.pop("_id", None)
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


def strip_id(doc: dict) -> dict:
    if doc is None:
        return None
    doc.pop("_id", None)
    return doc


# ----- Models -----
LEAD_SOURCES = ["Referral", "Marketing", "Website", "Email Campaign", "Personal"]
PROJECT_TYPES = ["Repair", "Roof Restoration", "Roof Replacement", "Maintenance", "New Construction", "Other"]
ROOF_TYPES = [
    "FARM (Fluid Applied Reinforced Membrane)",
    "Silicone w/ Granules",
    "Silicone",
    "Tile",
    "Shingle",
    "Metal",
    "BUR (Built-Up)",
    "ModBit Over-Lay",
    "ModBit Replacement",
    "ModBit",
    "EPDM Over-Lay",
    "EPDM Replacement",
    "EPDM w/ Ballast",
    "EPDM",
    "PVC Over-Lay",
    "PVC Replacement",
    "PVC",
    "TPO Over-Lay",
    "TPO Replacement",
    "TPO",
    "Construction Project",
    "Other",
]

# CURRENT roof type dropdown — what's on the building today.
# Excludes Over-Lay / Replacement variants (those are install-method labels, not existing surfaces).
# Includes "None" for new construction and "Other Construction Work" for non-roofing scopes.
CURRENT_ROOF_TYPES = [
    "None (new construction)",
    "FARM (Fluid Applied Reinforced Membrane)",
    "Silicone w/ Granules",
    "Silicone",
    "Tile",
    "Shingle",
    "Metal",
    "BUR (Built-Up)",
    "ModBit",
    "EPDM w/ Ballast",
    "EPDM",
    "PVC",
    "TPO",
    "Other Construction Work",
]
DEAL_STATUSES = ["Lead", "Sent", "Won", "Lost", "Past Lead"]
DEAL_TYPES = ["Assessment", "Scope"]
VENDOR_KINDS = ["Vendor", "Subcontractor"]
VENDOR_CATEGORIES = [
    "Material Supplier",
    "Equipment Supplier",
    "Porta Potty Supplier",
    "Dumpster Supplier",
    "Storage Container Supplier",
    "Labor",
    "Subcontractor",
    "Other",
]
COST_CATEGORIES = ["Materials", "Labor", "Subcontractor", "Other"]
MILESTONE_STATUSES = ["Pending", "Invoiced", "Paid"]
COST_ITEM_STATUSES = ["Pending", "Paid"]
CONTACT_TYPES = ["Owner", "Property Manager", "Tenant", "Other"]
DOCUMENT_CATEGORIES = ["Measurement Report", "Assessment", "Scope", "Proposal", "Invoice", "Photo", "Insurance/COI", "W-9", "Other"]
PARENT_TYPES = ["project", "vendor", "subcontractor", "contact", "property", "library"]

# Document Library taxonomy — shared resource folder for all users.
LIBRARY_TAXONOMY = [
    {"category": "SealTech Documents", "subcategories": ["Assessment & Reporting Documents", "Property Owner Guides", "Insurance & Storm Education", "Brochures"]},
    {"category": "Western Colloid", "subcategories": ["Specifications", "Safety Data", "Brochures"]},
    {"category": "Everest Systems", "subcategories": ["Specifications", "Safety Data", "Brochures"]},
    {"category": "Certificates & Credentials", "subcategories": ["Insurance / COI", "W-9", "Business License", "Manufacturer Certifications"]},
    {"category": "Contracts & Legal", "subcategories": ["Master Service Agreement", "Lien Waivers", "Change Orders", "Terms & Conditions"]},
    {"category": "Manufacturer Warranties", "subcategories": ["Sample Warranties", "Issued Warranties", "Warranty Reference"]},
    {"category": "Books", "subcategories": ["Period Close Snapshots", "Tax & Audit Packets", "Bank Statements"]},
]
IMPORT_CATEGORIES = ["contacts", "properties", "projects", "vendors", "subcontractors"]
DUPLICATE_MODES = ["skip", "update", "create"]
USER_ROLES = ["admin", "manager", "sales"]
FINANCIAL_FIELDS = ["proposal_option_1", "proposal_option_2", "proposal_option_3", "proposal_option_25yr", "chosen_amount", "materials_cost", "labor_cost", "subcontractor_cost", "other_expenses", "payment_milestones", "cost_items"]

# Late-fee policy: 1.5% per month (18% APR) on balances 30+ days past due.
LATE_FEE_MONTHLY_RATE = 0.015
LATE_FEE_GRACE_DAYS = 30
LATE_FEE_POLICY_TEXT = (
    "A late fee of 1.5% per month (18% APR) is applied to any balance more than "
    "30 days past due. Fees compound monthly and are reflected on each Statement of Account."
)


def compute_late_fee(invoice: dict, as_of, rate: float = LATE_FEE_MONTHLY_RATE) -> float:
    """Late fee for a single invoice: `rate`/month on balance_due once it crosses 30 days past due.
    `as_of` is a date. `rate` is a DECIMAL (e.g. 0.015 for 1.5%) — defaults to the global policy
    rate but callers should pass the resolved per-customer/per-entity rate when known.
    Returns dollars (rounded to 2dp). Returns 0 if not yet past the grace window or no
    due_date/balance is present."""
    try:
        bal = float(invoice.get("balance_due") or 0)
    except (TypeError, ValueError):
        bal = 0.0
    if bal <= 0.01:
        return 0.0
    due_raw = invoice.get("due_date") or invoice.get("invoice_date") or ""
    if not due_raw:
        return 0.0
    try:
        due = datetime.strptime(due_raw[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return 0.0
    days_past = (as_of - due).days
    if days_past < LATE_FEE_GRACE_DAYS:
        return 0.0
    # 30-59 days = 1 month, 60-89 = 2 months, ...
    months_past = days_past // 30
    return round(bal * rate * months_past, 2)


async def _resolve_invoice_late_fee_rate(invoice: dict) -> tuple[float, float]:
    """Look up the (decimal_rate, percent_rate) applicable to a specific invoice.
    Pulls customer override + entity default; falls back to global 1.5%."""
    from gl import resolve_late_fee_rate, resolve_late_fee_rate_pct
    ent_doc = None
    cust_doc = None
    ent_id = invoice.get("entity_id")
    if ent_id:
        ent_doc = await db.entities.find_one({"id": ent_id}, {"_id": 0, "late_fee_rate_pct": 1})
    cust_id = invoice.get("bill_to_contact_id") or invoice.get("contact_id")
    if cust_id:
        cust_doc = await db.contacts.find_one({"id": cust_id}, {"_id": 0, "late_fee_rate_pct": 1})
    return (resolve_late_fee_rate(ent_doc, cust_doc), resolve_late_fee_rate_pct(ent_doc, cust_doc))


async def _customer_statement_late_fee_rate(contact: dict, invoices: list) -> tuple[float, float]:
    """For a customer-level statement covering multiple invoices, pick a single rate:
       customer override → entity from first invoice → global 1.5%.
    Returns (decimal_rate, percent_rate)."""
    from gl import resolve_late_fee_rate, resolve_late_fee_rate_pct
    ent_doc = None
    for inv in invoices or []:
        eid = inv.get("entity_id")
        if eid:
            ent_doc = await db.entities.find_one({"id": eid}, {"_id": 0, "late_fee_rate_pct": 1})
            if ent_doc:
                break
    return (resolve_late_fee_rate(ent_doc, contact), resolve_late_fee_rate_pct(ent_doc, contact))


def proposal_mid_amount(d: dict) -> float:
    """Return the middle (median by value) of the proposal options, ignoring zeros.
    Falls back to the single non-zero value if only one is set.
    Includes the optional 4th 25-yr tier (FARM) when present.
    Used as the default tracking amount before a chosen_amount is locked in."""
    opts = []
    for i in (1, 2, 3):
        v = float(d.get(f"proposal_option_{i}", 0) or 0)
        if v > 0:
            opts.append(v)
    v25 = float(d.get("proposal_option_25yr", 0) or 0)
    if v25 > 0:
        opts.append(v25)
    if not opts:
        return 0.0
    opts.sort()
    return opts[len(opts) // 2]


def generate_password(length: int = 12) -> str:
    alphabet = string.ascii_letters + string.digits
    return "".join(secrets.choice(alphabet) for _ in range(length))


def is_admin(user: dict) -> bool:
    return user.get("role") == "admin"


def scrub_deal(doc: dict, user: dict) -> dict:
    """For non-admin viewers, hide financial fields unless they own/created the deal."""
    if is_admin(user):
        return doc
    owns = doc.get("assigned_to_user_id") == user["id"] or doc.get("created_by_user_id") == user["id"]
    if owns:
        return doc
    out = dict(doc)
    for f in FINANCIAL_FIELDS:
        if f in out:
            out[f] = [] if isinstance(out[f], list) else 0
    out["_financial_hidden"] = True
    return out


class RegisterReq(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)
    name: str = Field(min_length=1)


class LoginReq(BaseModel):
    email: EmailStr
    password: str


class UserOut(BaseModel):
    id: str
    email: str
    name: str
    role: str = "admin"
    phone: str = ""
    title: str = ""
    credentials: str = ""


class UserCreateReq(BaseModel):
    email: EmailStr
    name: str
    role: str = "sales"  # admin | manager | sales
    phone: str = ""
    title: str = ""
    credentials: str = ""


class UserUpdateReq(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    phone: Optional[str] = None
    title: Optional[str] = None
    credentials: Optional[str] = None


class TokenOut(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserOut


class ContactIn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    contact_name: str
    company_name: str = ""
    contact_type: str = "Owner"  # Owner | Property Manager | Tenant | Other
    phone: str = ""
    work_phone: str = ""
    mobile_phone: str = ""
    fax: str = ""
    email: str = ""
    address: str = ""
    address_line2: str = ""
    city: str = ""
    state: str = ""
    zip_code: str = ""
    billing_same_as_address: bool = True
    billing_address: str = ""
    billing_address_line2: str = ""
    billing_city: str = ""
    billing_state: str = ""
    billing_zip: str = ""
    website: str = ""
    late_fee_rate_pct: Optional[float] = None  # Per-customer override (None = inherit entity default)


class Contact(ContactIn):
    id: str
    created_at: str


class PropertyIn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    property_name: str
    property_address: str = ""
    property_address_line2: str = ""
    property_city: str = ""
    property_state: str = ""
    property_zip: str = ""
    property_contact_id: Optional[str] = None
    property_contact_name: str = ""
    property_contact_phone: str = ""
    notes: str = ""


class Property(PropertyIn):
    id: str
    created_at: str


class PaymentMilestone(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    label: str = ""
    percent: float = 0.0
    amount: float = 0.0
    due_date: str = ""
    status: str = "Pending"  # Pending | Invoiced | Paid
    paid_date: str = ""
    notes: str = ""


class CostItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    category: str = "Materials"  # Materials | Labor | Subcontractor | Other
    vendor_id: Optional[str] = None
    vendor_name: str = ""
    description: str = ""
    amount: float = 0.0
    date: str = ""
    status: str = "Pending"  # Pending | Paid


class VendorIn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    name: str
    kind: str = "Vendor"  # Vendor | Subcontractor
    category: str = "Other"
    contact_name: str = ""
    contact_title: str = ""
    website: str = ""
    phone: str = ""
    work_phone: str = ""
    mobile_phone: str = ""
    fax: str = ""
    email: str = ""
    tin_ein: str = ""
    tin_kind: str = "EIN"  # "EIN" | "SSN" — controls input mask + display formatting
    address: str = ""
    address_line2: str = ""
    city: str = ""
    state: str = ""
    zip_code: str = ""
    notes: str = ""
    # --- Certificates of Insurance (COI) — used primarily for Subcontractors ---
    gl_coi_on_file: bool = False
    gl_coi_issued_date: str = ""   # YYYY-MM-DD
    gl_coi_expiry_date: str = ""   # YYYY-MM-DD
    wc_coi_on_file: bool = False
    wc_coi_issued_date: str = ""
    wc_coi_expiry_date: str = ""


class Vendor(VendorIn):
    id: str
    created_at: str


class DealIn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    title: str
    deal_type: str = "Scope"  # Assessment | Scope
    contact_id: Optional[str] = None
    customer_contact_id: Optional[str] = None
    owner_contact_id: Optional[str] = None
    property_id: Optional[str] = None
    lead_source: str = "Personal"
    referral_source: str = ""
    project_type: str = "Repair"
    current_roof_type: str = "TPO"
    proposed_roof_type: str = "TPO"
    custom_scope: str = ""  # Free-form scope body — used on the PDF when proposed_roof_type is "Construction Project" or "Other"
    # Construction Project — 3-bucket scope used by the 2-page Construction PDF template
    construction_project_requirements: str = ""  # one bullet per line
    construction_other_requirements: str = ""    # materials / equipment / metal — one bullet per line
    construction_exclusions: str = ""            # one bullet per line
    construction_scope_subtitle: str = ""        # blue label printed next to "Scope of Work" on PDF (e.g. "Landscape Scope")
    project_type_override: str = ""              # overrides PROJECT TYPE label on the construction PDF
    property_sqft: float = 0.0
    perimeter_lnft: float = 0.0
    avg_parapet_height: float = 0.0
    total_sqft: float = 0.0
    # Free-form rep-entered Calculator add-ons (e.g. "Skylight curb flashing $850").
    # List of {label, cost} dicts; rendered on the spec sheet as separate
    # scope-inclusion bullets so the customer sees what they're paying for.
    calc_custom_addons: List[Dict[str, Any]] = Field(default_factory=list)
    proposal_option_1: float = 0.0
    proposal_option_2: float = 0.0
    proposal_option_3: float = 0.0
    # Optional 4th tier — used by templates with a tier_table (e.g. FARM 25-yr w/ Hail Rider)
    proposal_option_25yr: float = 0.0
    chosen_amount: float = 0.0
    chosen_date: str = ""
    date_sent: str = ""
    status: str = "Lead"
    materials_cost: float = 0.0
    labor_cost: float = 0.0
    subcontractor_cost: float = 0.0
    other_expenses: float = 0.0
    payment_milestones: List[PaymentMilestone] = Field(default_factory=list)
    cost_items: List[CostItem] = Field(default_factory=list)
    notes: str = ""
    assigned_to_user_id: Optional[str] = None
    product_description: str = ""
    warranty_20yr_add: float = 0.0
    warranty_15yr_add: float = 0.0
    warranty_10yr_add: float = 0.0
    # Optional 25-yr warranty add-on (e.g., FARM Hail Rider). Mirrors the other tiers.
    warranty_25yr_add: float = 0.0
    # NDL (No-Dollar-Limit) flags — used by the Material Calculator to auto-apply
    # Everest's tiered warranty pricing ($1,000 standard / $3,500 NDL). True means
    # the rep ticked the NDL toggle for that warranty band.
    warranty_10yr_ndl: bool = False
    warranty_15yr_ndl: bool = False
    warranty_20yr_ndl: bool = False
    warranty_25yr_ndl: bool = False
    warranty_color: str = "white"
    cover_photo_file_id: Optional[str] = None
    # Change orders — approved scope additions/deductions that affect the contract total
    change_orders: List[dict] = Field(default_factory=list)
    # Maintenance plan tracking
    maintenance_plan: bool = False
    maintenance_rate: float = 0.0
    maintenance_start_date: str = ""
    next_maintenance_date: str = ""
    last_maintenance_date: str = ""
    maintenance_visits: List[dict] = Field(default_factory=list)
    # Material Take-Off — pull from materials catalog, drives PO PDFs and Estimated Materials cost
    material_takeoff: List[dict] = Field(default_factory=list)
    # Project Calendar scheduling
    scheduled_start_date: str = ""   # YYYY-MM-DD — when crews start on site
    scheduled_end_date: str = ""     # YYYY-MM-DD — anticipated completion
    material_order_date: str = ""    # YYYY-MM-DD — when materials should be delivered / PO needs to fire
    # Scope send bookkeeping — populated when POST /deals/{id}/spec-sheet/email
    # succeeds; drives the "Scope Sent" pipeline dot on Deal Detail.
    last_scope_sent_at: str = ""
    last_scope_sent_to: str = ""
    scope_send_count: int = 0
    # Free-form append-only timeline (status changes, scope emails, etc.) used
    # by the Deal Detail "Activity Timeline" card and pipeline derivations.
    status_history: List[dict] = Field(default_factory=list)
    # Set when the customer counter-signs the proposal — promotes the
    # "Scope Sent" pipeline derivation to a stronger checkpoint.
    scope_signed_at: str = ""
    scope_signed_by_name: str = ""
    scope_signed_by_email: str = ""
    scope_signed_ip: str = ""
    scope_signed_user_agent: str = ""
    scope_signature_file_id: str = ""
    # Opaque URL-safe token gating the public /sign/{token} proposal viewer.
    # Minted on first scope-email send (see proposal_signing.ensure_proposal_token).
    proposal_sign_token: str = ""
    # Per-deal scope-bullet overrides — populated by the in-app Scope Editor.
    # Schema: { title?: str, scope_1_title?: str, scope_1?: List[str],
    #           scope_2_title?: str, scope_2?: List[str], key_advantages?: List[str] }
    # Empty / missing values fall back to the spec_sheet.py template default.
    scope_overrides: dict = Field(default_factory=dict)


class Deal(DealIn):
    id: str
    created_at: str


# ----- Invoice Models -----
class InvoiceLineItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    description: str = ""
    quantity: float = 1.0
    unit_price: float = 0.0
    amount: float = 0.0  # qty * unit_price (server-computed)


class InvoiceIn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    deal_id: Optional[str] = None
    customer_contact_id: Optional[str] = None
    invoice_type: str = ""  # Project Amount | Deposit | Mid-Project | Final | Maintenance | Repair | (blank)
    entity_id: Optional[str] = None  # Books — which legal entity this invoice belongs to (GL routes here)
    counter_entity_id: Optional[str] = None  # If billing another SealTech entity, auto-mirror via 1900/2900
    # Bill-to snapshot (frozen at creation time)
    bill_to_company: str = ""
    bill_to_name: str = ""
    bill_to_address: str = ""
    bill_to_address_line2: str = ""
    bill_to_city: str = ""
    bill_to_state: str = ""
    bill_to_zip: str = ""
    bill_to_email: str = ""
    # Email send config
    cc_email: str = ""
    # Invoice fields
    invoice_date: str = ""  # ISO yyyy-mm-dd
    due_date: str = ""
    terms: str = "Due Upon Receipt"
    project_title: str = ""
    project_address: str = ""
    project_total: float = 0.0  # Contract total of the parent project (for context on partial invoices)
    notes: str = ""
    line_items: List[InvoiceLineItem] = Field(default_factory=list)
    status: str = "Draft"  # Draft | Sent | Paid | Partial | Void | Overdue
    # Payment
    amount_paid: float = 0.0
    payment_date: str = ""
    payment_method: str = ""
    payment_reference: str = ""
    # Source link (which milestone or maintenance visit created this)
    source_type: str = ""  # milestone | maintenance_visit | manual
    source_id: str = ""


class Invoice(InvoiceIn):
    id: str
    invoice_number: str
    subtotal: float = 0.0
    total: float = 0.0
    balance_due: float = 0.0
    created_at: str
    created_by_user_id: Optional[str] = None
    last_sent_at: str = ""
    pdf_generated_at: str = ""
    gl_warnings: Optional[List[dict]] = None  # Set transiently on create/update when GL posting deferred


# ----- Vendor Bill (Payables) Models -----
class VendorBillLine(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    description: str = ""
    sku: str = ""  # vendor's product code if visible on the bill — drives auto-match to take-off lines
    project_id: Optional[str] = None
    project_title: str = ""
    quantity: float = 1.0
    unit_price: float = 0.0
    amount: float = 0.0
    # Variance tracking — when this bill line pays for a specific take-off line, store its id
    takeoff_line_id: Optional[str] = None


class VendorBillIn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    vendor_id: Optional[str] = None
    vendor_name: str = ""  # snapshot for display
    entity_id: Optional[str] = None  # Books — which legal entity this bill belongs to (GL routes here)
    counter_entity_id: Optional[str] = None  # If buying from another SealTech entity, auto-mirror via 1900/2900
    bill_number: str = ""
    bill_date: str = ""
    received_date: str = ""
    due_date: str = ""
    terms: str = "Due on Receipt"
    total: float = 0.0
    subtotal: float = 0.0
    tax: float = 0.0
    shipping: float = 0.0
    status: str = "Pending"  # Pending | Approved | Paid | Disputed | Void
    notes: str = ""
    attached_file_id: Optional[str] = None
    parsed_by_ai: bool = False
    line_items: List[VendorBillLine] = Field(default_factory=list)
    # Payment tracking
    paid_amount: float = 0.0
    paid_date: str = ""
    paid_method: str = ""
    paid_reference: str = ""


class VendorBill(VendorBillIn):
    id: str
    created_at: str
    created_by_user_id: Optional[str] = None
    gl_warnings: Optional[List[dict]] = None  # Set transiently on create/update when GL posting deferred


# ----- Materials Catalog -----
MATERIAL_CATEGORIES = ["Coating", "Primer", "Fabric", "Mastic", "Fasteners", "Sealant", "Equipment", "Tools", "Other"]


class MaterialIn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    sku: str = ""
    name: str = Field(min_length=1)
    category: str = "Other"
    unit: str = "each"
    default_price: float = 0.0
    shipping_pct: float = 0.0  # typical shipping load %, default per material
    markup_pct: float = 0.0    # internal target markup % (for your P&L planning only — never shown to customers)
    vendor_id: Optional[str] = None
    vendor_name: str = ""
    notes: str = ""


class Material(MaterialIn):
    id: str
    created_at: str
    updated_at: str


# ----- Auth Routes -----
@api_router.post("/auth/register", response_model=TokenOut)
async def register(body: RegisterReq):
    # Self-registration disabled if any users exist (admin must create users)
    count = await db.users.count_documents({})
    if count > 0:
        raise HTTPException(status_code=403, detail="Registration disabled. Ask an admin to add you.")
    email = body.email.lower()
    user_id = str(uuid.uuid4())
    doc = {
        "id": user_id,
        "email": email,
        "name": body.name,
        "role": "admin",
        "phone": "",
        "title": "",
        "password_hash": hash_password(body.password),
        "created_at": now_iso(),
    }
    await db.users.insert_one(doc)
    token = create_access_token(user_id, email)
    return TokenOut(access_token=token, user=UserOut(id=user_id, email=email, name=body.name, role="admin"))


@api_router.post("/auth/login", response_model=TokenOut)
async def login(body: LoginReq):
    email = body.email.lower()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    token = create_access_token(user["id"], email)
    return TokenOut(
        access_token=token,
        user=UserOut(
            id=user["id"], email=email, name=user.get("name", ""), role=user.get("role", "admin"),
            phone=user.get("phone", ""), title=user.get("title", ""), credentials=user.get("credentials", ""),
        ),
    )


@api_router.get("/auth/me", response_model=UserOut)
async def me(current=Depends(get_current_user)):
    return UserOut(
        id=current["id"], email=current["email"], name=current.get("name", ""),
        role=current.get("role", "admin"), phone=current.get("phone", ""), title=current.get("title", ""),
        credentials=current.get("credentials", ""),
    )


# ----- Magic-link for "Get App on My Phone" -----------------------------------
# A logged-in user generates a single-use token; scanning the QR / opening the
# link on their phone consumes the token and signs them in without re-typing
# the password. Token expires in 5 minutes and is invalidated on first use.
@api_router.post("/auth/magic-link")
async def issue_magic_link(current=Depends(get_current_user)):
    token = secrets.token_urlsafe(24)
    now = datetime.now(timezone.utc)
    await db.magic_links.insert_one({
        "token": token,
        "user_id": current["id"],
        "email": current["email"],
        "created_at": now.isoformat(),
        "expires_at": (now + timedelta(minutes=5)).isoformat(),
        "consumed_at": None,
    })
    # Index for fast lookup + TTL cleanup (best-effort; idempotent)
    try:
        await db.magic_links.create_index("token", unique=True)
        await db.magic_links.create_index("expires_at", expireAfterSeconds=600)
    except Exception:
        pass
    return {"token": token, "expires_in": 300}


@api_router.post("/auth/magic-link/consume")
async def consume_magic_link(body: dict = Body(...)):
    """Public endpoint — exchanges a one-time magic-link token for a JWT.
    No auth required (you ARE authenticating). Idempotent within the 5-minute
    window? No — single-use. Once consumed, returns 401 on retry."""
    token = (body.get("token") or "").strip()
    if not token:
        raise HTTPException(400, "token required")
    rec = await db.magic_links.find_one({"token": token})
    if not rec:
        raise HTTPException(401, "Invalid or expired link")
    if rec.get("consumed_at"):
        raise HTTPException(401, "This link has already been used")
    expires_at = rec.get("expires_at") or ""
    try:
        exp_dt = datetime.fromisoformat(expires_at.replace("Z", "+00:00"))
    except Exception:
        exp_dt = None
    if exp_dt and datetime.now(timezone.utc) > exp_dt:
        raise HTTPException(401, "This link has expired — generate a new one")

    # Mark consumed (best-effort race guard via $set with filter)
    upd = await db.magic_links.update_one(
        {"token": token, "consumed_at": None},
        {"$set": {"consumed_at": datetime.now(timezone.utc).isoformat()}},
    )
    if upd.modified_count == 0:
        raise HTTPException(401, "This link has already been used")

    user = await db.users.find_one({"id": rec["user_id"]})
    if not user:
        raise HTTPException(401, "User no longer exists")

    access_token = create_access_token(user["id"], user["email"])
    return TokenOut(
        access_token=access_token,
        user=UserOut(
            id=user["id"], email=user["email"], name=user.get("name", ""),
            role=user.get("role", "admin"), phone=user.get("phone", ""),
            title=user.get("title", ""), credentials=user.get("credentials", ""),
        ),
    )



class ProfileUpdateReq(BaseModel):
    """Self-edit: a logged-in user updates their OWN profile (name / title / phone / credentials).
    Email and role are intentionally NOT mutable here — those require admin."""
    model_config = ConfigDict(extra="ignore")
    name: Optional[str] = None
    title: Optional[str] = None
    phone: Optional[str] = None
    credentials: Optional[str] = None


class ChangePasswordReq(BaseModel):
    model_config = ConfigDict(extra="ignore")
    current_password: str
    new_password: str


@api_router.put("/auth/me", response_model=UserOut)
async def update_my_profile(body: ProfileUpdateReq, current=Depends(get_current_user)):
    patch = {k: v for k, v in body.model_dump().items() if v is not None}
    if not patch:
        raise HTTPException(status_code=400, detail="Nothing to update")
    # Strip leading/trailing whitespace; reject anything that looks like a password hash
    for k, v in list(patch.items()):
        if isinstance(v, str):
            patch[k] = v.strip()
            if patch[k].startswith("$2") and len(patch[k]) > 50:
                # Looks like a bcrypt hash — almost certainly user error
                raise HTTPException(status_code=400, detail=f"Field '{k}' looks like a password hash; ignoring.")
    await db.users.update_one({"id": current["id"]}, {"$set": patch})
    user = await db.users.find_one({"id": current["id"]}, {"_id": 0, "password_hash": 0})
    return user


@api_router.post("/auth/change-password")
async def change_my_password(body: ChangePasswordReq, current=Depends(get_current_user)):
    user = await db.users.find_one({"id": current["id"]})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    # Verify current password
    if not verify_password(body.current_password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Current password is incorrect")
    new_pw = (body.new_password or "").strip()
    if len(new_pw) < 8:
        raise HTTPException(status_code=400, detail="New password must be at least 8 characters")
    if new_pw == body.current_password:
        raise HTTPException(status_code=400, detail="New password must be different from current password")
    await db.users.update_one(
        {"id": current["id"]},
        {"$set": {"password_hash": hash_password(new_pw), "password_changed_at": now_iso()}},
    )
    return {"ok": True, "message": "Password changed successfully"}



# ----- Users (admin only) -----
def require_admin(current=Depends(get_current_user)) -> dict:
    if not is_admin(current):
        raise HTTPException(status_code=403, detail="Admin access required")
    return current


@api_router.get("/users", response_model=List[UserOut])
async def list_users(current=Depends(require_admin)):
    cursor = db.users.find({}, {"_id": 0, "password_hash": 0}).sort("name", 1)
    return await cursor.to_list(1000)


@api_router.post("/users")
async def create_user(body: UserCreateReq, current=Depends(require_admin)):
    email = body.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email already exists")
    if body.role not in USER_ROLES:
        raise HTTPException(status_code=400, detail="Invalid role")
    user_id = str(uuid.uuid4())
    generated = generate_password(12)
    doc = {
        "id": user_id,
        "email": email,
        "name": body.name,
        "role": body.role,
        "phone": body.phone or "",
        "title": body.title or "",
        "credentials": body.credentials or "",
        "password_hash": hash_password(generated),
        "created_at": now_iso(),
        "created_by": current["id"],
    }
    await db.users.insert_one(doc)
    return {
        "user": {"id": user_id, "email": email, "name": body.name, "role": body.role, "phone": doc["phone"], "title": doc["title"], "credentials": doc["credentials"]},
        "generated_password": generated,
    }


@api_router.put("/users/{user_id}", response_model=UserOut)
async def update_user(user_id: str, body: UserUpdateReq, current=Depends(require_admin)):
    patch = {k: v for k, v in body.model_dump().items() if v is not None}
    if "role" in patch and patch["role"] not in USER_ROLES:
        raise HTTPException(status_code=400, detail="Invalid role")
    if not patch:
        raise HTTPException(status_code=400, detail="Nothing to update")
    result = await db.users.update_one({"id": user_id}, {"$set": patch})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    return user


@api_router.post("/users/{user_id}/regenerate-password")
async def regenerate_password(user_id: str, current=Depends(require_admin)):
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    new_pw = generate_password(12)
    await db.users.update_one({"id": user_id}, {"$set": {"password_hash": hash_password(new_pw)}})
    return {"generated_password": new_pw}


@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current=Depends(require_admin)):
    if user_id == current["id"]:
        raise HTTPException(status_code=400, detail="You cannot delete yourself")
    admin_count = await db.users.count_documents({"role": "admin"})
    target = await db.users.find_one({"id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    if target.get("role") == "admin" and admin_count <= 1:
        raise HTTPException(status_code=400, detail="Cannot delete the last admin")
    await db.users.delete_one({"id": user_id})
    return {"ok": True}


# ----- Options Route -----
@api_router.get("/options")
async def options(current=Depends(get_current_user)):
    return {
        "lead_sources": LEAD_SOURCES,
        "project_types": PROJECT_TYPES,
        "material_categories": MATERIAL_CATEGORIES,
        "roof_types": ROOF_TYPES,
        "current_roof_types": CURRENT_ROOF_TYPES,
        "deal_statuses": DEAL_STATUSES,
        "deal_types": DEAL_TYPES,
        "vendor_kinds": VENDOR_KINDS,
        "vendor_categories": VENDOR_CATEGORIES,
        "cost_categories": COST_CATEGORIES,
        "milestone_statuses": MILESTONE_STATUSES,
        "cost_item_statuses": COST_ITEM_STATUSES,
        "contact_types": CONTACT_TYPES,
        "document_categories": DOCUMENT_CATEGORIES,
        "import_categories": IMPORT_CATEGORIES,
        "duplicate_modes": DUPLICATE_MODES,
        "user_roles": USER_ROLES,
        "milestone_templates": {
            "50/50": [{"label": "Deposit", "percent": 50}, {"label": "Completion", "percent": 50}],
            "50/25/25": [{"label": "Deposit", "percent": 50}, {"label": "Mid-Job", "percent": 25}, {"label": "Completion", "percent": 25}],
        },
    }




@api_router.get("/options/scope-preview")
async def preview_scope_title(
    proposed: str = Query("", description="Proposed roof type"),
    current: str = Query("", description="Current roof type"),
    current_user=Depends(get_current_user),
):
    """Resolve which spec-sheet template would render for a given proposed/current combo.

    Returns the template title and the Product Type line, so the deal form can show
    a "Will generate: ..." chip before the user clicks Generate.
    """
    from spec_sheet import _resolve_template
    template = _resolve_template(proposed or None, current or None)
    # Derive product_type using the same logic as the spec sheet route
    PRODUCT_TYPE_DEFAULTS = {
        "TPO Over-Lay": "TPO Roof System Over Existing TPO Over-Lay",
        "TPO Replacement": "TPO Roof System Replacing TPO",
        "EPDM Over-Lay": "EPDM Roof System Over Existing EPDM Over-Lay",
        "EPDM Replacement": "EPDM Roof System Replacing EPDM",
        "ModBit Over-Lay": "Modified Bitumen Roof System Over Existing Modified Bitumen Over-Lay",
        "ModBit Replacement": "Modified Bitumen Roof System Replacing Modified Bitumen",
        "PVC Over-Lay": "PVC Roof System Over Existing PVC Over-Lay",
        "PVC Replacement": "PVC Roof System Replacing PVC",
    }
    NEW_CONSTRUCTION_LABELS = {
        "TPO": "TPO", "TPO Over-Lay": "TPO", "TPO Replacement": "TPO",
        "EPDM": "EPDM", "EPDM Over-Lay": "EPDM", "EPDM Replacement": "EPDM", "EPDM w/ Ballast": "EPDM",
        "PVC": "PVC", "PVC Over-Lay": "PVC", "PVC Replacement": "PVC",
        "ModBit": "Modified Bitumen", "ModBit Over-Lay": "Modified Bitumen", "ModBit Replacement": "Modified Bitumen",
        "BUR (Built-Up)": "Built-Up Roof",
    }
    is_new = bool(current) and (current.strip().lower().startswith("none") or "new construction" in current.lower())
    # Custom-scope paths (non-roofing) shouldn't be labeled "Roof System"
    if template.get("dynamic_scope") or proposed in ("Construction Project", "Other"):
        if current and current.lower() != "other construction work":
            product_desc = f"Construction Project — Custom Scope (existing: {current})"
        else:
            product_desc = "Construction Project — Custom Scope"
    elif is_new and proposed in NEW_CONSTRUCTION_LABELS:
        product_desc = f"{NEW_CONSTRUCTION_LABELS[proposed]} Roof System on New Construction"
    elif proposed in PRODUCT_TYPE_DEFAULTS:
        product_desc = PRODUCT_TYPE_DEFAULTS[proposed]
    else:
        product_desc = f"{proposed} Roof System Over Existing {current}".strip() if proposed else ""
    return {
        "title": template.get("title", ""),
        "product_type": product_desc,
        "is_new_construction": is_new,
    }


# ----- Contacts -----
@api_router.get("/contacts", response_model=List[Contact])
async def list_contacts(current=Depends(get_current_user)):
    query = {"is_deleted": {"$ne": True}}
    if current.get("role") == "sales":
        query["$or"] = [{"assigned_to_user_id": current["id"]}, {"created_by_user_id": current["id"]}]
    cursor = db.contacts.find(query, {"_id": 0}).sort("created_at", -1)
    return await cursor.to_list(1000)


@api_router.post("/contacts", response_model=Contact)
async def create_contact(body: ContactIn, current=Depends(get_current_user)):
    data = body.model_dump()
    if data["billing_same_as_address"]:
        data["billing_address"] = data["address"]
        data["billing_address_line2"] = data["address_line2"]
        data["billing_city"] = data["city"]
        data["billing_state"] = data["state"]
        data["billing_zip"] = data["zip_code"]
    data["id"] = str(uuid.uuid4())
    data["created_at"] = now_iso()
    data["created_by_user_id"] = current["id"]
    # Only admins may assign a contact to another rep — non-admins always
    # own the contact they enter. Prevents reps from re-routing each other's
    # customers and keeps the office-admin handoff workflow centralized.
    requested_assignee = (data.get("assigned_to_user_id") or "").strip() or None
    if requested_assignee and requested_assignee != current["id"] and not is_admin(current):
        raise HTTPException(403, "Only an admin can assign a contact to a different rep")
    if not requested_assignee:
        data["assigned_to_user_id"] = current["id"]
    else:
        data["assigned_to_user_id"] = requested_assignee
    await db.contacts.insert_one(data.copy())
    return strip_id(data)


@api_router.get("/contacts/{contact_id}", response_model=Contact)
async def get_contact(contact_id: str, current=Depends(get_current_user)):
    doc = await db.contacts.find_one({"id": contact_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Contact not found")
    return doc


@api_router.put("/contacts/{contact_id}", response_model=Contact)
async def update_contact(contact_id: str, body: ContactIn, current=Depends(get_current_user)):
    existing = await db.contacts.find_one({"id": contact_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Contact not found")
    data = body.model_dump()
    if data["billing_same_as_address"]:
        data["billing_address"] = data["address"]
        data["billing_address_line2"] = data["address_line2"]
        data["billing_city"] = data["city"]
        data["billing_state"] = data["state"]
        data["billing_zip"] = data["zip_code"]
    # Lock the assignee field on update for non-admins. We accept their
    # submission only when it matches the existing value (i.e. they didn't
    # change it). Any divergence requires admin role.
    new_assignee = (data.get("assigned_to_user_id") or "").strip() or None
    prev_assignee = existing.get("assigned_to_user_id")
    if new_assignee != prev_assignee and not is_admin(current):
        # Silently restore to the existing value rather than 403 — keeps the
        # form submission resilient when non-admins POST the full record.
        data["assigned_to_user_id"] = prev_assignee
    result = await db.contacts.update_one({"id": contact_id}, {"$set": data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Contact not found")
    doc = await db.contacts.find_one({"id": contact_id}, {"_id": 0})
    return doc


@api_router.delete("/contacts/{contact_id}")
async def delete_contact(contact_id: str, current=Depends(get_current_user)):
    if is_admin(current):
        result = await db.contacts.delete_one({"id": contact_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Contact not found")
    else:
        result = await db.contacts.update_one(
            {"id": contact_id},
            {"$set": {"is_deleted": True, "deleted_at": now_iso(), "deleted_by": current["id"]}},
        )
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Contact not found")
    return {"ok": True}


# ----- Properties -----
@api_router.get("/properties", response_model=List[Property])
async def list_properties(current=Depends(get_current_user)):
    cursor = db.properties.find({"is_deleted": {"$ne": True}}, {"_id": 0}).sort("created_at", -1)
    return await cursor.to_list(1000)


@api_router.post("/properties", response_model=Property)
async def create_property(body: PropertyIn, current=Depends(get_current_user)):
    data = body.model_dump()
    data["id"] = str(uuid.uuid4())
    data["created_at"] = now_iso()
    await db.properties.insert_one(data.copy())
    return strip_id(data)


@api_router.get("/properties/{property_id}", response_model=Property)
async def get_property(property_id: str, current=Depends(get_current_user)):
    doc = await db.properties.find_one({"id": property_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Property not found")
    return doc


@api_router.put("/properties/{property_id}", response_model=Property)
async def update_property(property_id: str, body: PropertyIn, current=Depends(get_current_user)):
    data = body.model_dump()
    result = await db.properties.update_one({"id": property_id}, {"$set": data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Property not found")
    doc = await db.properties.find_one({"id": property_id}, {"_id": 0})
    return doc


@api_router.delete("/properties/{property_id}")
async def delete_property(property_id: str, current=Depends(get_current_user)):
    if is_admin(current):
        result = await db.properties.delete_one({"id": property_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Property not found")
    else:
        result = await db.properties.update_one(
            {"id": property_id},
            {"$set": {"is_deleted": True, "deleted_at": now_iso(), "deleted_by": current["id"]}},
        )
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Property not found")
    return {"ok": True}


def normalize_deal(data: dict) -> dict:
    """Auto-fill cost line item ids, milestone ids/amounts, and roll up aggregate cost buckets."""
    chosen = float(data.get("chosen_amount") or 0)

    # Measurements: auto-compute total_sqft = property_sqft + (perimeter_lnft * avg_parapet_height)
    try:
        psqft = float(data.get("property_sqft") or 0)
        plnft = float(data.get("perimeter_lnft") or 0)
        pht = float(data.get("avg_parapet_height") or 0)
    except (TypeError, ValueError):
        psqft = plnft = pht = 0.0
    data["property_sqft"] = round(psqft, 2)
    data["perimeter_lnft"] = round(plnft, 2)
    data["avg_parapet_height"] = round(pht, 2)
    data["total_sqft"] = round(psqft + (plnft * pht), 2)

    # Milestones: ensure id, recompute amount from percent * chosen
    milestones = data.get("payment_milestones") or []
    for m in milestones:
        if not m.get("id"):
            m["id"] = str(uuid.uuid4())
        try:
            pct = float(m.get("percent") or 0)
        except (TypeError, ValueError):
            pct = 0.0
        m["percent"] = pct
        m["amount"] = round(chosen * pct / 100.0, 2)
    data["payment_milestones"] = milestones

    # Cost items: ensure id; aggregate buckets
    items = data.get("cost_items") or []
    buckets = {"Materials": 0.0, "Labor": 0.0, "Subcontractor": 0.0, "Other": 0.0}
    for it in items:
        if not it.get("id"):
            it["id"] = str(uuid.uuid4())
        try:
            amt = float(it.get("amount") or 0)
        except (TypeError, ValueError):
            amt = 0.0
        it["amount"] = amt
        cat = it.get("category") or "Other"
        if cat not in buckets:
            cat = "Other"
            it["category"] = cat
        buckets[cat] += amt
    data["cost_items"] = items
    data["materials_cost"] = round(buckets["Materials"], 2)
    data["labor_cost"] = round(buckets["Labor"], 2)
    data["subcontractor_cost"] = round(buckets["Subcontractor"], 2)
    data["other_expenses"] = round(buckets["Other"], 2)

    # Maintenance: derive next_maintenance_date / last_maintenance_date from visits + start date
    visits = data.get("maintenance_visits") or []
    cleaned_visits = []
    for v in visits:
        if not v.get("id"):
            v["id"] = str(uuid.uuid4())
        try:
            v["amount"] = float(v.get("amount") or 0)
        except (TypeError, ValueError):
            v["amount"] = 0.0
        cleaned_visits.append(v)
    # Sort visits by date descending
    cleaned_visits.sort(key=lambda x: x.get("visit_date", ""), reverse=True)
    data["maintenance_visits"] = cleaned_visits
    last_visit_date = cleaned_visits[0].get("visit_date", "") if cleaned_visits else ""
    data["last_maintenance_date"] = last_visit_date

    # Normalize change orders
    change_orders = data.get("change_orders") or []
    cleaned_cos = []
    for co in change_orders:
        if not co.get("id"):
            co["id"] = str(uuid.uuid4())
        try:
            co["amount"] = float(co.get("amount") or 0)
        except (TypeError, ValueError):
            co["amount"] = 0.0
        if not co.get("status"):
            co["status"] = "Approved"
        cleaned_cos.append(co)
    data["change_orders"] = cleaned_cos

    # Normalize material take-off lines (snapshots from catalog)
    takeoff = data.get("material_takeoff") or []
    cleaned_takeoff = []
    for ln in takeoff:
        if not ln.get("id"):
            ln["id"] = str(uuid.uuid4())
        try:
            ln["quantity"] = float(ln.get("quantity") or 0)
        except (TypeError, ValueError):
            ln["quantity"] = 0.0
        try:
            ln["unit_cost"] = float(ln.get("unit_cost") or 0)
        except (TypeError, ValueError):
            ln["unit_cost"] = 0.0
        ln["line_total"] = round(ln["quantity"] * ln["unit_cost"], 2)
        ln.setdefault("ordered", False)
        ln.setdefault("received", False)
        cleaned_takeoff.append(ln)
    data["material_takeoff"] = cleaned_takeoff

    # Compute next_maintenance_date: last_visit + 1 year, else start_date + 1 year
    start = data.get("maintenance_start_date", "") or ""
    base = last_visit_date or start
    next_due = ""
    if base:
        try:
            base_dt = datetime.fromisoformat(base[:10])
            next_dt = base_dt.replace(year=base_dt.year + 1)
            next_due = next_dt.date().isoformat()
        except (ValueError, TypeError):
            next_due = ""
    data["next_maintenance_date"] = next_due
    return data


# ----- Deals -----
@api_router.get("/deals", response_model=List[Deal])
async def list_deals(current=Depends(get_current_user)):
    query = {"is_deleted": {"$ne": True}}
    if current.get("role") == "sales":
        query["$or"] = [{"assigned_to_user_id": current["id"]}, {"created_by_user_id": current["id"]}]
    cursor = db.deals.find(query, {"_id": 0}).sort("created_at", -1)
    items = await cursor.to_list(1000)
    return [scrub_deal(d, current) for d in items]


@api_router.post("/deals", response_model=Deal)
async def create_deal(body: DealIn, current=Depends(get_current_user)):
    data = normalize_deal(body.model_dump())
    data["id"] = str(uuid.uuid4())
    data["created_at"] = now_iso()
    data["created_by_user_id"] = current["id"]
    # Same admin-gate as contacts: only an admin can route a new deal to
    # another rep. Non-admins always own what they enter.
    requested_assignee = (data.get("assigned_to_user_id") or "").strip() or None
    if requested_assignee and requested_assignee != current["id"] and not is_admin(current):
        raise HTTPException(403, "Only an admin can assign a deal to a different rep")
    if not requested_assignee:
        data["assigned_to_user_id"] = current["id"]
    else:
        data["assigned_to_user_id"] = requested_assignee
    await db.deals.insert_one(data.copy())
    return strip_id(data)


@api_router.get("/deals/{deal_id}", response_model=Deal)
async def get_deal(deal_id: str, current=Depends(get_current_user)):
    doc = await db.deals.find_one({"id": deal_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Deal not found")
    if current.get("role") == "sales":
        owns = doc.get("assigned_to_user_id") == current["id"] or doc.get("created_by_user_id") == current["id"]
        if not owns:
            raise HTTPException(status_code=403, detail="Not your project")
    return scrub_deal(doc, current)


@api_router.get("/deals/{deal_id}/activity")
async def deal_activity(deal_id: str, current=Depends(get_current_user)):
    """Reconstructed activity timeline for a deal — derived from existing
    collections so we don't need a dedicated audit trail. Returns most-recent first."""
    deal = await db.deals.find_one({"id": deal_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")
    if current.get("role") == "sales":
        owns = deal.get("assigned_to_user_id") == current["id"] or deal.get("created_by_user_id") == current["id"]
        if not owns:
            raise HTTPException(status_code=403, detail="Not your project")

    EM_DASH = "\u2014"
    items = []
    scope_send_running_count = 0
    if deal.get("created_at"):
        items.append({"ts": deal["created_at"], "kind": "deal_created", "title": "Project created", "color": "#71717A"})
    for entry in (deal.get("status_history") or []):
        label = entry.get("label") or ""
        # Scope / Assessment email-send entries: rendered with recipient + a
        # running send-count so reps can see who got the proposal and how many
        # times it has gone out.
        if label in ("Scope emailed", "Assessment emailed"):
            scope_send_running_count += 1
            recipient = entry.get("to") or EM_DASH
            attach = int(entry.get("attachments_count") or 0)
            actor = entry.get("user_name") or ""
            attach_bit = f" {EM_DASH} {attach} attachment{'s' if attach != 1 else ''}" if attach else ""
            actor_bit = f" by {actor}" if actor else ""
            pdf_file_id = entry.get("pdf_file_id") or ""
            items.append({
                "ts": entry.get("at") or "",
                "kind": "invoice_sent" if label.startswith("Scope") else "assessment_created",
                "title": f"{label} (send #{scope_send_running_count})",
                "subtitle": f"to {recipient}{attach_bit}{actor_bit}",
                "color": "#062B67",
                "pdf_file_id": pdf_file_id,
            })
            continue
        # Default status-change entry (existing pattern: {from, to, at})
        items.append({
            "ts": entry.get("at") or "",
            "kind": "status_change",
            "title": f"Status \u2192 {entry.get('to') or '?'}",
            "subtitle": (f"From {entry.get('from')}" if entry.get("from") else ""),
            "color": "#062B67",
        })

    async for inv in db.invoices.find({"deal_id": deal_id, "is_deleted": {"$ne": True}}, {"_id": 0}):
        if inv.get("created_at"):
            items.append({
                "ts": inv["created_at"],
                "kind": "invoice_created",
                "title": f"Invoice {inv.get('invoice_number','')} drafted",
                "subtitle": f"${float(inv.get('total') or 0):,.2f}",
                "color": "#7E22CE",
            })
        if inv.get("last_sent_at"):
            email_to = inv.get("bill_to_email") or EM_DASH
            items.append({
                "ts": inv["last_sent_at"],
                "kind": "invoice_sent",
                "title": f"Invoice {inv.get('invoice_number','')} emailed",
                "subtitle": f"to {email_to}",
                "color": "#7E22CE",
            })
        if float(inv.get("amount_paid") or 0) > 0 and inv.get("paid_at"):
            paid_amt = float(inv['amount_paid'])
            items.append({
                "ts": inv["paid_at"],
                "kind": "payment_received",
                "title": f"Payment received {EM_DASH} ${paid_amt:,.2f}",
                "subtitle": f"Invoice {inv.get('invoice_number','')}",
                "color": "#16A34A",
            })

    for v in (deal.get("maintenance_visits") or []):
        if v.get("visit_date"):
            items.append({
                "ts": v["visit_date"],
                "kind": "maintenance_visit",
                "title": "Maintenance visit",
                "subtitle": (v.get("notes") or "")[:140],
                "color": "#16A34A",
            })

    async for ph in db.project_photos.find({"deal_id": deal_id}, {"_id": 0, "uploaded_at": 1, "label": 1}):
        if ph.get("uploaded_at"):
            items.append({
                "ts": ph["uploaded_at"],
                "kind": "photo_uploaded",
                "title": "Photo uploaded",
                "subtitle": ph.get("label") or "",
                "color": "#0EA5E9",
            })

    async for a in db.assessments.find({"deal_id": deal_id, "is_deleted": {"$ne": True}}, {"_id": 0}):
        if a.get("created_at"):
            items.append({
                "ts": a["created_at"],
                "kind": "assessment_created",
                "title": "Assessment started",
                "subtitle": a.get("property_address") or a.get("property_name") or "",
                "color": "#D97706",
            })

    items = [i for i in items if i.get("ts")]
    items.sort(key=lambda x: str(x["ts"]), reverse=True)
    return {"items": items[:200]}




@api_router.put("/deals/{deal_id}", response_model=Deal)
async def update_deal(deal_id: str, body: DealIn, current=Depends(get_current_user)):
    existing = await db.deals.find_one({"id": deal_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Deal not found")
    if current.get("role") == "sales":
        owns = existing.get("assigned_to_user_id") == current["id"] or existing.get("created_by_user_id") == current["id"]
        if not owns:
            raise HTTPException(status_code=403, detail="Not your project")
    data = normalize_deal(body.model_dump())
    # Lock the assignee field on update for non-admins. Silently restore to
    # the existing value rather than 403 — keeps the form submission
    # resilient when non-admins POST the full record without re-rendering
    # the dropdown.
    new_assignee = (data.get("assigned_to_user_id") or "").strip() or None
    prev_assignee = existing.get("assigned_to_user_id")
    if new_assignee != prev_assignee and not is_admin(current):
        data["assigned_to_user_id"] = prev_assignee
    # Track status changes for the activity timeline
    prev_status = (existing or {}).get("status")
    new_status = data.get("status")
    if new_status and prev_status and new_status != prev_status:
        history = list((existing or {}).get("status_history") or [])
        history.append({
            "from": prev_status,
            "to": new_status,
            "at": datetime.now(timezone.utc).isoformat(),
            "user_id": current["id"],
        })
        data["status_history"] = history
    # Track reassignments separately so admins can audit who moved deals.
    if (data.get("assigned_to_user_id") or None) != (prev_assignee or None):
        reassign_log = list((existing or {}).get("reassignment_history") or [])
        reassign_log.append({
            "from": prev_assignee,
            "to": data.get("assigned_to_user_id"),
            "at": datetime.now(timezone.utc).isoformat(),
            "by_user_id": current["id"],
        })
        data["reassignment_history"] = reassign_log
    await db.deals.update_one({"id": deal_id}, {"$set": data})
    doc = await db.deals.find_one({"id": deal_id}, {"_id": 0})
    # Fire-and-forget push to Google Calendar
    try:
        await gcal_module.push_deal(db, current["id"], doc, _PUBLIC_BASE_URL)
    except Exception:
        pass
    return scrub_deal(doc, current)


class DealScheduleIn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    scheduled_start_date: Optional[str] = None
    scheduled_end_date: Optional[str] = None
    material_order_date: Optional[str] = None


@api_router.put("/deals/{deal_id}/schedule", response_model=Deal)
async def update_deal_schedule(deal_id: str, body: DealScheduleIn, current=Depends(get_current_user)):
    """Partial update for Project Calendar reschedules — only the three date fields.
    Used by drag-to-reschedule on the calendar grid."""
    existing = await db.deals.find_one({"id": deal_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Deal not found")
    if current.get("role") == "sales":
        owns = existing.get("assigned_to_user_id") == current["id"] or existing.get("created_by_user_id") == current["id"]
        if not owns:
            raise HTTPException(status_code=403, detail="Not your project")
    patch = {k: v for k, v in body.model_dump().items() if v is not None}
    if patch:
        await db.deals.update_one({"id": deal_id}, {"$set": patch})
    doc = await db.deals.find_one({"id": deal_id}, {"_id": 0})
    try:
        await gcal_module.push_deal(db, current["id"], doc, _PUBLIC_BASE_URL)
    except Exception:
        pass
    return scrub_deal(doc, current)


class MaintenanceRescheduleIn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    visit_id: Optional[str] = None   # if provided → reschedule that logged visit
    date: str                         # new YYYY-MM-DD


@api_router.put("/deals/{deal_id}/maintenance-reschedule", response_model=Deal)
async def reschedule_maintenance(deal_id: str, body: MaintenanceRescheduleIn, current=Depends(get_current_user)):
    """Drag-to-reschedule for maintenance events on the Project Calendar.
       - If visit_id is provided, updates that specific entry in maintenance_visits[].
       - Otherwise (tentative slot), updates the deal's next_maintenance_date."""
    existing = await db.deals.find_one({"id": deal_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Deal not found")
    if current.get("role") == "sales":
        owns = existing.get("assigned_to_user_id") == current["id"] or existing.get("created_by_user_id") == current["id"]
        if not owns:
            raise HTTPException(status_code=403, detail="Not your project")
    moved_visit = None
    if body.visit_id:
        visits = list(existing.get("maintenance_visits") or [])
        found = False
        for v in visits:
            if v.get("id") == body.visit_id:
                v["visit_date"] = body.date
                moved_visit = v
                found = True
                break
        if not found:
            raise HTTPException(status_code=404, detail="Maintenance visit not found")
        await db.deals.update_one({"id": deal_id}, {"$set": {"maintenance_visits": visits}})
    else:
        await db.deals.update_one({"id": deal_id}, {"$set": {"next_maintenance_date": body.date}})
    doc = await db.deals.find_one({"id": deal_id}, {"_id": 0})
    # Push the reschedule to Google Calendar
    if moved_visit:
        try:
            await gcal_module.push_maintenance_visit(db, current["id"], deal_id, moved_visit, doc.get("title") or "Project", _PUBLIC_BASE_URL)
        except Exception:
            pass
    return scrub_deal(doc, current)



@api_router.delete("/deals/{deal_id}")
async def delete_deal(deal_id: str, current=Depends(get_current_user)):
    if is_admin(current):
        result = await db.deals.delete_one({"id": deal_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Deal not found")
    else:
        result = await db.deals.update_one(
            {"id": deal_id},
            {"$set": {"is_deleted": True, "deleted_at": now_iso(), "deleted_by": current["id"]}},
        )
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Deal not found")
    return {"ok": True}


# ----- Maintenance Plan -----
class MaintenanceVisitIn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    visit_date: str
    amount: float = 0.0
    subcontractor_id: Optional[str] = None
    subcontractor_name: str = ""
    notes: str = ""


@api_router.post("/deals/{deal_id}/maintenance-visits", response_model=Deal)
async def add_maintenance_visit(deal_id: str, body: MaintenanceVisitIn, current=Depends(get_current_user)):
    existing = await db.deals.find_one({"id": deal_id, "is_deleted": {"$ne": True}})
    if not existing:
        raise HTTPException(status_code=404, detail="Project not found")
    if current.get("role") == "sales":
        owns = existing.get("assigned_to_user_id") == current["id"] or existing.get("created_by_user_id") == current["id"]
        if not owns:
            raise HTTPException(status_code=403, detail="Not your project")
    visit = body.model_dump()
    visit["id"] = str(uuid.uuid4())
    # Auto-fill subcontractor_name from id if not provided
    if visit.get("subcontractor_id") and not visit.get("subcontractor_name"):
        sub = await db.vendors.find_one({"id": visit["subcontractor_id"]}, {"_id": 0})
        if sub:
            visit["subcontractor_name"] = sub.get("name", "")
    visits = list(existing.get("maintenance_visits") or [])
    visits.append(visit)
    # Re-run normalize via update_deal to recompute next_maintenance_date
    merged = {**existing, "maintenance_visits": visits, "maintenance_plan": True}
    # Strip server-managed before re-normalizing
    for k in ("id", "created_at", "_id", "is_deleted", "deleted_at", "deleted_by"):
        merged.pop(k, None)
    cleaned = normalize_deal(merged)
    await db.deals.update_one({"id": deal_id}, {"$set": cleaned})
    doc = await db.deals.find_one({"id": deal_id}, {"_id": 0})
    # Push the newly added maintenance visit to Google Calendar
    try:
        await gcal_module.push_maintenance_visit(db, current["id"], deal_id, visit, doc.get("title") or "Project", _PUBLIC_BASE_URL)
    except Exception:
        pass
    return scrub_deal(doc, current)


@api_router.delete("/deals/{deal_id}/maintenance-visits/{visit_id}", response_model=Deal)
async def delete_maintenance_visit(deal_id: str, visit_id: str, current=Depends(get_current_user)):
    existing = await db.deals.find_one({"id": deal_id, "is_deleted": {"$ne": True}})
    if not existing:
        raise HTTPException(status_code=404, detail="Project not found")
    if current.get("role") == "sales":
        owns = existing.get("assigned_to_user_id") == current["id"] or existing.get("created_by_user_id") == current["id"]
        if not owns:
            raise HTTPException(status_code=403, detail="Not your project")
    visits = [v for v in (existing.get("maintenance_visits") or []) if v.get("id") != visit_id]
    merged = {**existing, "maintenance_visits": visits}
    for k in ("id", "created_at", "_id", "is_deleted", "deleted_at", "deleted_by"):
        merged.pop(k, None)
    cleaned = normalize_deal(merged)
    await db.deals.update_one({"id": deal_id}, {"$set": cleaned})
    doc = await db.deals.find_one({"id": deal_id}, {"_id": 0})
    return scrub_deal(doc, current)


# ----- Material Take-Off -----
class TakeoffLineIn(BaseModel):
    """Single line being added to a project's take-off. Server snapshots the catalog fields."""
    model_config = ConfigDict(extra="ignore")
    material_id: str
    quantity: float = 0.0
    notes: str = ""


class TakeoffBulkAddIn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    lines: List[TakeoffLineIn] = Field(default_factory=list)


class TakeoffLinePatch(BaseModel):
    model_config = ConfigDict(extra="ignore")
    quantity: Optional[float] = None
    notes: Optional[str] = None
    ordered: Optional[bool] = None
    received: Optional[bool] = None


async def _check_deal_owner(deal_id: str, current: dict) -> dict:
    existing = await db.deals.find_one({"id": deal_id, "is_deleted": {"$ne": True}})
    if not existing:
        raise HTTPException(status_code=404, detail="Project not found")
    if current.get("role") == "sales":
        owns = existing.get("assigned_to_user_id") == current["id"] or existing.get("created_by_user_id") == current["id"]
        if not owns:
            raise HTTPException(status_code=403, detail="Not your project")
    return existing


def _persist_deal_takeoff(existing: dict, new_takeoff: list) -> dict:
    merged = {**existing, "material_takeoff": new_takeoff}
    for k in ("id", "created_at", "_id", "is_deleted", "deleted_at", "deleted_by"):
        merged.pop(k, None)
    return normalize_deal(merged)


@api_router.post("/deals/{deal_id}/takeoff", response_model=Deal)
async def add_takeoff_lines(deal_id: str, body: TakeoffBulkAddIn, current=Depends(get_current_user)):
    """Bulk-add take-off lines. Snapshots catalog fields (sku/name/unit/vendor/loaded cost)
    so price/name changes in the catalog do NOT silently affect previously-built take-offs."""
    existing = await _check_deal_owner(deal_id, current)
    if not body.lines:
        raise HTTPException(status_code=400, detail="No lines provided")

    # Resolve material snapshots
    new_lines = list(existing.get("material_takeoff") or [])
    for ln in body.lines:
        if ln.quantity <= 0:
            continue
        mat = await db.materials.find_one({"id": ln.material_id, "is_deleted": {"$ne": True}}, {"_id": 0})
        if not mat:
            raise HTTPException(status_code=404, detail=f"Material {ln.material_id} not found")
        loaded = float(mat.get("default_price", 0) or 0) * (1 + float(mat.get("shipping_pct", 0) or 0) / 100)
        new_lines.append({
            "id": str(uuid.uuid4()),
            "material_id": mat["id"],
            "sku": mat.get("sku", "") or "",
            "name": mat.get("name", "") or "",
            "unit": mat.get("unit", "") or "",
            "category": mat.get("category", "") or "",
            "vendor_id": mat.get("vendor_id"),
            "vendor_name": mat.get("vendor_name", "") or "",
            "quantity": float(ln.quantity),
            "unit_cost": round(loaded, 2),
            "line_total": round(loaded * float(ln.quantity), 2),
            "notes": ln.notes or "",
            "ordered": False,
            "received": False,
            "added_at": now_iso(),
        })

    cleaned = _persist_deal_takeoff(existing, new_lines)
    await db.deals.update_one({"id": deal_id}, {"$set": cleaned})
    doc = await db.deals.find_one({"id": deal_id}, {"_id": 0})
    return scrub_deal(doc, current)


@api_router.put("/deals/{deal_id}/takeoff/{line_id}", response_model=Deal)
async def update_takeoff_line(deal_id: str, line_id: str, body: TakeoffLinePatch, current=Depends(get_current_user)):
    existing = await _check_deal_owner(deal_id, current)
    lines = list(existing.get("material_takeoff") or [])
    found = False
    patch = {k: v for k, v in body.model_dump().items() if v is not None}
    for ln in lines:
        if ln.get("id") == line_id:
            ln.update(patch)
            found = True
            break
    if not found:
        raise HTTPException(status_code=404, detail="Take-off line not found")
    cleaned = _persist_deal_takeoff(existing, lines)
    await db.deals.update_one({"id": deal_id}, {"$set": cleaned})
    doc = await db.deals.find_one({"id": deal_id}, {"_id": 0})
    return scrub_deal(doc, current)


@api_router.delete("/deals/{deal_id}/takeoff/{line_id}", response_model=Deal)
async def delete_takeoff_line(deal_id: str, line_id: str, current=Depends(get_current_user)):
    existing = await _check_deal_owner(deal_id, current)
    lines = [ln for ln in (existing.get("material_takeoff") or []) if ln.get("id") != line_id]
    cleaned = _persist_deal_takeoff(existing, lines)
    await db.deals.update_one({"id": deal_id}, {"$set": cleaned})
    doc = await db.deals.find_one({"id": deal_id}, {"_id": 0})
    return scrub_deal(doc, current)


# ----- Take-Off ↔ Bill Linking + Variance -----
def _normalize_str(s: str) -> str:
    """Lowercase, alphanumeric-only normalization for fuzzy SKU matching."""
    return "".join(ch for ch in (s or "").lower() if ch.isalnum())


@api_router.get("/deals/{deal_id}/takeoff-variance")
async def deal_takeoff_variance(deal_id: str, current=Depends(get_current_user)):
    """Estimated vs Actual breakdown per take-off line.

    Actual = sum of vendor_bill line items whose `takeoff_line_id` equals the take-off line's id.
    Returns per-line, per-vendor, and project totals.
    """
    deal = await _check_deal_owner(deal_id, current)
    takeoff = list(deal.get("material_takeoff") or [])
    takeoff_ids = [ln["id"] for ln in takeoff]

    # Aggregate all bill line items linked to these take-off lines
    linked_amounts: dict = {tid: 0.0 for tid in takeoff_ids}
    linked_bills_by_line: dict = {tid: [] for tid in takeoff_ids}
    if takeoff_ids:
        cursor = db.vendor_bills.find(
            {"is_deleted": {"$ne": True}, "line_items.takeoff_line_id": {"$in": takeoff_ids}},
            {"_id": 0},
        )
        async for bill in cursor:
            for li in bill.get("line_items", []) or []:
                tid = li.get("takeoff_line_id")
                if tid in linked_amounts:
                    linked_amounts[tid] += float(li.get("amount") or 0)
                    linked_bills_by_line[tid].append({
                        "bill_id": bill.get("id"),
                        "bill_number": bill.get("bill_number") or "",
                        "vendor_name": bill.get("vendor_name") or "",
                        "bill_date": bill.get("bill_date") or "",
                        "line_id": li.get("id"),
                        "line_description": li.get("description") or "",
                        "line_amount": float(li.get("amount") or 0),
                        "line_quantity": float(li.get("quantity") or 0),
                    })

    out_lines = []
    by_vendor: dict = {}
    for ln in takeoff:
        est = float(ln.get("line_total") or 0)
        act = round(linked_amounts.get(ln["id"], 0.0), 2)
        var = round(act - est, 2)
        var_pct = round((var / est) * 100, 1) if est > 0 else None
        out_lines.append({
            "id": ln["id"],
            "name": ln.get("name", ""),
            "sku": ln.get("sku", ""),
            "vendor_name": ln.get("vendor_name", ""),
            "vendor_id": ln.get("vendor_id"),
            "unit": ln.get("unit", ""),
            "quantity": float(ln.get("quantity") or 0),
            "ordered": bool(ln.get("ordered")),
            "received": bool(ln.get("received")),
            "estimated": round(est, 2),
            "actual": act,
            "variance": var,
            "variance_pct": var_pct,
            "linked_bills": linked_bills_by_line.get(ln["id"], []),
        })

        v = ln.get("vendor_name") or "Unassigned"
        bv = by_vendor.setdefault(v, {"vendor_name": v, "estimated": 0.0, "actual": 0.0, "lines": 0, "linked_lines": 0})
        bv["estimated"] += est
        bv["actual"] += act
        bv["lines"] += 1
        if act > 0:
            bv["linked_lines"] += 1

    total_est = round(sum(l["estimated"] for l in out_lines), 2)
    total_act = round(sum(l["actual"] for l in out_lines), 2)
    total_var = round(total_act - total_est, 2)
    total_var_pct = round((total_var / total_est) * 100, 1) if total_est > 0 else None

    by_vendor_arr = []
    for v in by_vendor.values():
        v["estimated"] = round(v["estimated"], 2)
        v["actual"] = round(v["actual"], 2)
        v["variance"] = round(v["actual"] - v["estimated"], 2)
        v["variance_pct"] = round((v["variance"] / v["estimated"]) * 100, 1) if v["estimated"] > 0 else None
        by_vendor_arr.append(v)

    return {
        "lines": out_lines,
        "by_vendor": sorted(by_vendor_arr, key=lambda x: x["vendor_name"]),
        "totals": {
            "estimated": total_est,
            "actual": total_act,
            "variance": total_var,
            "variance_pct": total_var_pct,
            "lines": len(out_lines),
            "linked_lines": sum(1 for l in out_lines if l["actual"] > 0),
        },
    }


class LinkBillLineReq(BaseModel):
    model_config = ConfigDict(extra="ignore")
    takeoff_line_id: Optional[str] = None  # null = clear the link


@api_router.put("/vendor-bills/{bill_id}/lines/{line_id}/link")
async def link_bill_line_to_takeoff(bill_id: str, line_id: str, body: LinkBillLineReq, current=Depends(get_current_user)):
    """Link (or unlink with takeoff_line_id=null) a single vendor bill line item
    to a specific take-off line on a project."""
    bill = await db.vendor_bills.find_one({"id": bill_id, "is_deleted": {"$ne": True}})
    if not bill:
        raise HTTPException(status_code=404, detail="Vendor bill not found")

    # Validate target take-off line exists (if linking) and infer project_id from it
    target_deal = None
    target_line = None
    if body.takeoff_line_id:
        target_deal = await db.deals.find_one({"material_takeoff.id": body.takeoff_line_id, "is_deleted": {"$ne": True}})
        if not target_deal:
            raise HTTPException(status_code=404, detail="Take-off line not found")
        for ln in target_deal.get("material_takeoff") or []:
            if ln.get("id") == body.takeoff_line_id:
                target_line = ln
                break

    items = list(bill.get("line_items") or [])
    found = False
    for li in items:
        if li.get("id") == line_id:
            li["takeoff_line_id"] = body.takeoff_line_id
            # Snapshot the project on the bill line for filtering
            if target_deal:
                li["project_id"] = target_deal["id"]
                li["project_title"] = target_deal.get("title") or li.get("project_title", "")
            found = True
            break
    if not found:
        raise HTTPException(status_code=404, detail="Bill line item not found")

    await db.vendor_bills.update_one(
        {"id": bill_id},
        {"$set": {"line_items": items, "updated_at": now_iso()}},
    )
    return {
        "ok": True,
        "linked_to": body.takeoff_line_id,
        "takeoff_line_name": (target_line or {}).get("name", "") if target_line else "",
    }


@api_router.get("/deals/{deal_id}/linkable-bill-lines")
async def list_linkable_bill_lines(deal_id: str, current=Depends(get_current_user)):
    """Return vendor bill line items that could be linked to take-off lines on this project.

    Includes bills whose vendor matches a vendor in the project's take-off OR bills already
    tagged with this project_id on at least one line. For each line, returns an auto-match
    suggestion (the take-off line whose SKU normalizes to the same value).
    """
    deal = await _check_deal_owner(deal_id, current)
    takeoff = list(deal.get("material_takeoff") or [])
    if not takeoff:
        return {"lines": []}

    project_vendor_ids = list({ln.get("vendor_id") for ln in takeoff if ln.get("vendor_id")})
    takeoff_line_ids = {ln["id"] for ln in takeoff}
    # SKU map for fuzzy auto-suggest
    sku_to_takeoff = {}
    for ln in takeoff:
        key = _normalize_str(ln.get("sku", ""))
        if key:
            sku_to_takeoff.setdefault(key, []).append(ln)

    or_filters = []
    if project_vendor_ids:
        or_filters.append({"vendor_id": {"$in": project_vendor_ids}})
    or_filters.append({"line_items.project_id": deal_id})
    cursor = db.vendor_bills.find(
        {"is_deleted": {"$ne": True}, "$or": or_filters},
        {"_id": 0},
    ).sort("bill_date", -1)
    bills = await cursor.to_list(500)

    out = []
    for bill in bills:
        for li in bill.get("line_items") or []:
            current_link = li.get("takeoff_line_id")
            # Skip lines already linked to a take-off on a DIFFERENT project
            if current_link and current_link not in takeoff_line_ids:
                continue
            sugg = None
            if not current_link:
                key = _normalize_str(li.get("sku", ""))
                if key and key in sku_to_takeoff:
                    sugg = sku_to_takeoff[key][0]["id"]
            out.append({
                "bill_id": bill.get("id"),
                "bill_number": bill.get("bill_number") or "",
                "bill_date": bill.get("bill_date") or "",
                "vendor_id": bill.get("vendor_id"),
                "vendor_name": bill.get("vendor_name") or "",
                "line_id": li.get("id"),
                "description": li.get("description") or "",
                "sku": li.get("sku") or "",
                "quantity": float(li.get("quantity") or 0),
                "amount": float(li.get("amount") or 0),
                "linked_to": current_link,
                "suggested_takeoff_line_id": sugg,
            })
    return {"lines": out}



@api_router.get("/materials/grouped")
async def list_materials_grouped(current=Depends(get_current_user)):
    """Return materials grouped by vendor and product family for the picker.

    Product family is derived by splitting the material name on the em-dash separator
    (`Silkoxy H3 — 55 Gal Drum` → family `Silkoxy H3`, variant `55 Gal Drum`).
    Falls back to the full name if no em-dash is present.
    """
    cursor = db.materials.find({"is_deleted": {"$ne": True}}, {"_id": 0}).sort("name", 1)
    mats = await cursor.to_list(2000)
    vendors_map: dict = {}
    for m in mats:
        vname = (m.get("vendor_name") or "").strip() or "Unassigned"
        v = vendors_map.setdefault(vname, {"vendor_name": vname, "vendor_id": m.get("vendor_id"), "families": {}})
        # Split family / variant on " — " (em-dash) — the convention used by our importers
        full = (m.get("name") or "").strip()
        if " — " in full:
            family_name, variant_label = full.split(" — ", 1)
        else:
            family_name, variant_label = full, m.get("unit", "") or ""
        family = v["families"].setdefault(family_name, {
            "family": family_name,
            "category": m.get("category") or "",
            "variants": [],
        })
        loaded = float(m.get("default_price", 0) or 0) * (1 + float(m.get("shipping_pct", 0) or 0) / 100)
        family["variants"].append({
            "material_id": m["id"],
            "sku": m.get("sku") or "",
            "label": variant_label,
            "unit": m.get("unit", "") or "",
            "default_price": float(m.get("default_price", 0) or 0),
            "loaded_cost": round(loaded, 2),
            "notes": m.get("notes", "") or "",
        })

    # Materialize to list and sort
    out = []
    for v in vendors_map.values():
        families = sorted(v["families"].values(), key=lambda f: f["family"].lower())
        out.append({
            "vendor_name": v["vendor_name"],
            "vendor_id": v["vendor_id"],
            "family_count": len(families),
            "variant_count": sum(len(f["variants"]) for f in families),
            "families": families,
        })
    out.sort(key=lambda x: (x["vendor_name"] == "Unassigned", x["vendor_name"].lower()))
    return out


async def _build_po_dict(deal_id: str, vendor_id: str, current: dict) -> dict:
    """Build the dict that purchase_order_pdf.build_purchase_order_pdf consumes."""
    deal = await _check_deal_owner(deal_id, current)
    all_lines = [ln for ln in (deal.get("material_takeoff") or []) if ln.get("vendor_id") == vendor_id]
    if not all_lines:
        raise HTTPException(status_code=404, detail="No take-off lines for this vendor on this project")

    # Filter to only un-ordered lines? Keep all for now; future flag could split.
    vendor = await db.vendors.find_one({"id": vendor_id}, {"_id": 0}) or {"name": all_lines[0].get("vendor_name") or "Vendor"}

    # Resolve project address from property
    ship_to = {}
    if deal.get("property_id"):
        prop = await db.properties.find_one({"id": deal["property_id"]}, {"_id": 0})
        if prop:
            ship_to = {
                "address": prop.get("property_address", ""),
                "address_line2": prop.get("property_address_line2", ""),
                "city": prop.get("property_city", ""),
                "state": prop.get("property_state", ""),
                "zip": prop.get("property_zip", ""),
            }

    # Build the project_name / PO# convention: "<street>_<city>"
    street = (ship_to.get("address") or "").strip()
    city = (ship_to.get("city") or "").strip()
    project_name_fallback = (deal.get("title") or "").strip()
    if street and city:
        po_number = f"{street}_{city}"
    else:
        po_number = project_name_fallback or deal_id[:8]
    project_name = po_number

    # Requested-by = current user
    requested_by = {
        "name": current.get("name", ""),
        "title": current.get("title", ""),
        "phone": current.get("phone", ""),
        "email": current.get("email", ""),
    }

    po = {
        "po_number": po_number,
        "project_name": project_name,
        "po_date": datetime.now(timezone.utc).date().isoformat(),
        "ship_to": ship_to,
        "vendor": {
            "name": vendor.get("name", "") or "",
            "contact_name": vendor.get("contact_name", "") or "",
            "phone": vendor.get("phone", "") or vendor.get("mobile_phone", "") or vendor.get("work_phone", "") or "",
            "email": vendor.get("email", "") or "",
            "address": vendor.get("address", "") or "",
            "address_line2": vendor.get("address_line2", "") or "",
            "city": vendor.get("city", "") or "",
            "state": vendor.get("state", "") or "",
            "zip": vendor.get("zip_code", "") or "",
        },
        "requested_by": requested_by,
        "notes": "",
        "lines": [{
            "sku": ln.get("sku", ""),
            "name": ln.get("name", ""),
            "unit": ln.get("unit", ""),
            "quantity": ln.get("quantity", 0),
            "notes": ln.get("notes", ""),
        } for ln in all_lines],
    }
    return po


@api_router.get("/deals/{deal_id}/purchase-order/{vendor_id}.pdf")
async def deal_purchase_order_pdf(
    deal_id: str,
    vendor_id: str,
    token: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    """Generate a per-vendor PO / Material Take-Off PDF for a project. No prices on the PDF."""
    raw = None
    if authorization and authorization.startswith("Bearer "):
        raw = authorization[7:]
    elif token:
        raw = token
    if not raw:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(raw, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["sub"]})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        user.pop("_id", None)
        user.pop("password_hash", None)
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

    po = await _build_po_dict(deal_id, vendor_id, user)
    from purchase_order_pdf import build_purchase_order_pdf
    pdf_bytes = build_purchase_order_pdf(po)
    safe_name = po["po_number"].replace("/", "-").replace("\\", "-")
    fname = f"PO_{safe_name}_{po['vendor']['name'].replace(' ', '_')}.pdf"
    return Response(content=pdf_bytes, media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@api_router.post("/deals/{deal_id}/purchase-order/{vendor_id}/email")
async def email_purchase_order(deal_id: str, vendor_id: str, body: dict = Body(default={}), current=Depends(get_current_user)):
    """Email the per-vendor PO PDF to the vendor's contact email."""
    po = await _build_po_dict(deal_id, vendor_id, current)
    from purchase_order_pdf import build_purchase_order_pdf
    pdf_bytes = build_purchase_order_pdf(po)

    to_email = (body.get("to_email") or po["vendor"].get("email") or "").strip()
    cc_email = (body.get("cc_email") or "").strip()
    from_email = (body.get("from_email") or "").strip() or None
    if not from_email:
        # Default to the "projects" Send-As alias for purchase orders — POs
        # are project-side comms (vendor handoff, materials ordering) so they
        # group under projects@ per the team's correspondence-routing rule.
        from email_routing import get_from_for_category
        from_email = await get_from_for_category(db, "projects") or None
    if not to_email:
        raise HTTPException(status_code=400, detail="No recipient email — please provide one or set the vendor's email.")

    po_num = po["po_number"]
    vendor_name = po["vendor"]["name"] or "Vendor"
    project_name = po["project_name"]
    subject = f"Purchase Order — {po_num}"

    body_text = (
        f"Hi {po['vendor'].get('contact_name') or vendor_name},\n\n"
        f"Please find attached Purchase Order {po_num} for project {project_name}.\n\n"
        f"Could you confirm receipt, lead time, and pricing? Please call Darren Oliver at 720-715-9955 "
        f"if you have any questions or to discuss volume pricing.\n\n"
        f"Thank you,\n"
        f"{po['requested_by'].get('name') or 'SealTech Building Solutions'}\n"
        f"SealTech Building Solutions  ·  720-715-9955"
    )

    body_html = f"""
    <html><body style="font-family: Arial, Helvetica, sans-serif; color: #0A0A0A; max-width: 620px;">
      <p style="margin: 0 0 16px;">Hi {po['vendor'].get('contact_name') or vendor_name},</p>
      <p style="margin: 0 0 16px;">Please find attached <b>Purchase Order {po_num}</b> for project <b>{project_name}</b>.</p>
      <p style="margin: 16px 0;">Could you confirm receipt, lead time, and pricing? Please call <b>Darren Oliver at 720-715-9955</b> if you have any questions or to discuss volume pricing.</p>
      <p style="margin: 24px 0 0; padding-top: 16px; border-top: 1px solid #E4E4E7; color: #52525B; font-size: 12px;">
        <b style="color: #0A0A0A;">{po['requested_by'].get('name') or 'SealTech Building Solutions'}</b><br/>
        SealTech Building Solutions  ·  720-715-9955  ·  projects@sealtechsolutions.co
      </p>
    </body></html>
    """

    try:
        from email_sender import send_email, EmailNotConfigured
        result = send_email(
            to=to_email,
            cc=cc_email,
            subject=subject,
            body_text=body_text,
            body_html=body_html,
            reply_to=os.environ.get("GMAIL_FROM_EMAIL") or None,
            attachments=[{"filename": f"PO_{po_num}.pdf", "data": pdf_bytes, "mime": "application/pdf"}],
            from_email=from_email,
        )
    except EmailNotConfigured as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Email send failed: {type(e).__name__}: {e}")

    # Mark all this vendor's lines as ordered
    deal = await db.deals.find_one({"id": deal_id})
    lines = list(deal.get("material_takeoff") or [])
    for ln in lines:
        if ln.get("vendor_id") == vendor_id:
            ln["ordered"] = True
            ln["ordered_at"] = now_iso()
    cleaned = _persist_deal_takeoff(deal, lines)
    await db.deals.update_one({"id": deal_id}, {"$set": cleaned})

    return {
        "ok": True,
        "message": f"PO {po_num} emailed to {to_email}",
        "to_email": to_email,
        "message_id": result.get("message_id"),
    }




@api_router.get("/maintenance")
async def list_maintenance(current=Depends(get_current_user)):
    """Return all projects with maintenance_plan=True, with denormalized contact + property info."""
    query = {"is_deleted": {"$ne": True}, "maintenance_plan": True}
    if current.get("role") == "sales":
        query["$or"] = [{"assigned_to_user_id": current["id"]}, {"created_by_user_id": current["id"]}]
    deals = await db.deals.find(query, {"_id": 0}).sort("next_maintenance_date", 1).to_list(2000)
    today = datetime.now(timezone.utc).date().isoformat()
    soon_cutoff = (datetime.now(timezone.utc) + timedelta(days=30)).date().isoformat()
    out = []
    for d in deals:
        # Status
        nxt = d.get("next_maintenance_date") or ""
        if not nxt:
            status = "Unscheduled"
        elif nxt < today:
            status = "Overdue"
        elif nxt <= soon_cutoff:
            status = "Due Soon"
        else:
            status = "Upcoming"
        # Pull denormalized contact + property snippets
        contact_name = ""
        contact_phone = ""
        if d.get("customer_contact_id") or d.get("contact_id"):
            cid = d.get("customer_contact_id") or d.get("contact_id")
            cust = await db.contacts.find_one({"id": cid}, {"_id": 0})
            if cust:
                contact_name = cust.get("contact_name", "") or ""
                contact_phone = (cust.get("mobile_phone") or cust.get("phone") or cust.get("work_phone") or "").strip()
        property_name = ""
        property_address = ""
        if d.get("property_id"):
            prop = await db.properties.find_one({"id": d["property_id"]}, {"_id": 0})
            if prop:
                property_name = prop.get("property_name", "") or ""
                addr = prop.get("property_address", "") or ""
                city = prop.get("property_city", "") or ""
                st = prop.get("property_state", "") or ""
                zp = prop.get("property_zip", "") or ""
                tail = ", ".join([p for p in [city, st] if p])
                if zp:
                    tail = f"{tail} {zp}".strip()
                property_address = " · ".join([p for p in [addr, tail] if p])
        out.append({
            "id": d["id"],
            "title": d.get("title", ""),
            "contact_name": contact_name,
            "contact_phone": contact_phone,
            "property_name": property_name,
            "property_address": property_address,
            "maintenance_rate": d.get("maintenance_rate", 0),
            "maintenance_start_date": d.get("maintenance_start_date", ""),
            "last_maintenance_date": d.get("last_maintenance_date", ""),
            "next_maintenance_date": nxt,
            "status": status,
            "visit_count": len(d.get("maintenance_visits") or []),
        })
    return out


# ----- Invoices -----
# Number sequence seed: INV-2026-1100 (user-specified starting point)
INVOICE_SEQ_SEED = {"year": 2026, "next": 1100}


async def _next_invoice_number() -> str:
    """Atomic counter per year stored in `settings` collection (key=invoice_seq_{year})."""
    year = datetime.now(timezone.utc).year
    key = f"invoice_seq_{year}"
    seed = INVOICE_SEQ_SEED["next"] if year == INVOICE_SEQ_SEED["year"] else 1
    doc = await db.settings.find_one_and_update(
        {"key": key},
        {"$inc": {"value": 1}, "$setOnInsert": {"key": key}},
        upsert=True,
        return_document=True,
    )
    raw = doc.get("value", 1)
    # If we just inserted and value is 1 but year matches seed, jump to seed
    if year == INVOICE_SEQ_SEED["year"] and raw < INVOICE_SEQ_SEED["next"]:
        await db.settings.update_one({"key": key}, {"$set": {"value": INVOICE_SEQ_SEED["next"]}})
        raw = INVOICE_SEQ_SEED["next"]
    return f"INV-{year}-{raw:04d}"


def _recalc_invoice(inv: dict) -> dict:
    """Compute line-item amounts + subtotal + total + balance_due."""
    items = inv.get("line_items") or []
    sub = 0.0
    for it in items:
        if not it.get("id"):
            it["id"] = str(uuid.uuid4())
        try:
            qty = float(it.get("quantity") or 0)
            unit = float(it.get("unit_price") or 0)
        except (TypeError, ValueError):
            qty = 0.0
            unit = 0.0
        it["quantity"] = qty
        it["unit_price"] = unit
        it["amount"] = round(qty * unit, 2)
        sub += it["amount"]
    inv["line_items"] = items
    inv["subtotal"] = round(sub, 2)
    inv["total"] = round(sub, 2)  # no tax
    paid = float(inv.get("amount_paid") or 0)
    inv["amount_paid"] = round(paid, 2)
    inv["balance_due"] = round(inv["total"] - paid, 2)
    # Status auto-rules
    today_iso = datetime.now(timezone.utc).date().isoformat()
    status = inv.get("status", "Draft")
    # Always recompute from money flow when payment has been recorded, even on
    # Draft invoices — recording a payment is itself the "send" gesture, so a
    # paid-in-full Draft should never sit at Draft.
    if status != "Void":
        if paid > 0 and inv["balance_due"] <= 0.01:
            status = "Paid"
        elif paid > 0:
            status = "Partial"
        elif status != "Draft" and inv.get("due_date") and inv["due_date"] < today_iso:
            status = "Overdue"
    inv["status"] = status
    return inv


async def _build_bill_to_from_contact(contact_id: str) -> dict:
    """Pull billing address snapshot from a contact, defaulting to primary address if billing_same."""
    if not contact_id:
        return {}
    c = await db.contacts.find_one({"id": contact_id}, {"_id": 0})
    if not c:
        return {}
    same = c.get("billing_same_as_address", True)
    if same:
        return {
            "bill_to_company": c.get("company_name", ""),
            "bill_to_name": c.get("contact_name", ""),
            "bill_to_address": c.get("address", ""),
            "bill_to_address_line2": c.get("address_line2", ""),
            "bill_to_city": c.get("city", ""),
            "bill_to_state": c.get("state", ""),
            "bill_to_zip": c.get("zip_code", ""),
            "bill_to_email": c.get("email", ""),
        }
    return {
        "bill_to_company": c.get("company_name", ""),
        "bill_to_name": c.get("contact_name", ""),
        "bill_to_address": c.get("billing_address", ""),
        "bill_to_address_line2": c.get("billing_address_line2", ""),
        "bill_to_city": c.get("billing_city", ""),
        "bill_to_state": c.get("billing_state", ""),
        "bill_to_zip": c.get("billing_zip", ""),
        "bill_to_email": c.get("email", ""),
    }


@api_router.get("/invoices")
async def list_invoices(status: Optional[str] = None, deal_id: Optional[str] = None, current=Depends(get_current_user)):
    query = {"is_deleted": {"$ne": True}}
    if status:
        query["status"] = status
    if deal_id:
        query["deal_id"] = deal_id
    if current.get("role") == "sales":
        query["created_by_user_id"] = current["id"]
    cursor = db.invoices.find(query, {"_id": 0}).sort("created_at", -1)
    items = await cursor.to_list(2000)
    return items


async def _invoice_gl_warnings(inv: dict) -> list:
    """Build GL warnings for an invoice — currently surfaces period-lock blocks.
    Checked BEFORE issuing the hook so the toast reflects what the user just attempted."""
    warnings = []
    # Use payment_date if a payment was applied, else invoice_date for the issue posting
    inv_date = inv.get("invoice_date") or now_iso()[:10]
    pay_date = inv.get("payment_date") or inv_date
    paid = float(inv.get("amount_paid") or 0)
    status = (inv.get("status") or "").lower()
    if inv.get("entity_id") and status not in ("draft", "void"):
        lock = await gl.check_period_lock(db, inv["entity_id"], inv_date)
        if lock:
            warnings.append({
                "type": "period_locked",
                "side": "issuer",
                "entity_id": inv["entity_id"],
                "posting_date": inv_date,
                "lock_through": lock,
                "kind": "issue",
                "message": f"Invoice issuance posting deferred — entity is locked through {lock}. Reopen the period (Books → Period Close) to record this in the ledger.",
            })
        if paid > 0:
            lock_p = await gl.check_period_lock(db, inv["entity_id"], pay_date)
            if lock_p:
                warnings.append({
                    "type": "period_locked",
                    "side": "issuer",
                    "entity_id": inv["entity_id"],
                    "posting_date": pay_date,
                    "lock_through": lock_p,
                    "kind": "payment",
                    "message": f"Payment posting deferred — entity is locked through {lock_p}. Reopen the period to record this in the ledger.",
                })
    if inv.get("counter_entity_id") and status not in ("draft", "void"):
        lock_c = await gl.check_period_lock(db, inv["counter_entity_id"], inv_date)
        if lock_c:
            warnings.append({
                "type": "period_locked",
                "side": "counter",
                "entity_id": inv["counter_entity_id"],
                "posting_date": inv_date,
                "lock_through": lock_c,
                "kind": "issue_mirror",
                "message": f"Inter-Co mirror deferred — counter entity is locked through {lock_c}. Reopen the counter-entity's period to mirror this entry.",
            })
    return warnings


async def _bill_gl_warnings(bill: dict) -> list:
    warnings = []
    bill_date = bill.get("bill_date") or bill.get("received_date") or now_iso()[:10]
    pay_date = bill.get("paid_date") or bill_date
    paid = float(bill.get("paid_amount") or 0)
    status = (bill.get("status") or "").lower()
    if bill.get("entity_id") and status not in ("draft", "void"):
        lock = await gl.check_period_lock(db, bill["entity_id"], bill_date)
        if lock:
            warnings.append({
                "type": "period_locked",
                "side": "buyer",
                "entity_id": bill["entity_id"],
                "posting_date": bill_date,
                "lock_through": lock,
                "kind": "bill_received",
                "message": f"Bill posting deferred — entity is locked through {lock}. Reopen the period (Books → Period Close) to record this in the ledger.",
            })
        if paid > 0:
            lock_p = await gl.check_period_lock(db, bill["entity_id"], pay_date)
            if lock_p:
                warnings.append({
                    "type": "period_locked",
                    "side": "buyer",
                    "entity_id": bill["entity_id"],
                    "posting_date": pay_date,
                    "lock_through": lock_p,
                    "kind": "bill_payment",
                    "message": f"Bill payment posting deferred — entity is locked through {lock_p}. Reopen the period to record this in the ledger.",
                })
    if bill.get("counter_entity_id") and status not in ("draft", "void"):
        lock_c = await gl.check_period_lock(db, bill["counter_entity_id"], bill_date)
        if lock_c:
            warnings.append({
                "type": "period_locked",
                "side": "counter",
                "entity_id": bill["counter_entity_id"],
                "posting_date": bill_date,
                "lock_through": lock_c,
                "kind": "bill_received_mirror",
                "message": f"Inter-Co mirror deferred — counter entity is locked through {lock_c}. Reopen the counter-entity's period to mirror this entry.",
            })
    return warnings


@api_router.post("/invoices", response_model=Invoice)
async def create_invoice(body: InvoiceIn, current=Depends(get_current_user)):
    data = body.model_dump()
    # Pre-fill bill-to from contact if not provided
    if data.get("customer_contact_id") and not data.get("bill_to_company") and not data.get("bill_to_name"):
        bt = await _build_bill_to_from_contact(data["customer_contact_id"])
        for k, v in bt.items():
            if not data.get(k):
                data[k] = v
    # Pre-fill project info if deal_id supplied
    if data.get("deal_id"):
        deal = await db.deals.find_one({"id": data["deal_id"]}, {"_id": 0})
        if deal:
            if not data.get("project_title"):
                data["project_title"] = deal.get("title", "")
            if not data.get("project_total"):
                # Use chosen_amount if set, else MID proposal option (typical buy point)
                pt = float(deal.get("chosen_amount", 0) or 0)
                if pt <= 0:
                    pt = proposal_mid_amount(deal)
                # Add approved change orders
                co_total = sum(
                    float(co.get("amount", 0) or 0)
                    for co in (deal.get("change_orders") or [])
                    if (co.get("status") or "Approved") == "Approved"
                )
                data["project_total"] = round(pt + co_total, 2)
            if not data.get("customer_contact_id"):
                data["customer_contact_id"] = deal.get("customer_contact_id") or deal.get("contact_id")
            # Auto-fill bill-to from deal's contact if still empty
            if not data.get("bill_to_company") and not data.get("bill_to_name") and data.get("customer_contact_id"):
                bt = await _build_bill_to_from_contact(data["customer_contact_id"])
                for k, v in bt.items():
                    if not data.get(k):
                        data[k] = v
            # Property address
            if not data.get("project_address") and deal.get("property_id"):
                prop = await db.properties.find_one({"id": deal["property_id"]}, {"_id": 0})
                if prop:
                    addr1 = " ".join([p for p in [prop.get("property_address", ""), prop.get("property_address_line2", "")] if p]).strip()
                    line2 = ", ".join([p for p in [prop.get("property_city", ""), prop.get("property_state", "")] if p])
                    if prop.get("property_zip"):
                        line2 = f"{line2} {prop.get('property_zip')}".strip()
                    data["project_address"] = "  ·  ".join([p for p in [addr1, line2] if p])
    # Defaults
    if not data.get("invoice_date"):
        data["invoice_date"] = datetime.now(timezone.utc).date().isoformat()
    if not data.get("due_date"):
        data["due_date"] = data["invoice_date"]  # "Due Upon Receipt"
    if not data.get("terms"):
        data["terms"] = "Due Upon Receipt"
    data = _recalc_invoice(data)
    data["id"] = str(uuid.uuid4())
    data["invoice_number"] = await _next_invoice_number()
    data["created_at"] = now_iso()
    data["created_by_user_id"] = current["id"]
    data["is_deleted"] = False
    await db.invoices.insert_one(data.copy())
    # Books — auto-journal (no-op if entity_id is not set)
    gl_warnings = await _invoice_gl_warnings(data)
    try:
        await gl.post_invoice_issue(db, data, posted_by_user_id=current["id"])
        await gl.post_invoice_payment(db, data, posted_by_user_id=current["id"])
    except Exception as e:
        logger.warning(f"GL post (invoice create) failed: {type(e).__name__}: {e}")
    out = strip_id(data)
    if gl_warnings:
        out["gl_warnings"] = gl_warnings
    return out


@api_router.get("/invoices/{invoice_id}", response_model=Invoice)
async def get_invoice(invoice_id: str, current=Depends(get_current_user)):
    doc = await db.invoices.find_one({"id": invoice_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return doc


@api_router.put("/invoices/{invoice_id}", response_model=Invoice)
async def update_invoice(invoice_id: str, body: InvoiceIn, current=Depends(get_current_user)):
    existing = await db.invoices.find_one({"id": invoice_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Invoice not found")
    data = body.model_dump()
    # Preserve immutable fields
    data["id"] = existing["id"]
    data["invoice_number"] = existing["invoice_number"]
    data["created_at"] = existing["created_at"]
    data["created_by_user_id"] = existing.get("created_by_user_id")
    data["last_sent_at"] = existing.get("last_sent_at", "")
    data["pdf_generated_at"] = existing.get("pdf_generated_at", "")
    data = _recalc_invoice(data)
    await db.invoices.update_one({"id": invoice_id}, {"$set": data})
    # If entity changed, reverse old journals first
    gl_warnings = await _invoice_gl_warnings(data)
    try:
        if existing.get("entity_id") and existing.get("entity_id") != data.get("entity_id"):
            await gl.reverse_journals(db, source_type="invoice", source_id=invoice_id)
            await gl.reverse_journals(db, source_type="invoice_ic_mirror", source_id=invoice_id)
        # If counter-entity changed, also reverse the mirror so it doesn't dangle on the old entity
        if existing.get("counter_entity_id") != data.get("counter_entity_id"):
            await gl.reverse_journals(db, source_type="invoice_ic_mirror", source_id=invoice_id)
        await gl.post_invoice_issue(db, data, posted_by_user_id=current["id"])
        await gl.post_invoice_payment(db, data, posted_by_user_id=current["id"])
    except Exception as e:
        logger.warning(f"GL post (invoice update) failed: {type(e).__name__}: {e}")
    out = strip_id(data)
    if gl_warnings:
        out["gl_warnings"] = gl_warnings
    return out


@api_router.delete("/invoices/{invoice_id}")
async def delete_invoice(invoice_id: str, current=Depends(get_current_user)):
    if is_admin(current):
        await db.invoices.delete_one({"id": invoice_id})
    else:
        await db.invoices.update_one({"id": invoice_id}, {"$set": {"is_deleted": True, "deleted_at": now_iso(), "deleted_by": current["id"]}})
    # Reverse any GL postings tied to this invoice
    try:
        await gl.reverse_journals(db, source_type="invoice", source_id=invoice_id)
        await gl.reverse_journals(db, source_type="invoice_ic_mirror", source_id=invoice_id)
    except Exception as e:
        logger.warning(f"GL reverse (invoice delete) failed: {type(e).__name__}: {e}")
    return {"ok": True}


@api_router.post("/invoices/from-milestone")
async def invoice_from_milestone(deal_id: str = Body(...), milestone_id: str = Body(...), current=Depends(get_current_user)):
    """Auto-generate a draft invoice for a single milestone of a deal."""
    deal = await db.deals.find_one({"id": deal_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")
    ms = next((m for m in (deal.get("payment_milestones") or []) if m.get("id") == milestone_id), None)
    if not ms:
        raise HTTPException(status_code=404, detail="Milestone not found")
    body = InvoiceIn(
        deal_id=deal_id,
        customer_contact_id=deal.get("customer_contact_id") or deal.get("contact_id"),
        source_type="milestone",
        source_id=milestone_id,
        project_title=deal.get("title", ""),
        line_items=[InvoiceLineItem(description=f"{ms.get('label') or 'Project Milestone'} — {ms.get('percent', 0)}% of contract", quantity=1, unit_price=float(ms.get("amount") or 0))],
    )
    return await create_invoice(body, current)


@api_router.post("/invoices/from-maintenance-visit")
async def invoice_from_maintenance_visit(deal_id: str = Body(...), visit_id: str = Body(...), current=Depends(get_current_user)):
    """Auto-generate a draft invoice for a maintenance visit."""
    deal = await db.deals.find_one({"id": deal_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")
    visit = next((v for v in (deal.get("maintenance_visits") or []) if v.get("id") == visit_id), None)
    if not visit:
        raise HTTPException(status_code=404, detail="Visit not found")
    visit_date = visit.get("visit_date", "")
    desc_bits = [f"Annual Maintenance Visit ({visit_date})" if visit_date else "Annual Maintenance Visit"]
    if visit.get("notes"):
        desc_bits.append(visit["notes"])
    body = InvoiceIn(
        deal_id=deal_id,
        customer_contact_id=deal.get("customer_contact_id") or deal.get("contact_id"),
        source_type="maintenance_visit",
        source_id=visit_id,
        project_title=deal.get("title", ""),
        invoice_date=visit_date or datetime.now(timezone.utc).date().isoformat(),
        line_items=[InvoiceLineItem(description=" — ".join(desc_bits), quantity=1, unit_price=float(visit.get("amount") or 0))],
    )
    return await create_invoice(body, current)


@api_router.get("/invoices/{invoice_id}/pdf")
async def invoice_pdf(invoice_id: str, token: Optional[str] = Query(None), authorization: Optional[str] = Header(None)):
    # Allow ?token= for browser downloads
    raw = None
    if authorization and authorization.startswith("Bearer "):
        raw = authorization[7:]
    elif token:
        raw = token
    if not raw:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(raw, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["sub"]})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

    inv = await db.invoices.find_one({"id": invoice_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    from invoice_pdf import build_invoice_pdf
    _, inv_rate_pct = await _resolve_invoice_late_fee_rate(inv)
    pdf_bytes = build_invoice_pdf(inv, late_fee_rate_pct=inv_rate_pct)
    # Mark PDF generated
    await db.invoices.update_one({"id": invoice_id}, {"$set": {"pdf_generated_at": now_iso()}})
    fname = f"{inv['invoice_number']}.pdf"
    return Response(content=pdf_bytes, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@api_router.post("/invoices/{invoice_id}/email")
async def email_invoice(invoice_id: str, body: dict = Body(...), current=Depends(get_current_user)):
    """Send the invoice via Gmail SMTP with PDF attached. Marks invoice as Sent."""
    inv = await db.invoices.find_one({"id": invoice_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    to_email = (body.get("to_email") or inv.get("bill_to_email") or "").strip()
    cc_email = (body.get("cc_email") or inv.get("cc_email") or "").strip()
    from_email = (body.get("from_email") or "").strip() or None
    if not from_email:
        # Default to the "finance" Send-As alias for invoices.
        from email_routing import get_from_for_category
        from_email = await get_from_for_category(db, "finance") or None
    if not to_email:
        raise HTTPException(status_code=400, detail="No recipient email — please provide one.")

    # Generate the PDF in-memory
    from invoice_pdf import build_invoice_pdf
    # Resolve per-invoice late-fee rate (customer override → entity default → 1.5%)
    inv_rate, inv_rate_pct = await _resolve_invoice_late_fee_rate(inv)
    inv_rate_pct_str = (f"{inv_rate_pct:.2f}").rstrip("0").rstrip(".")
    pdf_bytes = build_invoice_pdf(inv, late_fee_rate_pct=inv_rate_pct)

    # Compose email
    inv_num = inv.get("invoice_number", "")
    total = float(inv.get("total") or 0)
    balance = float(inv.get("balance_due") or 0)
    bill_to = inv.get("bill_to_company") or inv.get("bill_to_name") or "—"
    project = inv.get("project_title") or ""
    due_date = inv.get("due_date") or "Due upon receipt"

    subject = f"Invoice {inv_num} from SealTech Building Solutions"
    if project:
        subject += f" — {project}"

    body_text = (
        f"Hello,\n\n"
        f"Please find attached Invoice {inv_num} for {bill_to}.\n\n"
        f"  Total:        ${total:,.2f}\n"
        f"  Balance Due:  ${balance:,.2f}\n"
        f"  Due Date:     {due_date}\n"
        f"  Terms:        {inv.get('terms', 'Due Upon Receipt')}\n\n"
        f"Remit payment to:\n"
        f"  SealTech Building Solutions\n"
        f"  2278 Mannatt Ct, Castle Rock, CO 80104\n\n"
        f"LATE FEE POLICY: A late fee of {inv_rate_pct_str}% per month is applied to any\n"
        f"balance more than 30 days past due. Fees compound monthly.\n\n"
        f"If you have any questions, please reply to this email.\n\n"
        f"Thank you for your business,\n"
        f"SealTech Building Solutions\n"
        f"720-715-9955  ·  finance@sealtechsolutions.co"
    )

    body_html = f"""
    <html><body style="font-family: Arial, Helvetica, sans-serif; color: #0A0A0A; max-width: 620px;">
      <p style="margin: 0 0 16px;">Hello,</p>
      <p style="margin: 0 0 16px;">Please find attached <b>Invoice {inv_num}</b> for {bill_to}.</p>
      <table style="border-collapse: collapse; margin: 16px 0;">
        <tr><td style="padding: 4px 16px 4px 0; color: #52525B; font-size: 13px;">Total</td><td style="padding: 4px 0; font-weight: bold; font-family: monospace;">${total:,.2f}</td></tr>
        <tr><td style="padding: 4px 16px 4px 0; color: #52525B; font-size: 13px;">Balance Due</td><td style="padding: 4px 0; font-weight: bold; font-family: monospace; color: #062B67;">${balance:,.2f}</td></tr>
        <tr><td style="padding: 4px 16px 4px 0; color: #52525B; font-size: 13px;">Due Date</td><td style="padding: 4px 0; font-family: monospace;">{due_date}</td></tr>
        <tr><td style="padding: 4px 16px 4px 0; color: #52525B; font-size: 13px;">Terms</td><td style="padding: 4px 0;">{inv.get('terms', 'Due Upon Receipt')}</td></tr>
      </table>
      <p style="margin: 16px 0 8px; color: #52525B; font-size: 13px;"><b>Remit payment to:</b></p>
      <p style="margin: 0 0 16px; line-height: 1.5;">
        SealTech Building Solutions<br/>
        2278 Mannatt Ct, Castle Rock, CO 80104
      </p>
      <p style="margin: 16px 0; padding: 10px 14px; background: #FFFBEB; border-left: 3px solid #B45309; color: #52525B; font-size: 12px;">
        <b style="color: #B45309;">Late Fee Policy:</b> A late fee of <b>{inv_rate_pct_str}% per month</b> is applied to any balance more than <b>30 days past due</b>. Fees compound monthly and are reflected on each Statement of Account.
      </p>
      <p style="margin: 16px 0;">If you have any questions, please reply to this email.</p>
      <p style="margin: 24px 0 0; padding-top: 16px; border-top: 1px solid #E4E4E7; color: #52525B; font-size: 12px;">
        <b style="color: #0A0A0A;">SealTech Building Solutions</b><br/>
        720-715-9955  ·  finance@sealtechsolutions.co
      </p>
    </body></html>
    """

    try:
        from email_sender import send_email, EmailNotConfigured
        result = send_email(
            to=to_email,
            cc=cc_email,
            subject=subject,
            body_text=body_text,
            body_html=body_html,
            reply_to=os.environ.get("GMAIL_FROM_EMAIL") or None,
            attachments=[{"filename": f"{inv_num}.pdf", "data": pdf_bytes, "mime": "application/pdf"}],
            from_email=from_email,
        )
    except EmailNotConfigured as e:
        raise HTTPException(status_code=500, detail=str(e))
    except smtplib.SMTPAuthenticationError as e:
        raise HTTPException(status_code=500, detail=f"Gmail authentication failed — check the App Password. ({e.smtp_code}: {e.smtp_error.decode() if e.smtp_error else ''})")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Email send failed: {type(e).__name__}: {e}")

    # Mark invoice as Sent (preserves Paid/Partial)
    patch = {
        "bill_to_email": to_email,
        "cc_email": cc_email,
        "last_sent_at": now_iso(),
    }
    if inv.get("status") in ("Draft", "Overdue"):
        patch["status"] = "Sent"
    await db.invoices.update_one({"id": invoice_id}, {"$set": patch})

    return {
        "ok": True,
        "mocked": False,
        "message": f"Invoice {inv_num} emailed to {to_email}" + (f" (cc: {cc_email})" if cc_email else ""),
        "to_email": to_email,
        "cc_email": cc_email,
        "message_id": result.get("message_id"),
    }


# ----- Statement of Account (per-customer aging) -----
async def _open_invoices_for_contact(contact_id: str) -> list:
    """All invoices that bill this contact and still have a balance (not Paid/Void/Draft)."""
    cursor = db.invoices.find(
        {
            "customer_contact_id": contact_id,
            "is_deleted": {"$ne": True},
            "status": {"$nin": ["Paid", "Void", "Draft"]},
        },
        {"_id": 0},
    ).sort("invoice_date", 1)
    out = []
    async for inv in cursor:
        bal = float(inv.get("balance_due") or 0)
        if bal > 0.01:
            out.append(inv)
    return out


@api_router.get("/contacts/{contact_id}/statement-summary")
async def statement_summary(contact_id: str, current=Depends(get_current_user)):
    """JSON preview of the statement (aging buckets + invoice count + total)."""
    contact = await db.contacts.find_one({"id": contact_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not contact:
        raise HTTPException(status_code=404, detail="Customer not found")
    invs = await _open_invoices_for_contact(contact_id)
    from statement_pdf import compute_aging
    as_of = datetime.now(timezone.utc).date()
    rate, rate_pct = await _customer_statement_late_fee_rate(contact, invs)
    aging = compute_aging(invs, as_of, rate=rate)
    return {
        "customer": {
            "id": contact["id"],
            "contact_name": contact.get("contact_name", ""),
            "company_name": contact.get("company_name", ""),
            "email": contact.get("email", ""),
        },
        "invoice_count": len(invs),
        "as_of": as_of.isoformat(),
        "late_fee_rate_pct": rate_pct,
        "aging": aging,
    }


@api_router.get("/customers-with-open-balance")
async def customers_with_open_balance(current=Depends(get_current_user)):
    """List every contact who has at least one open invoice, with their total balance.
    Drives the Statement of Account picker on the Invoices page."""
    pipeline = [
        {"$match": {
            "is_deleted": {"$ne": True},
            "status": {"$nin": ["Paid", "Void", "Draft"]},
            "balance_due": {"$gt": 0.01},
        }},
        {"$group": {
            "_id": "$customer_contact_id",
            "open_balance": {"$sum": "$balance_due"},
            "invoice_count": {"$sum": 1},
            "oldest_due": {"$min": "$due_date"},
        }},
        {"$match": {"_id": {"$ne": None}}},
        {"$sort": {"open_balance": -1}},
    ]
    rows = await db.invoices.aggregate(pipeline).to_list(500)
    # Hydrate contact details
    out = []
    for r in rows:
        cid = r["_id"]
        c = await db.contacts.find_one({"id": cid, "is_deleted": {"$ne": True}}, {"_id": 0, "id": 1, "contact_name": 1, "company_name": 1, "email": 1})
        if not c:
            continue
        out.append({
            "customer_id": cid,
            "contact_name": c.get("contact_name", ""),
            "company_name": c.get("company_name", ""),
            "email": c.get("email", ""),
            "open_balance": round(float(r.get("open_balance") or 0), 2),
            "invoice_count": int(r.get("invoice_count") or 0),
            "oldest_due": r.get("oldest_due") or "",
        })
    return out


@api_router.get("/contacts/{contact_id}/statement.pdf")
async def statement_pdf(
    contact_id: str,
    token: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    """Download the Statement of Account PDF for this customer (supports ?token= for browser links)."""
    raw = None
    if authorization and authorization.startswith("Bearer "):
        raw = authorization[7:]
    elif token:
        raw = token
    if not raw:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(raw, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["sub"]})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

    contact = await db.contacts.find_one({"id": contact_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not contact:
        raise HTTPException(status_code=404, detail="Customer not found")
    invs = await _open_invoices_for_contact(contact_id)
    from statement_pdf import build_statement_pdf
    as_of = datetime.now(timezone.utc).date().isoformat()
    rate, _ = await _customer_statement_late_fee_rate(contact, invs)
    pdf_bytes = build_statement_pdf(contact, invs, as_of, rate=rate)
    label = (contact.get("company_name") or contact.get("contact_name") or "customer").replace(" ", "_")
    fname = f"statement-{label}-{as_of}.pdf"
    return Response(content=pdf_bytes, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@api_router.post("/contacts/{contact_id}/statement/email")
async def email_statement(contact_id: str, body: dict = Body(...), current=Depends(get_current_user)):
    """Email the Statement of Account PDF to the customer via Gmail SMTP."""
    contact = await db.contacts.find_one({"id": contact_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not contact:
        raise HTTPException(status_code=404, detail="Customer not found")
    to_email = (body.get("to_email") or contact.get("email") or "").strip()
    cc_email = (body.get("cc_email") or "").strip()
    from_email = (body.get("from_email") or "").strip() or None
    if not from_email:
        # Default to the "finance" Send-As alias for statements / late notices.
        from email_routing import get_from_for_category
        from_email = await get_from_for_category(db, "finance") or None
    if not to_email:
        raise HTTPException(status_code=400, detail="No recipient email — please provide one.")

    invs = await _open_invoices_for_contact(contact_id)
    from statement_pdf import build_statement_pdf, compute_aging
    as_of_date = datetime.now(timezone.utc).date()
    as_of = as_of_date.isoformat()
    rate, rate_pct = await _customer_statement_late_fee_rate(contact, invs)
    pdf_bytes = build_statement_pdf(contact, invs, as_of, rate=rate)
    aging = compute_aging(invs, as_of_date, rate=rate)
    rate_pct_str = (f"{rate_pct:.2f}").rstrip("0").rstrip(".")

    cust_label = contact.get("company_name") or contact.get("contact_name") or "your account"
    grand = float(aging.get("total") or 0)
    late_fees = float(aging.get("late_fees") or 0)
    grand_with_fees = float(aging.get("total_due_with_fees") or grand)
    grand_str = f"${grand:,.2f}"
    fees_str = f"${late_fees:,.2f}"
    grand_with_fees_str = f"${grand_with_fees:,.2f}"
    has_fees = late_fees > 0.01

    subject = f"Statement of Account — {cust_label} — {as_of_date.strftime('%B %d, %Y')}"

    fees_line_text = f"  Late Fees ({rate_pct_str}%/mo): {fees_str}\n  TOTAL DUE:           {grand_with_fees_str}\n" if has_fees else ""
    body_text = (
        f"Hello,\n\n"
        f"Please find attached your current Statement of Account from SealTech Building Solutions, "
        f"covering all open invoices as of {as_of_date.strftime('%B %d, %Y')}.\n\n"
        f"  Open Invoices:       {len(invs)}\n"
        f"  Open Balance:        {grand_str}\n"
        f"{fees_line_text}"
        f"\n"
        f"Remit payment to:\n"
        f"  SealTech Building Solutions\n"
        f"  2278 Mannatt Ct, Castle Rock, CO 80104\n\n"
        f"LATE FEE POLICY: A late fee of {rate_pct_str}% per month is applied to any\n"
        f"balance more than 30 days past due. Fees compound monthly and are shown\n"
        f"on each Statement of Account.\n\n"
        f"If any of the invoices listed have already been paid, or if you have questions about any of them, "
        f"please reply to this email or call us at 720-715-9955 so we can reconcile your account.\n\n"
        f"Thank you for your business,\n"
        f"SealTech Building Solutions\n"
        f"720-715-9955  ·  finance@sealtechsolutions.co"
    )

    fees_html = ""
    if has_fees:
        fees_html = (
            f'<tr><td style="padding: 4px 16px 4px 0; color: #B45309; font-size: 13px;">Late Fees ({rate_pct_str}%/mo)</td>'
            f'<td style="padding: 4px 0; font-weight: bold; font-family: monospace; color: #B45309;">{fees_str}</td></tr>'
            f'<tr><td style="padding: 8px 16px 4px 0; color: #0A0A0A; font-size: 13px; font-weight: bold; border-top: 1px solid #0A0A0A;">TOTAL DUE</td>'
            f'<td style="padding: 8px 0 4px; font-weight: bold; font-family: monospace; color: #062B67; font-size: 15px; border-top: 1px solid #0A0A0A;">{grand_with_fees_str}</td></tr>'
        )

    body_html = f"""
    <html><body style="font-family: Arial, Helvetica, sans-serif; color: #0A0A0A; max-width: 620px;">
      <p style="margin: 0 0 16px;">Hello,</p>
      <p style="margin: 0 0 16px;">Please find attached your current <b>Statement of Account</b> from SealTech Building Solutions, covering all open invoices as of {as_of_date.strftime('%B %d, %Y')}.</p>
      <table style="border-collapse: collapse; margin: 16px 0;">
        <tr><td style="padding: 4px 16px 4px 0; color: #52525B; font-size: 13px;">Open Invoices</td><td style="padding: 4px 0; font-family: monospace;">{len(invs)}</td></tr>
        <tr><td style="padding: 4px 16px 4px 0; color: #52525B; font-size: 13px;">Open Balance</td><td style="padding: 4px 0; font-weight: bold; font-family: monospace;">{grand_str}</td></tr>
        {fees_html}
      </table>
      <p style="margin: 16px 0 8px; color: #52525B; font-size: 13px;"><b>Remit payment to:</b></p>
      <p style="margin: 0 0 16px; line-height: 1.5;">
        SealTech Building Solutions<br/>
        2278 Mannatt Ct, Castle Rock, CO 80104
      </p>
      <p style="margin: 16px 0; padding: 10px 14px; background: #FFFBEB; border-left: 3px solid #B45309; color: #52525B; font-size: 12px;">
        <b style="color: #B45309;">Late Fee Policy:</b> A late fee of <b>{rate_pct_str}% per month</b> is applied to any balance more than <b>30 days past due</b>. Fees compound monthly and are reflected on each Statement of Account.
      </p>
      <p style="margin: 16px 0;">If any of the invoices listed have already been paid, or if you have questions about any of them, please reply to this email or call us at 720-715-9955 so we can reconcile your account.</p>
      <p style="margin: 24px 0 0; padding-top: 16px; border-top: 1px solid #E4E4E7; color: #52525B; font-size: 12px;">
        <b style="color: #0A0A0A;">SealTech Building Solutions</b><br/>
        720-715-9955  ·  finance@sealtechsolutions.co
      </p>
    </body></html>
    """

    label = cust_label.replace(" ", "_")
    fname = f"statement-{label}-{as_of}.pdf"
    try:
        from email_sender import send_email, EmailNotConfigured
        result = send_email(
            to=to_email,
            cc=cc_email,
            subject=subject,
            body_text=body_text,
            body_html=body_html,
            reply_to=os.environ.get("GMAIL_FROM_EMAIL") or None,
            attachments=[{"filename": fname, "data": pdf_bytes, "mime": "application/pdf"}],
            from_email=from_email,
        )
    except EmailNotConfigured as e:
        raise HTTPException(status_code=500, detail=str(e))
    except smtplib.SMTPAuthenticationError as e:
        raise HTTPException(status_code=500, detail=f"Gmail authentication failed — check the App Password. ({e.smtp_code}: {e.smtp_error.decode() if e.smtp_error else ''})")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Email send failed: {type(e).__name__}: {e}")

    return {
        "ok": True,
        "mocked": False,
        "message": f"Statement of Account emailed to {to_email}" + (f" (cc: {cc_email})" if cc_email else ""),
        "to_email": to_email,
        "cc_email": cc_email,
        "message_id": result.get("message_id"),
        "total_balance": round(grand, 2),
        "late_fees": round(late_fees, 2),
        "total_due_with_fees": round(grand_with_fees, 2),
        "invoice_count": len(invs),
    }


# ----- Vendor Bills (Payables) -----
def _recalc_bill(bill: dict) -> dict:
    """Compute subtotal/total from line items, and auto-fill due_date from terms if missing."""
    items = bill.get("line_items") or []
    sub = 0.0
    for it in items:
        if not it.get("id"):
            it["id"] = str(uuid.uuid4())
        try:
            qty = float(it.get("quantity") or 0)
            unit = float(it.get("unit_price") or 0)
        except (TypeError, ValueError):
            qty = 0.0
            unit = 0.0
        amt = float(it.get("amount") or 0)
        if not amt and qty and unit:
            amt = round(qty * unit, 2)
        it["quantity"] = qty
        it["unit_price"] = unit
        it["amount"] = round(amt, 2)
        sub += it["amount"]
    bill["line_items"] = items
    # If user provided a total, keep it; else compute from line items + tax + shipping
    if not float(bill.get("total") or 0):
        tax = float(bill.get("tax") or 0)
        shipping = float(bill.get("shipping") or 0)
        bill["subtotal"] = round(sub, 2)
        bill["total"] = round(sub + tax + shipping, 2)
    # Auto-fill paid_amount when status = Paid
    if (bill.get("status") or "").lower() == "paid":
        total = float(bill.get("total") or 0)
        paid = float(bill.get("paid_amount") or 0)
        if paid < total - 0.01:
            bill["paid_amount"] = total
        if not bill.get("paid_date"):
            bill["paid_date"] = datetime.now(timezone.utc).date().isoformat()
    # Auto-compute due_date from terms when missing
    if not bill.get("due_date") and bill.get("bill_date"):
        try:
            base = datetime.fromisoformat(bill["bill_date"][:10]).date()
            terms = (bill.get("terms") or "").lower()
            days = 0
            if "net" in terms:
                m = re.search(r"net\s*(\d+)", terms)
                if m:
                    days = int(m.group(1))
            if days > 0:
                bill["due_date"] = (base + timedelta(days=days)).isoformat()
            else:
                bill["due_date"] = base.isoformat()
        except (ValueError, TypeError):
            pass
    return bill


@api_router.get("/vendor-bills")
async def list_vendor_bills(status: Optional[str] = None, vendor_id: Optional[str] = None, project_id: Optional[str] = None, current=Depends(get_current_user)):
    query = {"is_deleted": {"$ne": True}}
    if status:
        query["status"] = status
    if vendor_id:
        query["vendor_id"] = vendor_id
    if project_id:
        query["line_items.project_id"] = project_id
    if current.get("role") == "sales":
        query["created_by_user_id"] = current["id"]
    items = await db.vendor_bills.find(query, {"_id": 0}).sort("bill_date", -1).to_list(2000)
    return items


@api_router.post("/vendor-bills", response_model=VendorBill)
async def create_vendor_bill(body: VendorBillIn, current=Depends(get_current_user)):
    data = body.model_dump()
    # Snapshot vendor name from vendor_id if missing
    if data.get("vendor_id") and not data.get("vendor_name"):
        v = await db.vendors.find_one({"id": data["vendor_id"]}, {"_id": 0})
        if v:
            data["vendor_name"] = v.get("name", "")
    # Snapshot project titles on each line item
    for li in data.get("line_items", []) or []:
        if li.get("project_id") and not li.get("project_title"):
            d = await db.deals.find_one({"id": li["project_id"]}, {"_id": 0})
            if d:
                li["project_title"] = d.get("title", "")
    if not data.get("received_date"):
        data["received_date"] = datetime.now(timezone.utc).date().isoformat()
    data = _recalc_bill(data)
    data["id"] = str(uuid.uuid4())
    data["created_at"] = now_iso()
    data["created_by_user_id"] = current["id"]
    data["is_deleted"] = False
    await db.vendor_bills.insert_one(data.copy())
    gl_warnings = await _bill_gl_warnings(data)
    try:
        await gl.post_bill_received(db, data, posted_by_user_id=current["id"])
        await gl.post_bill_payment(db, data, posted_by_user_id=current["id"])
    except Exception as e:
        logger.warning(f"GL post (vendor bill create) failed: {type(e).__name__}: {e}")
    out = strip_id(data)
    if gl_warnings:
        out["gl_warnings"] = gl_warnings
    return out


@api_router.get("/vendor-bills/{bill_id}", response_model=VendorBill)
async def get_vendor_bill(bill_id: str, current=Depends(get_current_user)):
    doc = await db.vendor_bills.find_one({"id": bill_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Bill not found")
    return doc


@api_router.put("/vendor-bills/{bill_id}", response_model=VendorBill)
async def update_vendor_bill(bill_id: str, body: VendorBillIn, current=Depends(get_current_user)):
    existing = await db.vendor_bills.find_one({"id": bill_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Bill not found")
    data = body.model_dump()
    # Snapshot project titles
    for li in data.get("line_items", []) or []:
        if li.get("project_id") and not li.get("project_title"):
            d = await db.deals.find_one({"id": li["project_id"]}, {"_id": 0})
            if d:
                li["project_title"] = d.get("title", "")
    data["id"] = existing["id"]
    data["created_at"] = existing["created_at"]
    data["created_by_user_id"] = existing.get("created_by_user_id")
    data = _recalc_bill(data)
    await db.vendor_bills.update_one({"id": bill_id}, {"$set": data})
    gl_warnings = await _bill_gl_warnings(data)
    try:
        if existing.get("entity_id") and existing.get("entity_id") != data.get("entity_id"):
            await gl.reverse_journals(db, source_type="vendor_bill", source_id=bill_id)
            await gl.reverse_journals(db, source_type="vendor_bill_ic_mirror", source_id=bill_id)
        if existing.get("counter_entity_id") != data.get("counter_entity_id"):
            await gl.reverse_journals(db, source_type="vendor_bill_ic_mirror", source_id=bill_id)
        await gl.post_bill_received(db, data, posted_by_user_id=current["id"])
        await gl.post_bill_payment(db, data, posted_by_user_id=current["id"])
    except Exception as e:
        logger.warning(f"GL post (vendor bill update) failed: {type(e).__name__}: {e}")
    out = strip_id(data)
    if gl_warnings:
        out["gl_warnings"] = gl_warnings
    return out


@api_router.delete("/vendor-bills/{bill_id}")
async def delete_vendor_bill(bill_id: str, current=Depends(get_current_user)):
    if is_admin(current):
        await db.vendor_bills.delete_one({"id": bill_id})
    else:
        await db.vendor_bills.update_one({"id": bill_id}, {"$set": {"is_deleted": True, "deleted_at": now_iso(), "deleted_by": current["id"]}})
    try:
        await gl.reverse_journals(db, source_type="vendor_bill", source_id=bill_id)
        await gl.reverse_journals(db, source_type="vendor_bill_ic_mirror", source_id=bill_id)
    except Exception as e:
        logger.warning(f"GL reverse (vendor bill delete) failed: {type(e).__name__}: {e}")
    return {"ok": True}


# ----- Bulk CSV Import (Vendor Bills) -----

CSV_COLUMN_ALIASES = {
    # canonical          -> accepted header variants (lowercased, stripped)
    "vendor":            ["vendor", "vendor_name", "vendor name", "supplier", "payee"],
    "bill_number":       ["bill_number", "bill number", "bill #", "invoice_number", "invoice number", "invoice #", "ref", "reference"],
    "bill_date":         ["bill_date", "bill date", "date", "invoice_date", "invoice date"],
    "due_date":          ["due_date", "due date", "due", "payment_due"],
    "description":       ["description", "memo", "notes", "line_description", "details"],
    "amount":            ["amount", "total", "amount_due", "total_amount", "subtotal"],
    "expense_account":   ["expense_account", "expense account", "account", "account_number", "gl_account", "gl"],
    "project":           ["project", "project_name", "project name", "deal", "job"],
}


def _normalize_csv_headers(fieldnames: list) -> dict[str, str]:
    """Maps canonical column names → the actual header string in the CSV."""
    out: dict[str, str] = {}
    if not fieldnames:
        return out
    lookup = {(h or "").strip().lower(): h for h in fieldnames}
    for canon, aliases in CSV_COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in lookup:
                out[canon] = lookup[alias]
                break
    return out


def _parse_csv_date(s: str) -> str:
    """Accept ISO YYYY-MM-DD, MM/DD/YYYY, M/D/YY, or empty. Returns ISO or empty."""
    if not s:
        return ""
    s = s.strip()
    if not s:
        return ""
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return ""


def _parse_csv_amount(s: str) -> Optional[float]:
    """Lenient amount parser: strips $, commas, parens (negative)."""
    if s is None:
        return None
    raw = str(s).strip()
    if not raw:
        return None
    neg = raw.startswith("(") and raw.endswith(")")
    cleaned = raw.replace("$", "").replace(",", "").replace("(", "").replace(")", "").strip()
    try:
        v = float(cleaned)
        return -v if neg else v
    except (ValueError, TypeError):
        return None


async def _resolve_csv_vendor(vendor_name: str) -> Optional[dict]:
    """Try to match a CSV vendor cell to an existing vendor (case-insensitive exact, then prefix)."""
    if not vendor_name:
        return None
    name = vendor_name.strip()
    # Exact case-insensitive
    v = await db.vendors.find_one({"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}}, {"_id": 0})
    if v:
        return v
    # Prefix
    v = await db.vendors.find_one({"name": {"$regex": f"^{re.escape(name[:20])}", "$options": "i"}}, {"_id": 0})
    return v


async def _resolve_csv_expense_account(entity_id: str, provided: str, vendor: Optional[dict]) -> dict:
    """Resolve to a real account doc. provided may be an account number, a name, or blank.
    Falls back to gl.cogs_account_for(vendor)."""
    candidates = []
    p = (provided or "").strip()
    if p:
        # Try as number first
        a = await db.chart_of_accounts.find_one({"entity_id": entity_id, "number": p, "is_active": True}, {"_id": 0})
        if a:
            return {"account": a, "source": "csv-number"}
        # Try as exact name
        a = await db.chart_of_accounts.find_one({"entity_id": entity_id, "name": {"$regex": f"^{re.escape(p)}$", "$options": "i"}, "is_active": True}, {"_id": 0})
        if a:
            return {"account": a, "source": "csv-name"}
        candidates.append(p)
    # Fallback to vendor category default
    from gl import cogs_account_for
    default_num = cogs_account_for(vendor)
    a = await db.chart_of_accounts.find_one({"entity_id": entity_id, "number": default_num, "is_active": True}, {"_id": 0})
    if a:
        return {"account": a, "source": "vendor-default"}
    return {"account": None, "source": "missing", "tried": candidates}


@api_router.post("/vendor-bills/csv-preview")
async def vendor_bills_csv_preview(
    file: UploadFile = File(...),
    entity_id: str = Form(...),
    current=Depends(get_current_user),
):
    """Parse a CSV of vendor bills and return a per-row preview with vendor match,
    suggested expense account, and the GL impact lines that WOULD post (DR expense,
    CR AP) — without committing anything. The frontend renders this as a review table
    where the user can fix any flagged rows before clicking "Commit".

    Required CSV columns (case-insensitive, header synonyms supported):
       vendor (or vendor_name), amount
    Optional:
       bill_number, bill_date, due_date, description, expense_account (number or name)
    """
    import csv
    import io as _io

    ent = await db.entities.find_one({"id": entity_id})
    if not ent:
        raise HTTPException(status_code=404, detail="Entity not found")

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Empty file")
    if len(raw) > 5 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="CSV too large (max 5 MB)")
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = raw.decode("latin-1")
        except UnicodeDecodeError:
            raise HTTPException(status_code=400, detail="Could not decode file — please save as UTF-8 CSV")

    reader = csv.DictReader(_io.StringIO(text))
    headers = _normalize_csv_headers(reader.fieldnames or [])
    missing_required = [c for c in ("vendor", "amount") if c not in headers]
    if missing_required:
        return {
            "ok": False,
            "header_error": f"CSV is missing required columns: {', '.join(missing_required)}. Found: {reader.fieldnames}",
            "preview": [],
            "summary": {"total_rows": 0, "valid_rows": 0, "error_rows": 0, "total_amount": 0.0},
        }

    preview = []
    total_amount = 0.0
    valid_rows = 0
    error_rows = 0

    for row_num, raw_row in enumerate(reader, start=2):  # row 1 was header
        errors = []
        vendor_str = (raw_row.get(headers.get("vendor", "")) or "").strip()
        amount = _parse_csv_amount(raw_row.get(headers.get("amount", ""), ""))
        if not vendor_str:
            errors.append("Vendor name is required")
        if amount is None:
            errors.append("Amount could not be parsed")
        elif amount <= 0:
            errors.append("Amount must be greater than zero")

        vendor = await _resolve_csv_vendor(vendor_str) if vendor_str else None
        if vendor_str and not vendor:
            errors.append(f'Vendor "{vendor_str}" not found — create it before importing, or fix the name')

        provided_acct = (raw_row.get(headers.get("expense_account", ""), "") or "").strip() if "expense_account" in headers else ""
        acct_result = await _resolve_csv_expense_account(entity_id, provided_acct, vendor)
        expense_acct = acct_result["account"]
        if not expense_acct:
            errors.append(f'Expense account not found (provided: "{provided_acct or "—"}", fallback also missing)')

        bill_date = _parse_csv_date(raw_row.get(headers.get("bill_date", ""), "")) if "bill_date" in headers else ""
        if not bill_date:
            bill_date = datetime.now(timezone.utc).date().isoformat()
        due_date = _parse_csv_date(raw_row.get(headers.get("due_date", ""), "")) if "due_date" in headers else ""
        if not due_date:
            due_date = bill_date  # Due-on-receipt

        bill_number = (raw_row.get(headers.get("bill_number", ""), "") or "").strip() if "bill_number" in headers else ""
        description = (raw_row.get(headers.get("description", ""), "") or "").strip() if "description" in headers else ""
        if not description:
            description = f"{vendor_str} bill" if vendor_str else "Imported bill"

        # GL preview lines (what would post on create)
        gl_lines = []
        if expense_acct and amount and amount > 0:
            gl_lines = [
                {"side": "DR", "account_number": expense_acct["number"], "account_name": expense_acct["name"], "amount": round(amount, 2)},
                {"side": "CR", "account_number": "2000", "account_name": "Accounts Payable", "amount": round(amount, 2)},
            ]

        is_valid = len(errors) == 0
        if is_valid:
            valid_rows += 1
            total_amount += amount or 0
        else:
            error_rows += 1

        preview.append({
            "row": row_num,
            "vendor_input": vendor_str,
            "vendor_id": (vendor or {}).get("id"),
            "vendor_name": (vendor or {}).get("name") or vendor_str,
            "vendor_matched": bool(vendor),
            "bill_number": bill_number,
            "bill_date": bill_date,
            "due_date": due_date,
            "description": description,
            "amount": round(amount or 0, 2),
            "expense_account_id": (expense_acct or {}).get("id"),
            "expense_account_number": (expense_acct or {}).get("number") or provided_acct,
            "expense_account_name": (expense_acct or {}).get("name") or "",
            "expense_account_source": acct_result.get("source"),
            "gl_lines": gl_lines,
            "errors": errors,
            "valid": is_valid,
        })

    return {
        "ok": True,
        "preview": preview,
        "summary": {
            "total_rows": len(preview),
            "valid_rows": valid_rows,
            "error_rows": error_rows,
            "total_amount": round(total_amount, 2),
        },
    }


class CsvCommitRow(BaseModel):
    model_config = ConfigDict(extra="ignore")
    vendor_id: Optional[str] = None
    vendor_name: str = ""
    bill_number: str = ""
    bill_date: str = ""
    due_date: str = ""
    description: str = ""
    amount: float
    expense_account_id: Optional[str] = None
    expense_account_number: str = ""


class CsvCommitIn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    entity_id: str
    rows: List[CsvCommitRow]


@api_router.post("/vendor-bills/csv-commit")
async def vendor_bills_csv_commit(body: CsvCommitIn, current=Depends(get_current_user)):
    """Commit a previewed CSV import. Each row is validated again (idempotency: rows missing
    a vendor_id or expense_account are skipped). Each created bill runs through the normal
    GL posting (DR expense / CR AP) — same as a manual Add Bill. Returns per-row results."""
    ent = await db.entities.find_one({"id": body.entity_id})
    if not ent:
        raise HTTPException(status_code=404, detail="Entity not found")
    if not body.rows:
        raise HTTPException(status_code=400, detail="No rows to commit")

    created = []
    skipped = []
    for r in body.rows:
        if not r.vendor_id:
            skipped.append({"vendor_name": r.vendor_name, "reason": "Vendor not matched"})
            continue
        if r.amount is None or r.amount <= 0:
            skipped.append({"vendor_name": r.vendor_name, "reason": "Invalid amount"})
            continue
        if not r.expense_account_id and not r.expense_account_number:
            skipped.append({"vendor_name": r.vendor_name, "reason": "Expense account missing"})
            continue

        # Build the VendorBill doc inline (mirrors create_vendor_bill but bypasses the body model
        # since we trust the previewed/validated rows from the same session).
        line_id = str(uuid.uuid4())
        bill_id = str(uuid.uuid4())
        bill_date = r.bill_date or datetime.now(timezone.utc).date().isoformat()
        due_date = r.due_date or bill_date
        data = {
            "id": bill_id,
            "vendor_id": r.vendor_id,
            "vendor_name": r.vendor_name,
            "entity_id": body.entity_id,
            "counter_entity_id": None,
            "bill_number": r.bill_number,
            "bill_date": bill_date,
            "received_date": datetime.now(timezone.utc).date().isoformat(),
            "due_date": due_date,
            "terms": "Due on Receipt",
            "total": round(r.amount, 2),
            "subtotal": round(r.amount, 2),
            "tax": 0.0,
            "shipping": 0.0,
            "status": "Pending",
            "notes": "Imported from CSV bulk upload",
            "attached_file_id": None,
            "parsed_by_ai": False,
            "line_items": [{
                "id": line_id,
                "description": r.description or r.vendor_name,
                "sku": "",
                "project_id": None,
                "project_title": "",
                "quantity": 1.0,
                "unit_price": round(r.amount, 2),
                "amount": round(r.amount, 2),
                "takeoff_line_id": None,
                # CSV path explicitly carries the expense override so GL routes correctly:
                "expense_account_id": r.expense_account_id,
                "expense_account_number": r.expense_account_number,
            }],
            "paid_amount": 0.0,
            "paid_date": "",
            "paid_method": "",
            "paid_reference": "",
            "created_at": now_iso(),
            "created_by_user_id": current["id"],
            "is_deleted": False,
            "csv_import": True,
        }
        await db.vendor_bills.insert_one(data.copy())
        try:
            await gl.post_bill_received(db, data, posted_by_user_id=current["id"])
        except Exception as e:
            logger.warning(f"GL post (CSV import bill {bill_id}) failed: {type(e).__name__}: {e}")
        created.append({
            "id": bill_id,
            "vendor_name": r.vendor_name,
            "bill_number": r.bill_number,
            "amount": r.amount,
        })

    return {
        "ok": True,
        "created_count": len(created),
        "skipped_count": len(skipped),
        "created": created,
        "skipped": skipped,
    }


@api_router.post("/vendor-bills/parse")
async def parse_vendor_bill(file: UploadFile = File(...), current=Depends(get_current_user)):
    """Upload a vendor invoice (PDF/image), parse with Gemini Vision, return suggested structured data
    + suggested vendor match + suggested project matches per line item.
    Does NOT save anything — the frontend opens an editor pre-filled for user review."""
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Empty file")
    if len(raw) > 25 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (max 25 MB)")
    try:
        from vendor_bill_parser import parse_invoice_bytes
        parsed = await parse_invoice_bytes(raw, file.filename or "invoice.pdf", file.content_type)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Parse failed: {type(e).__name__}: {e}")

    # Match vendor
    suggested_vendor_id = None
    vname = (parsed.get("vendor_name") or "").strip().lower()
    if vname:
        all_vendors = await db.vendors.find({"is_deleted": {"$ne": True}}, {"_id": 0, "id": 1, "name": 1}).to_list(500)
        # Exact match first
        for v in all_vendors:
            if (v.get("name") or "").strip().lower() == vname:
                suggested_vendor_id = v["id"]
                break
        # Substring match fallback
        if not suggested_vendor_id:
            for v in all_vendors:
                vn = (v.get("name") or "").strip().lower()
                if vn and (vn in vname or vname in vn):
                    suggested_vendor_id = v["id"]
                    break

    # Match projects against PO number / line item description
    suggested_project_matches = {}
    deals = await db.deals.find({"is_deleted": {"$ne": True}}, {"_id": 0, "id": 1, "title": 1}).to_list(1000)
    po = (parsed.get("po_number") or "").strip().lower()
    for i, line in enumerate(parsed.get("line_items") or []):
        desc_l = (line.get("description") or "").lower()
        haystack = f"{desc_l} {po}".strip()
        if not haystack:
            continue
        for d in deals:
            t = (d.get("title") or "").strip().lower()
            if not t:
                continue
            if t in haystack:
                suggested_project_matches[str(i)] = d["id"]
                break

    # Upload the file to object storage so it's available as an attachment
    attached_file_id = None
    try:
        ext = (file.filename.rsplit(".", 1)[-1] if file.filename and "." in file.filename else "pdf").lower()
        file_id = str(uuid.uuid4())
        storage_path = f"{APP_NAME}/uploads/vendor_bill/pending/{file_id}.{ext}"
        put_result = put_object(storage_path, raw, file.content_type or "application/pdf")
        await db.files.insert_one({
            "id": file_id,
            "parent_type": "vendor_bill",
            "parent_id": "pending",
            "category": "Invoice",
            "storage_path": put_result["path"],
            "original_filename": file.filename or "invoice.pdf",
            "content_type": file.content_type or "application/pdf",
            "size": len(raw),
            "is_deleted": False,
            "uploaded_by": current["id"],
            "created_at": now_iso(),
        })
        attached_file_id = file_id
    except Exception as e:
        logger.warning(f"Vendor bill PDF stash failed: {e}")
        attached_file_id = None

    return {
        "parsed": parsed,
        "suggested_vendor_id": suggested_vendor_id,
        "suggested_project_matches": suggested_project_matches,
        "attached_file_id": attached_file_id,
    }


@api_router.get("/payables/report")
async def payables_report(current=Depends(get_current_user)):
    """List bills due within the next 7 days plus all currently overdue, grouped by vendor."""
    today = datetime.now(timezone.utc).date()
    horizon = today + timedelta(days=7)
    today_iso = today.isoformat()
    horizon_iso = horizon.isoformat()

    query = {
        "is_deleted": {"$ne": True},
        "status": {"$in": ["Pending", "Approved"]},
    }
    if current.get("role") == "sales":
        query["created_by_user_id"] = current["id"]
    bills = await db.vendor_bills.find(query, {"_id": 0}).to_list(2000)

    overdue = []
    due_this_week = []
    for b in bills:
        dd = (b.get("due_date") or "")[:10]
        if not dd:
            continue
        if dd < today_iso:
            overdue.append(b)
        elif dd <= horizon_iso:
            due_this_week.append(b)

    def group_by_vendor(bs):
        groups = {}
        for b in bs:
            key = b.get("vendor_id") or b.get("vendor_name") or "Unknown"
            grp = groups.setdefault(key, {"vendor_id": b.get("vendor_id"), "vendor_name": b.get("vendor_name") or "Unknown", "bills": [], "total": 0.0})
            grp["bills"].append(b)
            grp["total"] += float(b.get("total") or 0) - float(b.get("paid_amount") or 0)
        rows = list(groups.values())
        for r in rows:
            r["total"] = round(r["total"], 2)
            r["bills"].sort(key=lambda x: (x.get("due_date") or ""))
        rows.sort(key=lambda x: -x["total"])
        return rows

    overdue.sort(key=lambda b: (b.get("due_date") or ""))
    due_this_week.sort(key=lambda b: (b.get("due_date") or ""))
    return {
        "today": today_iso,
        "horizon": horizon_iso,
        "overdue": group_by_vendor(overdue),
        "due_this_week": group_by_vendor(due_this_week),
        "overdue_count": len(overdue),
        "due_this_week_count": len(due_this_week),
        "overdue_total": round(sum(float(b.get("total") or 0) - float(b.get("paid_amount") or 0) for b in overdue), 2),
        "due_this_week_total": round(sum(float(b.get("total") or 0) - float(b.get("paid_amount") or 0) for b in due_this_week), 2),
    }


@api_router.get("/payables/report.{fmt}")
async def payables_report_export(fmt: str, current=Depends(get_current_user)):
    """Export the payables report as Excel or PDF."""
    report = await payables_report(current)
    if fmt == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "Payables"
        headers = ["Bucket", "Vendor", "Bill #", "Bill Date", "Due Date", "Terms", "Total", "Paid", "Balance"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill("solid", fgColor="B91C1C")
            cell.alignment = Alignment(horizontal="left")
        # Overdue
        for grp in report["overdue"]:
            for b in grp["bills"]:
                ws.append(["OVERDUE", grp["vendor_name"], b.get("bill_number", ""), b.get("bill_date", ""), b.get("due_date", ""), b.get("terms", ""), float(b.get("total") or 0), float(b.get("paid_amount") or 0), round(float(b.get("total") or 0) - float(b.get("paid_amount") or 0), 2)])
        for grp in report["due_this_week"]:
            for b in grp["bills"]:
                ws.append(["DUE THIS WEEK", grp["vendor_name"], b.get("bill_number", ""), b.get("bill_date", ""), b.get("due_date", ""), b.get("terms", ""), float(b.get("total") or 0), float(b.get("paid_amount") or 0), round(float(b.get("total") or 0) - float(b.get("paid_amount") or 0), 2)])
        ws.freeze_panes = "A2"
        for i in range(1, 10):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 18
        buf = BytesIO()
        wb.save(buf)
        return Response(content=buf.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": 'attachment; filename="sealtech-payables.xlsx"'})
    raise HTTPException(status_code=400, detail="Use fmt=xlsx")


@api_router.post("/payables/email")
async def email_payables_report(body: dict = Body(default={}), current=Depends(get_current_user)):
    """Manually trigger a 'this week's payables' email to the user (or any address provided)."""
    to_email = (body.get("to_email") or os.environ.get("GMAIL_FROM_EMAIL") or "").strip()
    if not to_email:
        raise HTTPException(status_code=400, detail="Recipient email required")
    report = await payables_report(current)
    # Build the email
    try:
        from email_sender import send_for_category
        html = _render_payables_email_html(report)
        text = _render_payables_email_text(report)
        result = await send_for_category(
            db, "finance",
            to=to_email,
            subject=f"SealTech Payables — Week of {report['today']} — {report['overdue_count']} overdue · {report['due_this_week_count']} due",
            body_text=text,
            body_html=html,
        )
        return {"ok": True, "message_id": result.get("message_id"), "to": to_email}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Email failed: {type(e).__name__}: {e}")


def _render_payables_email_text(report: dict) -> str:
    lines = [
        f"SealTech Payables — Week of {report['today']}",
        "",
        f"OVERDUE ({report['overdue_count']} bills · ${report['overdue_total']:,.2f})",
    ]
    for grp in report["overdue"]:
        lines.append(f"  {grp['vendor_name']}  ${grp['total']:,.2f}")
        for b in grp["bills"]:
            balance = float(b.get("total") or 0) - float(b.get("paid_amount") or 0)
            lines.append(f"    - {b.get('bill_number') or '-'}  due {b.get('due_date')}  ${balance:,.2f}")
    lines += ["", f"DUE THIS WEEK ({report['due_this_week_count']} bills · ${report['due_this_week_total']:,.2f})"]
    for grp in report["due_this_week"]:
        lines.append(f"  {grp['vendor_name']}  ${grp['total']:,.2f}")
        for b in grp["bills"]:
            balance = float(b.get("total") or 0) - float(b.get("paid_amount") or 0)
            lines.append(f"    - {b.get('bill_number') or '-'}  due {b.get('due_date')}  ${balance:,.2f}")
    return "\n".join(lines)


def _render_payables_email_html(report: dict) -> str:
    def grp_table(grps, color="#B91C1C"):
        if not grps:
            return '<p style="color:#52525B; font-style:italic;">None</p>'
        rows = []
        for grp in grps:
            rows.append(f'<tr style="background:#F4F4F5;"><td colspan="3" style="padding:8px 12px; font-weight:bold; color:{color};">{grp["vendor_name"]}</td><td style="padding:8px 12px; text-align:right; font-weight:bold; font-family:monospace;">${grp["total"]:,.2f}</td></tr>')
            for b in grp["bills"]:
                balance = float(b.get("total") or 0) - float(b.get("paid_amount") or 0)
                rows.append(f'<tr><td style="padding:6px 12px; padding-left:24px; font-size:13px;">{b.get("bill_number") or "—"}</td><td style="padding:6px 12px; font-size:13px; color:#52525B;">{b.get("bill_date") or "—"}</td><td style="padding:6px 12px; font-size:13px; color:#52525B;">Due {b.get("due_date") or "—"}</td><td style="padding:6px 12px; text-align:right; font-family:monospace;">${balance:,.2f}</td></tr>')
        return f'<table style="width:100%; border-collapse:collapse; margin:8px 0;"><thead><tr style="border-bottom:1px solid #E4E4E7;"><th style="text-align:left; padding:6px 12px; font-size:10px; color:#52525B; text-transform:uppercase; letter-spacing:1px;">Bill #</th><th style="text-align:left; padding:6px 12px; font-size:10px; color:#52525B; text-transform:uppercase; letter-spacing:1px;">Date</th><th style="text-align:left; padding:6px 12px; font-size:10px; color:#52525B; text-transform:uppercase; letter-spacing:1px;">Due</th><th style="text-align:right; padding:6px 12px; font-size:10px; color:#52525B; text-transform:uppercase; letter-spacing:1px;">Balance</th></tr></thead><tbody>{"".join(rows)}</tbody></table>'

    return f"""<html><body style="font-family: Arial, Helvetica, sans-serif; color:#0A0A0A; max-width:720px; margin:0 auto; padding:20px;">
<div style="border-bottom:3px solid #062B67; padding-bottom:12px; margin-bottom:20px;">
  <div style="font-size:10px; font-weight:bold; letter-spacing:3px; color:#A0703A; text-transform:uppercase;">Weekly Payables</div>
  <h1 style="font-size:24px; margin:4px 0 0;">Bills to Pay — {report['today']}</h1>
</div>
<div style="display:flex; gap:16px; margin-bottom:20px;">
  <div style="flex:1; border:1px solid #E4E4E7; padding:16px;">
    <div style="font-size:10px; color:#52525B; text-transform:uppercase; letter-spacing:1px;">Overdue</div>
    <div style="font-size:24px; font-weight:bold; color:#B91C1C;">${report['overdue_total']:,.2f}</div>
    <div style="font-size:11px; color:#52525B;">{report['overdue_count']} bills</div>
  </div>
  <div style="flex:1; border:1px solid #E4E4E7; padding:16px;">
    <div style="font-size:10px; color:#52525B; text-transform:uppercase; letter-spacing:1px;">Due This Week</div>
    <div style="font-size:24px; font-weight:bold; color:#A0703A;">${report['due_this_week_total']:,.2f}</div>
    <div style="font-size:11px; color:#52525B;">{report['due_this_week_count']} bills</div>
  </div>
</div>
<h2 style="color:#B91C1C; font-size:14px; text-transform:uppercase; letter-spacing:2px; margin:20px 0 4px;">Overdue</h2>
{grp_table(report['overdue'], color='#B91C1C')}
<h2 style="color:#A0703A; font-size:14px; text-transform:uppercase; letter-spacing:2px; margin:24px 0 4px;">Due This Week (Next 7 Days)</h2>
{grp_table(report['due_this_week'], color='#A0703A')}
<p style="margin-top:24px; padding-top:16px; border-top:1px solid #E4E4E7; color:#52525B; font-size:11px;">
  SealTech Building Solutions  -  720-715-9955  -  finance@sealtechsolutions.co
</p>
</body></html>"""


# ----- Vendors -----
@api_router.get("/vendors", response_model=List[Vendor])
async def list_vendors(kind: Optional[str] = None, current=Depends(get_current_user)):
    query = {"is_deleted": {"$ne": True}}
    if kind:
        query["kind"] = kind
    cursor = db.vendors.find(query, {"_id": 0}).sort("name", 1)
    return await cursor.to_list(1000)


@api_router.get("/coi-roster")
async def coi_roster(current=Depends(get_current_user)):
    """Return Subcontractors whose GL or WC COI is missing, expired, or expiring within 30 days.

    Response shape per row:
      { id, name, email, gl_status, gl_expiry, wc_status, wc_expiry, worst_status }
    `worst_status` ∈ {"expired", "expiring", "missing"} — rows where both COIs
    are current ("ok") are excluded.
    """
    from datetime import datetime, date as _date
    today = _date.today()
    rows = []
    cursor = db.vendors.find({"kind": "Subcontractor", "is_deleted": {"$ne": True}}, {"_id": 0})
    async for v in cursor:
        def _one(on_file, exp):
            if not on_file:
                return "missing", None
            if not exp:
                return "missing", None
            try:
                d = datetime.strptime(exp, "%Y-%m-%d").date()
            except Exception:
                return "missing", None
            days = (d - today).days
            if days < 0:
                return "expired", d.isoformat()
            if days <= 30:
                return "expiring", d.isoformat()
            return "ok", d.isoformat()
        gl_status, gl_exp = _one(v.get("gl_coi_on_file"), v.get("gl_coi_expiry_date"))
        wc_status, wc_exp = _one(v.get("wc_coi_on_file"), v.get("wc_coi_expiry_date"))
        # Worst-of ranking: expired > expiring > missing > ok
        rank = {"expired": 3, "expiring": 2, "missing": 1, "ok": 0}
        worst = max([gl_status, wc_status], key=lambda x: rank.get(x, 0))
        if worst == "ok":
            continue  # exclude fully-current subs
        rows.append({
            "id": v["id"],
            "name": v.get("name", ""),
            "email": v.get("email", ""),
            "contact_name": v.get("contact_name", ""),
            "gl_status": gl_status,
            "gl_expiry": gl_exp,
            "wc_status": wc_status,
            "wc_expiry": wc_exp,
            "worst_status": worst,
        })
    rows.sort(key=lambda r: ({"expired": 0, "expiring": 1, "missing": 2}.get(r["worst_status"], 9), r["name"].lower()))
    return rows


class CoiRenewalEmailIn(BaseModel):
    to_override: Optional[str] = None  # optional override email
    cc: Optional[str] = None


@api_router.post("/coi-roster/{vendor_id}/email-renewal")
async def email_coi_renewal(vendor_id: str, body: CoiRenewalEmailIn, current=Depends(get_current_user)):
    """Send a one-click "Please send updated COI" email to the subcontractor."""
    v = await db.vendors.find_one({"id": vendor_id, "is_deleted": {"$ne": True}})
    if not v:
        raise HTTPException(404, "Subcontractor not found")
    to = (body.to_override or v.get("email") or "").strip()
    if not to:
        raise HTTPException(400, "Subcontractor has no email on file — provide to_override or add an email to the contact record")

    name = v.get("contact_name") or v.get("name") or "there"
    company = v.get("name") or ""
    # Compose human-readable missing/expired summary
    parts = []
    if v.get("gl_coi_on_file"):
        gx = v.get("gl_coi_expiry_date") or "no expiry on file"
        parts.append(f"• General Liability COI — expires {gx}")
    else:
        parts.append("• General Liability COI — not on file")
    if v.get("wc_coi_on_file"):
        wx = v.get("wc_coi_expiry_date") or "no expiry on file"
        parts.append(f"• Workers' Comp COI — expires {wx}")
    else:
        parts.append("• Workers' Comp COI — not on file")
    bullets = "\n".join(parts)

    subject = f"Action Needed: Updated Certificates of Insurance for {company}"
    body_text = (
        f"Hi {name},\n\n"
        f"Our records show the following insurance documents need to be updated for {company}:\n\n"
        f"{bullets}\n\n"
        "Please send us a current Certificate of Insurance at your earliest convenience so we can keep "
        "your account active for upcoming projects.\n\n"
        "Thanks,\n"
        "SealTech Commercial Roofing"
    )
    body_html = (
        f"<p>Hi {name},</p>"
        f"<p>Our records show the following insurance documents need to be updated for <b>{company}</b>:</p>"
        f"<ul>{''.join(f'<li>{p[2:]}</li>' for p in parts)}</ul>"
        "<p>Please send us a current Certificate of Insurance at your earliest convenience so we can keep "
        "your account active for upcoming projects.</p>"
        "<p>Thanks,<br/>SealTech Commercial Roofing</p>"
    )
    try:
        from email_sender import send_for_category
        result = await send_for_category(
            db, "projects",
            to=to, subject=subject, body_text=body_text, body_html=body_html, cc=body.cc)
        # Log a touchpoint on the vendor for visibility
        await db.vendors.update_one(
            {"id": vendor_id},
            {"$set": {"coi_last_renewal_email_at": datetime.now(timezone.utc).isoformat(),
                      "coi_last_renewal_email_by": current.get("id")}},
        )
        return {"ok": True, "to": to, "result": result}
    except Exception as e:
        raise HTTPException(500, f"Failed to send: {e}")




@api_router.post("/vendors", response_model=Vendor)
async def create_vendor(body: VendorIn, current=Depends(get_current_user)):
    data = body.model_dump()
    data["id"] = str(uuid.uuid4())
    data["created_at"] = now_iso()
    await db.vendors.insert_one(data.copy())
    return strip_id(data)


@api_router.get("/vendors/{vendor_id}", response_model=Vendor)
async def get_vendor(vendor_id: str, current=Depends(get_current_user)):
    doc = await db.vendors.find_one({"id": vendor_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return doc


@api_router.put("/vendors/{vendor_id}", response_model=Vendor)
async def update_vendor(vendor_id: str, body: VendorIn, current=Depends(get_current_user)):
    data = body.model_dump()
    result = await db.vendors.update_one({"id": vendor_id}, {"$set": data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vendor not found")
    doc = await db.vendors.find_one({"id": vendor_id}, {"_id": 0})
    return doc


@api_router.delete("/vendors/{vendor_id}")
async def delete_vendor(vendor_id: str, current=Depends(get_current_user)):
    result = await db.vendors.delete_one({"id": vendor_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return {"ok": True}


@api_router.get("/email-aliases")
async def email_aliases(current=Depends(get_current_user)):
    """List of allowed sender addresses (Gmail aliases) for invoice / statement / PO / scope / assessment emails.
    Returns:
      - aliases: list of whitelisted FROM addresses
      - default: legacy single default (GMAIL_FROM_EMAIL)
      - defaults: per-doc-type defaults so each modal can preselect the right alias
    """
    try:
        from email_sender import get_from_aliases
        aliases = get_from_aliases()
        env_default = (os.environ.get("GMAIL_FROM_EMAIL") or "").strip() or (aliases[0] if aliases else "")

        def _pick(preferred: str) -> str:
            return preferred if preferred in aliases else env_default

        defaults = {
            "invoice":    _pick("finance@sealtechsolutions.co"),
            "statement":  _pick("finance@sealtechsolutions.co"),
            "po":         _pick("projects@sealtechsolutions.co"),
            "scope":      _pick("scope@sealtechsolutions.co"),
            "assessment": _pick("assessments@sealtechsolutions.co"),
        }
        return {"aliases": aliases, "default": env_default, "defaults": defaults}
    except Exception as e:
        return {"aliases": [], "default": "", "defaults": {}, "error": str(e)}


# ----- Document Library -----
@api_router.get("/library/taxonomy")
async def library_taxonomy(current=Depends(get_current_user)):
    return {"taxonomy": LIBRARY_TAXONOMY}


def _valid_library_subcat(category: str, subcategory: str) -> bool:
    for cat in LIBRARY_TAXONOMY:
        if cat["category"] == category:
            return subcategory in cat["subcategories"]
    return False


@api_router.post("/library/files")
async def upload_library_file(
    file: UploadFile = File(...),
    category: str = Form(...),
    subcategory: str = Form(...),
    display_name: str = Form(""),
    description: str = Form(""),
    current=Depends(get_current_user),
):
    if not _valid_library_subcat(category, subcategory):
        raise HTTPException(status_code=400, detail="Invalid category/subcategory")
    ext = (file.filename.rsplit(".", 1)[-1] if "." in file.filename else "bin").lower()
    file_id = str(uuid.uuid4())
    safe_cat = category.replace(" ", "_").replace("&", "and")
    safe_sub = subcategory.replace(" ", "_").replace("&", "and").replace("/", "_")
    storage_path = f"{APP_NAME}/library/{safe_cat}/{safe_sub}/{file_id}.{ext}"
    data = await file.read()
    if len(data) > 50 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (max 50MB)")
    try:
        result = put_object(storage_path, data, file.content_type or "application/octet-stream")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Upload failed: {e}")
    doc = {
        "id": file_id,
        "category": category,
        "subcategory": subcategory,
        "display_name": (display_name.strip() or file.filename),
        "description": description.strip(),
        "storage_path": result["path"],
        "original_filename": file.filename,
        "content_type": file.content_type or "application/octet-stream",
        "size": len(data),
        "is_deleted": False,
        "uploaded_by": current["id"],
        "uploader_name": current.get("name", ""),
        "created_at": now_iso(),
    }
    await db.library_files.insert_one(doc.copy())
    doc.pop("_id", None)
    return doc


@api_router.get("/library/files")
async def list_library_files(
    category: Optional[str] = Query(None),
    subcategory: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    current=Depends(get_current_user),
):
    q = {"is_deleted": False}
    if category:
        q["category"] = category
    if subcategory:
        q["subcategory"] = subcategory
    if search:
        s = search.strip()
        q["$or"] = [
            {"display_name": {"$regex": s, "$options": "i"}},
            {"description": {"$regex": s, "$options": "i"}},
            {"original_filename": {"$regex": s, "$options": "i"}},
        ]
    rows = await db.library_files.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return rows


@api_router.put("/library/files/{file_id}")
async def update_library_file(file_id: str, body: dict = Body(...), current=Depends(get_current_user)):
    existing = await db.library_files.find_one({"id": file_id, "is_deleted": False}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="File not found")
    update = {}
    if "display_name" in body:
        update["display_name"] = (body["display_name"] or "").strip() or existing["display_name"]
    if "description" in body:
        update["description"] = (body["description"] or "").strip()
    if "category" in body and "subcategory" in body:
        if not _valid_library_subcat(body["category"], body["subcategory"]):
            raise HTTPException(status_code=400, detail="Invalid category/subcategory")
        update["category"] = body["category"]
        update["subcategory"] = body["subcategory"]
    update["updated_at"] = now_iso()
    await db.library_files.update_one({"id": file_id}, {"$set": update})
    return {**existing, **update}


@api_router.delete("/library/files/{file_id}")
async def delete_library_file(file_id: str, current=Depends(get_current_user)):
    await db.library_files.update_one({"id": file_id}, {"$set": {"is_deleted": True, "deleted_at": now_iso(), "deleted_by": current["id"]}})
    return {"ok": True}


@api_router.get("/library/files/{file_id}/download")
async def download_library_file(
    file_id: str,
    token: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    raw = None
    if authorization and authorization.startswith("Bearer "):
        raw = authorization[7:]
    elif token:
        raw = token
    if not raw:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(raw, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["sub"]})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid token")
    rec = await db.library_files.find_one({"id": file_id, "is_deleted": False})
    if not rec:
        raise HTTPException(status_code=404, detail="File not found")
    data, content_type = get_object(rec["storage_path"])
    return Response(
        content=data,
        media_type=rec.get("content_type") or content_type,
        headers={"Content-Disposition": f'attachment; filename="{rec["original_filename"]}"'},
    )


# ----- Email Scope (with optional Library attachments) -----
import scope_suggestions as _scope_sugg

@api_router.get("/deals/{deal_id}/scope-suggestions")
async def get_scope_suggestions(deal_id: str, current=Depends(get_current_user)):
    """Smart library doc picks for the Email Scope modal — based on proposed_roof_type
    plus any user-curated smart_tags on library files. Returns the file_ids to
    pre-check and a reasons map for the UI badge."""
    deal = await db.deals.find_one({"id": deal_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")
    files = await db.library_files.find({"is_deleted": {"$ne": True}}, {"_id": 0}).to_list(5000)
    return _scope_sugg.suggest_library_files(files, deal)


@api_router.post("/deals/{deal_id}/spec-sheet/email")
async def email_spec_sheet(deal_id: str, body: dict = Body(default={}), current=Depends(get_current_user)):
    """Email the scope PDF to the customer with optional Library file attachments.
    Body: { to_email, cc_email, from_email, library_file_ids: [str], message }

    FROM-address routing (when not explicitly overridden in body):
      - deal_type == "Assessment"  → assessments@sealtechsolutions.co
      - deal_type == "Scope"       → scope@sealtechsolutions.co
      - Fallback to GMAIL_FROM_EMAIL if either alias isn't whitelisted.
    """
    deal = await db.deals.find_one({"id": deal_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not deal:
        raise HTTPException(status_code=404, detail="Project not found")
    to_email = (body.get("to_email") or "").strip()
    cc_email = (body.get("cc_email") or "").strip()
    from_email = (body.get("from_email") or "").strip() or None
    custom_message = (body.get("message") or "").strip()

    # Auto-pick FROM by deal type when caller didn't pin one — use the
    # email-routing settings doc so it stays in sync with the Settings UI.
    is_assessment = (deal.get("deal_type") or "").lower() == "assessment"
    if not from_email:
        from email_routing import get_from_for_category
        from email_sender import get_from_aliases
        allowed = set(get_from_aliases())
        category = "assessments" if is_assessment else "scope"
        resolved = await get_from_for_category(db, category)
        from_email = resolved if (resolved and resolved in allowed) else (os.environ.get("GMAIL_FROM_EMAIL") or "").strip() or None

    if not to_email:
        # Auto-populate from primary contact if available
        cid = deal.get("primary_contact_id")
        if cid:
            c = await db.contacts.find_one({"id": cid}, {"_id": 0, "email": 1})
            if c and c.get("email"):
                to_email = c["email"]
    if not to_email:
        raise HTTPException(status_code=400, detail="No recipient email — please provide one.")

    # Build the spec-sheet PDF using the shared helper
    pdf_bytes = await _build_spec_pdf_for_deal(deal, current)

    # Compose attachment list — scope PDF first, then any library files picked by user
    project_label = (deal.get("title") or "scope").replace(" ", "_")
    attachments = [{"filename": f"{project_label}-scope.pdf", "data": pdf_bytes, "mime": "application/pdf"}]

    lib_ids = body.get("library_file_ids") or []
    if isinstance(lib_ids, list) and lib_ids:
        for fid in lib_ids:
            rec = await db.library_files.find_one({"id": fid, "is_deleted": False})
            if not rec:
                continue
            try:
                data, ct = get_object(rec["storage_path"])
                attachments.append({"filename": rec.get("original_filename") or "document.pdf", "data": data, "mime": rec.get("content_type") or ct or "application/octet-stream"})
            except Exception:
                continue

    # Auto-attach project cover photo(s). The caller can pass an explicit
    # `cover_photo_ids` list to override; otherwise we look up any project
    # photo with is_cover=True for this deal. This is SAFE — the Material
    # Take-Off PDF lives in `material_takeoffs`, not `project_photos`, so it
    # cannot be picked up here (locked rule: takeoff NEVER goes to customer).
    photo_ids = body.get("cover_photo_ids")
    if photo_ids is None:
        cover_photos = await db.project_photos.find(
            {"deal_id": deal_id, "is_cover": True, "is_deleted": {"$ne": True}},
            {"_id": 0},
        ).to_list(5)
    elif isinstance(photo_ids, list) and photo_ids:
        cover_photos = await db.project_photos.find(
            {"deal_id": deal_id, "id": {"$in": photo_ids}, "is_deleted": {"$ne": True}},
            {"_id": 0},
        ).to_list(len(photo_ids))
    else:
        cover_photos = []
    for ph in cover_photos:
        try:
            data, ct = get_object(ph["storage_path"])
            ext = (ph.get("content_type") or "image/jpeg").split("/")[-1].split("+")[0] or "jpg"
            fname = ph.get("original_filename") or f"{project_label}-cover.{ext}"
            attachments.append({
                "filename": fname,
                "data": data,
                "mime": ph.get("content_type") or ct or "image/jpeg",
            })
        except Exception as e:
            logger.warning(f"could not attach cover photo {ph.get('id')}: {e}")
            continue

    cust_company = ""
    cid = deal.get("primary_contact_id")
    if cid:
        c = await db.contacts.find_one({"id": cid}, {"_id": 0, "company_name": 1, "contact_name": 1})
        if c:
            cust_company = c.get("company_name") or c.get("contact_name") or ""

    subject = (
        f"Property Assessment — {deal.get('title') or 'Project'}"
        if is_assessment
        else f"Roofing Proposal — {deal.get('title') or 'Project'}"
    )
    intro = custom_message if custom_message else (
        f"Please find attached the property assessment for {deal.get('title') or 'your project'}. "
        f"The report details our findings, recommendations, and next steps."
        if is_assessment
        else
        f"Please find attached the roofing proposal for {deal.get('title') or 'your project'}. "
        f"The scope details our recommended system, pricing tiers, and standard inclusions/exclusions."
    )
    # Mint or reuse the public proposal-signing token + build the Sign Off URL.
    # The url is embedded in the email body so the recipient can accept inline
    # without needing a CRM login. Assessment emails do not include this CTA.
    sign_off_url = ""
    if not is_assessment:
        try:
            token_val = await _proposal_signing.ensure_proposal_token(db, deal_id)
            if _PUBLIC_BASE_URL and token_val:
                sign_off_url = f"{_PUBLIC_BASE_URL.rstrip('/')}/sign/{token_val}"
        except Exception as e:
            logger.warning(f"could not mint proposal_sign_token for deal={deal_id}: {e}")
            sign_off_url = ""

    body_text = (
        f"Hello {(cust_company or '')},\n\n"
        f"{intro}\n\n"
        f"  Attachments: {len(attachments)} file{'s' if len(attachments) > 1 else ''}\n\n"
        + (
            f"Ready to move forward? Sign off here: {sign_off_url}\n\n"
            if sign_off_url else ""
        )
        + f"If you have any questions, please reply to this email or call us at 720-715-9955.\n\n"
        f"Thank you,\n"
        f"SealTech Building Solutions\n"
        f"720-715-9955  ·  scope@sealtechsolutions.co"
    )
    sign_off_html = (
        f'<p style="margin: 20px 0;"><a href="{sign_off_url}" style="background:#062B67;color:white;'
        f'padding:12px 24px;text-decoration:none;font-weight:bold;letter-spacing:1.5px;font-size:13px;'
        f'text-transform:uppercase;border-radius:2px;display:inline-block;">Review &amp; Sign Off</a></p>'
        f'<p style="margin: 0 0 16px; color:#52525B; font-size:12px;">'
        f'Or paste this link into your browser: <span style="word-break:break-all;color:#062B67;">{sign_off_url}</span>'
        f'</p>'
    ) if sign_off_url else ""
    body_html = f"""
    <html><body style="font-family: Arial, Helvetica, sans-serif; color: #0A0A0A; max-width: 620px;">
      <p style="margin: 0 0 16px;">Hello {cust_company or ''},</p>
      <p style="margin: 0 0 16px;">{intro.replace(chr(10), '<br/>')}</p>
      <p style="margin: 0 0 16px; color: #52525B; font-size: 13px;"><b>Attachments:</b> {len(attachments)} file{'s' if len(attachments) > 1 else ''}</p>
      {sign_off_html}
      <p style="margin: 16px 0;">If you have any questions, please reply to this email or call us at 720-715-9955.</p>
      <p style="margin: 24px 0 0; padding-top: 16px; border-top: 1px solid #E4E4E7; color: #52525B; font-size: 12px;">
        <b style="color: #0A0A0A;">SealTech Building Solutions</b><br/>
        720-715-9955  ·  scope@sealtechsolutions.co
      </p>
    </body></html>
    """

    try:
        from email_sender import send_email, EmailNotConfigured
        result = send_email(
            to=to_email,
            cc=cc_email,
            subject=subject,
            body_text=body_text,
            body_html=body_html,
            reply_to=os.environ.get("GMAIL_FROM_EMAIL") or None,
            attachments=attachments,
            from_email=from_email,
        )
    except EmailNotConfigured as e:
        raise HTTPException(status_code=500, detail=str(e))
    except smtplib.SMTPAuthenticationError as e:
        raise HTTPException(status_code=500, detail=f"Gmail authentication failed. ({e.smtp_code})")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Email send failed: {type(e).__name__}: {e}")

    # Stamp the deal so the "Scope Sent" pipeline dot turns green and the
    # Next-Step card moves forward. (Bug: previously the email fired but the
    # pipeline visualization stayed stuck at "Email the scope".)
    sent_at = now_iso()
    history_label = "Assessment emailed" if is_assessment else "Scope emailed"

    # Persist the exact PDF that went out to Object Storage so reps can re-open
    # it from the Activity Timeline ("Open the PDF that went out" link). The
    # raw `pdf_bytes` was generated a few lines above for the attachment.
    sent_pdf_file_id = ""
    try:
        sent_pdf_file_id = str(uuid.uuid4())
        kind_slug = "assessment-pdf" if is_assessment else "scope-pdf"
        sp = f"{APP_NAME}/uploads/deal/{deal_id}/{kind_slug}-{sent_pdf_file_id}.pdf"
        put_object(sp, pdf_bytes, "application/pdf")
        await db.files.insert_one({
            "id": sent_pdf_file_id,
            "parent_type": "deal",
            "parent_id": deal_id,
            "category": "Assessment" if is_assessment else "Scope",
            "storage_path": sp,
            "original_filename": f"{project_label}-{'assessment' if is_assessment else 'scope'}-{sent_at[:10]}.pdf",
            "content_type": "application/pdf",
            "size": len(pdf_bytes),
            "is_deleted": False,
            "uploaded_by": current["id"],
            "created_at": sent_at,
            "is_sent_snapshot": True,  # marks this as a frozen send-snapshot, not user-uploaded
        })
    except Exception as e:
        logger.warning(f"sent-PDF snapshot stash failed for deal={deal_id}: {e}")
        sent_pdf_file_id = ""

    history_entry = {
        "at": sent_at,
        "user_id": current.get("id", ""),
        "user_name": current.get("name", ""),
        "label": history_label,
        "to": to_email,
        "attachments_count": len(attachments),
        "pdf_file_id": sent_pdf_file_id,
    }
    try:
        await db.deals.update_one(
            {"id": deal_id},
            {
                "$set": {
                    "last_scope_sent_at": sent_at,
                    "last_scope_sent_to": to_email,
                },
                "$inc": {"scope_send_count": 1},
                "$push": {"status_history": history_entry},
            },
        )
    except Exception as e:
        # Email already went out — don't blow up the response on a bookkeeping
        # failure, just warn so it shows up in logs.
        logger.warning(f"deal stamp (scope-sent) failed for {deal_id}: {type(e).__name__}: {e}")

    return {
        "ok": True,
        "message": f"Scope emailed to {to_email}" + (f" (cc: {cc_email})" if cc_email else "") + f" with {len(attachments)} attachment{'s' if len(attachments) > 1 else ''}",
        "to_email": to_email,
        "cc_email": cc_email,
        "from_email": from_email,
        "attachments_count": len(attachments),
        "message_id": result.get("message_id"),
        "last_scope_sent_at": sent_at,
    }


# ----- Subcontractor Job Logs & Scorecards -----
SUB_JOB_STATUSES = ["Scheduled", "In Progress", "Completed", "Cancelled"]


class SubJobLogIn(BaseModel):
    """One logged job assignment for a subcontractor. Drives the scorecard metrics
    (on-time %, average quality, $ awarded)."""
    model_config = ConfigDict(extra="ignore")
    subcontractor_id: str
    deal_id: Optional[str] = None
    work_description: str = ""
    scheduled_date: Optional[str] = None  # yyyy-mm-dd
    completed_date: Optional[str] = None  # yyyy-mm-dd, set when status moves to Completed
    status: str = "Scheduled"
    quality_rating: Optional[int] = None  # 1-5, nullable until rated
    issues_count: int = 0  # punch-list / callback issues raised on this job
    contract_amount: float = 0.0
    notes: str = ""


def _normalize_sub_job(doc: dict) -> dict:
    if doc.get("status") not in SUB_JOB_STATUSES:
        doc["status"] = "Scheduled"
    qr = doc.get("quality_rating")
    if qr is not None:
        try:
            qr = int(qr)
            doc["quality_rating"] = max(1, min(5, qr))
        except (TypeError, ValueError):
            doc["quality_rating"] = None
    try:
        doc["issues_count"] = max(0, int(doc.get("issues_count") or 0))
    except (TypeError, ValueError):
        doc["issues_count"] = 0
    try:
        doc["contract_amount"] = float(doc.get("contract_amount") or 0)
    except (TypeError, ValueError):
        doc["contract_amount"] = 0.0
    # If marked Completed without a completed_date, stamp it today
    if doc.get("status") == "Completed" and not doc.get("completed_date"):
        doc["completed_date"] = datetime.now(timezone.utc).date().isoformat()
    # Derived on_time flag (only meaningful once completed)
    sched = doc.get("scheduled_date") or ""
    done = doc.get("completed_date") or ""
    on_time = None
    if doc.get("status") == "Completed" and sched and done:
        try:
            on_time = datetime.strptime(done[:10], "%Y-%m-%d").date() <= datetime.strptime(sched[:10], "%Y-%m-%d").date()
        except ValueError:
            on_time = None
    doc["on_time"] = on_time
    return doc


@api_router.get("/sub-jobs")
async def list_sub_jobs(
    subcontractor_id: Optional[str] = Query(None),
    deal_id: Optional[str] = Query(None),
    status_filter: Optional[str] = Query(None, alias="status"),
    current=Depends(get_current_user),
):
    q = {"is_deleted": {"$ne": True}}
    if subcontractor_id:
        q["subcontractor_id"] = subcontractor_id
    if deal_id:
        q["deal_id"] = deal_id
    if status_filter:
        q["status"] = status_filter
    rows = await db.sub_job_logs.find(q, {"_id": 0}).sort("scheduled_date", -1).to_list(1000)
    return rows


@api_router.post("/sub-jobs")
async def create_sub_job(body: SubJobLogIn, current=Depends(get_current_user)):
    sub = await db.vendors.find_one({"id": body.subcontractor_id, "kind": "Subcontractor"}, {"_id": 0})
    if not sub:
        raise HTTPException(status_code=404, detail="Subcontractor not found")
    deal_title = ""
    if body.deal_id:
        d = await db.deals.find_one({"id": body.deal_id, "is_deleted": {"$ne": True}}, {"_id": 0, "title": 1})
        if d:
            deal_title = d.get("title", "")
    doc = body.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["subcontractor_name"] = sub.get("name", "")
    doc["deal_title"] = deal_title
    doc["created_at"] = now_iso()
    doc["created_by"] = current["id"]
    doc = _normalize_sub_job(doc)
    await db.sub_job_logs.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.put("/sub-jobs/{job_id}")
async def update_sub_job(job_id: str, body: SubJobLogIn, current=Depends(get_current_user)):
    existing = await db.sub_job_logs.find_one({"id": job_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Sub job not found")
    updated = {**existing, **body.model_dump(exclude_none=False)}
    # Re-hydrate denormalized fields if subcontractor_id changed
    if body.subcontractor_id != existing.get("subcontractor_id"):
        sub = await db.vendors.find_one({"id": body.subcontractor_id, "kind": "Subcontractor"}, {"_id": 0})
        updated["subcontractor_name"] = sub.get("name", "") if sub else existing.get("subcontractor_name", "")
    if body.deal_id and body.deal_id != existing.get("deal_id"):
        d = await db.deals.find_one({"id": body.deal_id}, {"_id": 0, "title": 1})
        updated["deal_title"] = d.get("title", "") if d else ""
    updated["updated_at"] = now_iso()
    updated = _normalize_sub_job(updated)
    updated.pop("_id", None)
    await db.sub_job_logs.update_one({"id": job_id}, {"$set": updated})
    return updated


@api_router.delete("/sub-jobs/{job_id}")
async def delete_sub_job(job_id: str, current=Depends(get_current_user)):
    if is_admin(current):
        await db.sub_job_logs.delete_one({"id": job_id})
    else:
        await db.sub_job_logs.update_one({"id": job_id}, {"$set": {"is_deleted": True, "deleted_at": now_iso(), "deleted_by": current["id"]}})
    return {"ok": True}


def _scorecard_rating(metrics: dict) -> str:
    """Roll-up grade based on on-time % and average quality. Friendly badge label."""
    if metrics["completed_jobs"] == 0:
        return "No History"
    on_time_pct = metrics["on_time_pct"]
    avg_q = metrics["avg_quality"] or 0
    if on_time_pct >= 90 and avg_q >= 4.5:
        return "A+ — Top Performer"
    if on_time_pct >= 80 and avg_q >= 4.0:
        return "A — Strong"
    if on_time_pct >= 70 and avg_q >= 3.5:
        return "B — Solid"
    if on_time_pct >= 60 or avg_q >= 3.0:
        return "C — Needs Review"
    return "D — Caution"


@api_router.get("/subcontractor-scorecards")
async def subcontractor_scorecards(current=Depends(get_current_user)):
    """Aggregated scorecard for every subcontractor. Includes only Completed jobs in
    quality/on-time stats; total awarded counts all non-cancelled jobs."""
    # Pull all subs first so we report zeros for subs with no jobs yet
    subs = await db.vendors.find({"kind": "Subcontractor"}, {"_id": 0}).sort("name", 1).to_list(2000)
    # Pull every non-deleted job
    jobs = await db.sub_job_logs.find({"is_deleted": {"$ne": True}}, {"_id": 0}).to_list(20000)
    by_sub = {s["id"]: [] for s in subs}
    for j in jobs:
        sid = j.get("subcontractor_id")
        if sid in by_sub:
            by_sub[sid].append(j)
        else:
            by_sub.setdefault(sid, []).append(j)

    out = []
    for sub in subs:
        rows = by_sub.get(sub["id"], [])
        completed = [r for r in rows if r.get("status") == "Completed"]
        scheduled = [r for r in rows if r.get("status") == "Scheduled"]
        in_progress = [r for r in rows if r.get("status") == "In Progress"]
        cancelled = [r for r in rows if r.get("status") == "Cancelled"]
        completed_with_dates = [r for r in completed if r.get("on_time") is not None]
        on_time_count = sum(1 for r in completed_with_dates if r.get("on_time") is True)
        on_time_pct = round((on_time_count / len(completed_with_dates)) * 100, 1) if completed_with_dates else 0.0
        rated = [r for r in completed if r.get("quality_rating") is not None]
        avg_q = round(sum(r["quality_rating"] for r in rated) / len(rated), 2) if rated else 0.0
        total_awarded = round(sum(float(r.get("contract_amount") or 0) for r in rows if r.get("status") != "Cancelled"), 2)
        issues_total = sum(int(r.get("issues_count") or 0) for r in rows)
        # Last completed job date
        last_done = None
        for r in completed:
            d = r.get("completed_date") or r.get("scheduled_date") or ""
            if d and (last_done is None or d > last_done):
                last_done = d
        metrics = {
            "subcontractor_id": sub["id"],
            "subcontractor_name": sub.get("name", ""),
            "category": sub.get("category", ""),
            "total_jobs": len(rows),
            "completed_jobs": len(completed),
            "scheduled_jobs": len(scheduled),
            "in_progress_jobs": len(in_progress),
            "cancelled_jobs": len(cancelled),
            "on_time_pct": on_time_pct,
            "on_time_count": on_time_count,
            "rated_jobs": len(rated),
            "avg_quality": avg_q,
            "total_awarded": total_awarded,
            "issues_total": issues_total,
            "last_completed": last_done,
        }
        metrics["grade"] = _scorecard_rating(metrics)
        out.append(metrics)
    # Sort: graded subs first (by on_time desc, then avg_q desc), then no-history at bottom
    out.sort(key=lambda m: (m["completed_jobs"] == 0, -m["on_time_pct"], -(m["avg_quality"] or 0), m["subcontractor_name"]))
    return out


# ----- Spec Sheet -----
@api_router.get("/deals/{deal_id}/sign-link")
async def deal_sign_link(deal_id: str, current: dict = Depends(get_current_user)):
    """Mint (or return existing) the public proposal-signing URL for a deal.

    Powers the calculator's "Get Signed" button — Darren clicks once and gets
    the URL to either copy to clipboard or open on the tablet at the table.
    Idempotent: re-calling returns the same URL until the deal is signed.
    """
    deal = await db.deals.find_one({"id": deal_id, "is_deleted": {"$ne": True}},
                                   {"_id": 0, "id": 1, "title": 1,
                                    "scope_signed_at": 1, "proposal_sign_token": 1})
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")
    token_val = await _proposal_signing.ensure_proposal_token(db, deal_id)
    base = _PUBLIC_BASE_URL.rstrip("/") if _PUBLIC_BASE_URL else ""
    if not base:
        raise HTTPException(status_code=500,
                            detail="PUBLIC_BASE_URL not configured")
    return {
        "deal_id": deal_id,
        "title": deal.get("title") or "",
        "token": token_val,
        "sign_url": f"{base}/sign/{token_val}",
        "already_signed": bool(deal.get("scope_signed_at")),
    }


@api_router.get("/deals/{deal_id}/spec-sheet.pdf")
async def deal_spec_sheet(
    deal_id: str,
    token: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    # Auth (Bearer header OR ?token= for browser links)
    raw = None
    if authorization and authorization.startswith("Bearer "):
        raw = authorization[7:]
    elif token:
        raw = token
    if not raw:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(raw, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["sub"]})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

    deal = await db.deals.find_one({"id": deal_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not deal:
        raise HTTPException(status_code=404, detail="Project not found")

    pdf_bytes = await _build_spec_pdf_for_deal(deal, user)
    filename = f"sealtech-scope-{(deal.get('title') or 'project')}.pdf".replace(" ", "_")
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


async def _build_spec_pdf_for_deal(deal: dict, user: dict) -> bytes:
    """Shared PDF builder used by /spec-sheet.pdf (download) and
    /spec-sheet/email (send-with-attachments)."""

    # Build address from linked property
    project_address = "—"
    if deal.get("property_id"):
        prop = await db.properties.find_one({"id": deal["property_id"]}, {"_id": 0})
        if prop:
            parts = [prop.get("property_address", ""), prop.get("property_address_line2", "")]
            line1 = " ".join([p for p in parts if p]).strip()
            city = prop.get("property_city", "")
            st = prop.get("property_state", "")
            zp = prop.get("property_zip", "")
            line2 = ", ".join([p for p in [city, st] if p]).strip()
            if zp:
                line2 = f"{line2} {zp}".strip()
            project_address = "  ·  ".join([p for p in [line1, line2] if p]).strip() or "—"

    # Fetch customer contact (name + best phone) for personalization
    contact_name = ""
    contact_phone = ""
    customer_id = deal.get("customer_contact_id") or deal.get("contact_id")
    if customer_id:
        cust = await db.contacts.find_one({"id": customer_id}, {"_id": 0})
        if cust:
            contact_name = cust.get("contact_name", "") or ""
            contact_phone = (cust.get("mobile_phone") or cust.get("phone") or cust.get("work_phone") or "").strip()

    # Per-roof-type product type defaults (user-curated wording for each variant)
    PRODUCT_TYPE_DEFAULTS = {
        "TPO Over-Lay": "TPO Roof System Over Existing TPO Over-Lay",
        "TPO Replacement": "TPO Roof System Replacing TPO",
        "EPDM Over-Lay": "EPDM Roof System Over Existing EPDM Over-Lay",
        "EPDM Replacement": "EPDM Roof System Replacing EPDM",
        "ModBit Over-Lay": "Modified Bitumen Roof System Over Existing Modified Bitumen Over-Lay",
        "ModBit Replacement": "Modified Bitumen Roof System Replacing Modified Bitumen",
        "PVC Over-Lay": "PVC Roof System Over Existing PVC Over-Lay",
        "PVC Replacement": "PVC Roof System Replacing PVC",
        "FARM (Fluid Applied Reinforced Membrane)": (
            f"Fluid Applied Reinforced Membrane Roof System Over Existing {deal.get('current_roof_type','').strip() or '—'}"
        ),
    }
    # Membrane-system labels we recognize for new construction Product Type phrasing
    NEW_CONSTRUCTION_LABELS = {
        "TPO": "TPO",
        "TPO Over-Lay": "TPO",
        "TPO Replacement": "TPO",
        "EPDM": "EPDM",
        "EPDM Over-Lay": "EPDM",
        "EPDM Replacement": "EPDM",
        "EPDM w/ Ballast": "EPDM",
        "PVC": "PVC",
        "PVC Over-Lay": "PVC",
        "PVC Replacement": "PVC",
        "ModBit": "Modified Bitumen",
        "ModBit Over-Lay": "Modified Bitumen",
        "ModBit Replacement": "Modified Bitumen",
        "BUR (Built-Up)": "Built-Up Roof",
    }
    proposed = deal.get("proposed_roof_type") or "Silicone"
    current = deal.get("current_roof_type") or ""
    is_new = (
        current.strip().lower().startswith("none")
        or "new construction" in current.lower()
    )
    if deal.get("product_description"):
        product_desc = deal["product_description"]
    elif proposed in ("Construction Project", "Other") or current.lower() == "other construction work":
        # Non-roofing scope — keep the PDF header generic and let the free-form scope tell the story
        product_desc = "Construction Project — Custom Scope"
    elif is_new and proposed in NEW_CONSTRUCTION_LABELS:
        product_desc = f"{NEW_CONSTRUCTION_LABELS[proposed]} Roof System on New Construction"
    elif proposed in PRODUCT_TYPE_DEFAULTS:
        product_desc = PRODUCT_TYPE_DEFAULTS[proposed]
    else:
        product_desc = f"{proposed} Roof System Over Existing {current}".strip()
    color = deal.get("warranty_color") or "white"

    data = {
        "contact_name": contact_name,
        "contact_phone": contact_phone,
        "project_address": project_address,
        "product_type": product_desc,
        "date": datetime.now(timezone.utc).strftime("%m/%d/%Y"),
        "opt_20": float(deal.get("proposal_option_1") or 0),
        "opt_15": float(deal.get("proposal_option_2") or 0),
        "opt_10": float(deal.get("proposal_option_3") or 0),
        "opt_25": float(deal.get("proposal_option_25yr") or 0),
        "w20": float(deal.get("warranty_20yr_add") or 0),
        "w15": float(deal.get("warranty_15yr_add") or 0),
        "w10": float(deal.get("warranty_10yr_add") or 0),
        "w25": float(deal.get("warranty_25yr_add") or 0),
        "total_sqft": float(deal.get("total_sqft") or 0),
        "color": color,
        "roof_type_label": (deal.get("proposed_roof_type") or "silicone").lower(),
        "custom_scope": (deal.get("custom_scope") or "").strip(),
        "construction_project_requirements": (deal.get("construction_project_requirements") or "").strip(),
        "construction_other_requirements": (deal.get("construction_other_requirements") or "").strip(),
        "construction_exclusions": (deal.get("construction_exclusions") or "").strip(),
        "construction_scope_subtitle": (deal.get("construction_scope_subtitle") or "").strip(),
        "project_type_override": (deal.get("project_type_override") or "").strip(),
        # Per-deal bullet overrides — surface to build_spec_sheet so the editor
        # changes win over the template defaults on this project's PDF only.
        "scope_overrides": deal.get("scope_overrides") or {},
        # Free-form Custom Add-Ons typed in the Calculator (e.g. Metal Flashing
        # $650). Rendered as additional Inclusions bullets so the customer
        # sees what items are baked into the totals.
        "calc_custom_addons": deal.get("calc_custom_addons") or [],
    }

    # Fetch cover photo if set
    photo_bytes = None
    cover_id = deal.get("cover_photo_file_id")
    if cover_id:
        rec = await db.files.find_one({"id": cover_id, "is_deleted": False})
        if rec:
            try:
                photo_bytes, _ = get_object(rec["storage_path"])
            except Exception:
                photo_bytes = None
    # Fallback: if no `cover_photo_file_id` is set on the deal, use whatever
    # project photo is flagged `is_cover=True`. This keeps the in-PDF banner
    # in sync with the gallery's "Set as Cover" star, even when the deal's
    # dedicated field gets cleared during data recovery / album reshuffles.
    if not photo_bytes:
        flagged = await db.project_photos.find_one(
            {"deal_id": deal["id"], "is_cover": True, "is_deleted": {"$ne": True}},
            {"_id": 0, "storage_path": 1},
        )
        if flagged and flagged.get("storage_path"):
            try:
                photo_bytes, _ = get_object(flagged["storage_path"])
            except Exception:
                photo_bytes = None

    pdf_bytes = build_spec_sheet(
        data,
        cover_photo_bytes=photo_bytes,
        roof_type=deal.get("proposed_roof_type"),
        current_roof_type=deal.get("current_roof_type"),
        signer_name=(user.get("name") or "").strip(),
        signer_credentials=(user.get("credentials") or "").strip(),
    )
    return pdf_bytes


# ----- Scope Editor (P2): per-deal bullet overrides for the spec-sheet PDF -----
@api_router.get("/deals/{deal_id}/scope-bullets")
async def get_scope_bullets(deal_id: str, current=Depends(get_current_user)):
    """Return the effective scope bullets for this deal so the Scope Editor
    has something to pre-populate. Merges the template defaults with any
    per-deal `scope_overrides` and flags which keys have been overridden."""
    deal = await db.deals.find_one({"id": deal_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not deal:
        raise HTTPException(404, "Deal not found")
    if current.get("role") == "sales":
        owns = deal.get("assigned_to_user_id") == current["id"] or deal.get("created_by_user_id") == current["id"]
        if not owns:
            raise HTTPException(403, "Not your project")

    from spec_sheet import _resolve_template, _apply_scope_overrides
    base = _resolve_template(
        deal.get("proposed_roof_type"),
        deal.get("current_roof_type"),
    )
    overrides = deal.get("scope_overrides") or {}
    effective = _apply_scope_overrides(base, overrides)

    def _val(field, list_field=False):
        return effective.get(field, [] if list_field else "")

    return {
        "deal_id": deal_id,
        "roof_type": deal.get("proposed_roof_type") or "",
        "is_new_construction": (deal.get("current_roof_type") or "").lower().startswith("none") or "new construction" in (deal.get("current_roof_type") or "").lower(),
        "template_title": base.get("title", ""),
        "is_dynamic_scope": bool(base.get("dynamic_scope")),
        "effective": {
            "title": _val("title"),
            "scope_1_title": _val("scope_1_title"),
            "scope_1": _val("scope_1", list_field=True),
            "scope_2_title": _val("scope_2_title"),
            "scope_2": _val("scope_2", list_field=True),
            "key_advantages": _val("key_advantages", list_field=True),
        },
        "defaults": {
            "title": base.get("title", ""),
            "scope_1_title": base.get("scope_1_title", ""),
            "scope_1": base.get("scope_1", []),
            "scope_2_title": base.get("scope_2_title", ""),
            "scope_2": base.get("scope_2", []),
            "key_advantages": base.get("key_advantages", []),
        },
        "overrides": overrides,
        "overridden_keys": [k for k in ("title", "scope_1_title", "scope_1", "scope_2_title", "scope_2", "key_advantages") if k in overrides and overrides[k]],
    }


@api_router.put("/deals/{deal_id}/scope-bullets")
async def put_scope_bullets(deal_id: str, body: dict = Body(...), current=Depends(get_current_user)):
    """Save (or clear) per-deal scope overrides. Pass `null` or omit a key to
    revert that field to the template default; pass an empty list/string to
    revert. Returns the same shape as the GET so the UI can refresh in one round trip.
    """
    deal = await db.deals.find_one({"id": deal_id, "is_deleted": {"$ne": True}}, {"_id": 0, "id": 1, "assigned_to_user_id": 1, "created_by_user_id": 1})
    if not deal:
        raise HTTPException(404, "Deal not found")
    if current.get("role") == "sales":
        owns = deal.get("assigned_to_user_id") == current["id"] or deal.get("created_by_user_id") == current["id"]
        if not owns:
            raise HTTPException(403, "Not your project")

    overrides: dict = {}
    for key in ("title", "scope_1_title", "scope_2_title"):
        v = body.get(key)
        if isinstance(v, str) and v.strip():
            overrides[key] = v.strip()
    for key in ("scope_1", "scope_2", "key_advantages"):
        v = body.get(key)
        if isinstance(v, list):
            cleaned = [str(x).strip() for x in v if str(x).strip()]
            if cleaned:
                overrides[key] = cleaned
    await db.deals.update_one(
        {"id": deal_id},
        {"$set": {"scope_overrides": overrides, "updated_at": now_iso()}},
    )
    return await get_scope_bullets(deal_id, current)  # type: ignore[arg-type]


# ----- Files (Documents) -----
@api_router.post("/files/upload")
async def upload_file(
    file: UploadFile = File(...),
    parent_type: str = Form(...),
    parent_id: str = Form(...),
    category: str = Form("Other"),
    current=Depends(get_current_user),
):
    if parent_type not in PARENT_TYPES:
        raise HTTPException(status_code=400, detail="Invalid parent_type")
    if category not in DOCUMENT_CATEGORIES:
        category = "Other"
    ext = (file.filename.rsplit(".", 1)[-1] if "." in file.filename else "bin").lower()
    file_id = str(uuid.uuid4())
    storage_path = f"{APP_NAME}/uploads/{parent_type}/{parent_id}/{file_id}.{ext}"
    data = await file.read()
    if len(data) > 50 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (max 50MB)")
    try:
        result = put_object(storage_path, data, file.content_type or "application/octet-stream")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Upload failed: {e}")
    doc = {
        "id": file_id,
        "parent_type": parent_type,
        "parent_id": parent_id,
        "category": category,
        "storage_path": result["path"],
        "original_filename": file.filename,
        "content_type": file.content_type or "application/octet-stream",
        "size": len(data),
        "is_deleted": False,
        "uploaded_by": current["id"],
        "created_at": now_iso(),
    }
    await db.files.insert_one(doc.copy())
    doc.pop("_id", None)
    return doc


@api_router.get("/files")
async def list_files(
    parent_type: str = Query(...),
    parent_id: str = Query(...),
    current=Depends(get_current_user),
):
    cursor = db.files.find(
        {"parent_type": parent_type, "parent_id": parent_id, "is_deleted": False},
        {"_id": 0},
    ).sort("created_at", -1)
    return await cursor.to_list(1000)


@api_router.get("/files/{file_id}/download")
async def download_file(
    file_id: str,
    token: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    # Auth check - support either Bearer header or token query (for browser links)
    raw = None
    if authorization and authorization.startswith("Bearer "):
        raw = authorization[7:]
    elif token:
        raw = token
    if not raw:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(raw, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["sub"]})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid token")
    rec = await db.files.find_one({"id": file_id, "is_deleted": False})
    if not rec:
        raise HTTPException(status_code=404, detail="File not found")
    try:
        data, content_type = get_object(rec["storage_path"])
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Download failed: {e}")
    return Response(
        content=data,
        media_type=rec.get("content_type") or content_type,
        headers={"Content-Disposition": f'attachment; filename="{rec["original_filename"]}"'},
    )


@api_router.delete("/files/{file_id}")
async def delete_file(file_id: str, current=Depends(get_current_user)):
    result = await db.files.update_one({"id": file_id}, {"$set": {"is_deleted": True, "deleted_at": now_iso()}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="File not found")
    return {"ok": True}


# ----- Export -----
async def _records_for(category: str) -> list:
    if category == "contacts":
        return await db.contacts.find({}, {"_id": 0}).sort("contact_name", 1).to_list(5000)
    if category == "properties":
        return await db.properties.find({}, {"_id": 0}).sort("property_name", 1).to_list(5000)
    if category == "projects":
        return await db.deals.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    if category == "vendors":
        return await db.vendors.find({"kind": "Vendor"}, {"_id": 0}).sort("name", 1).to_list(5000)
    if category == "subcontractors":
        return await db.vendors.find({"kind": "Subcontractor"}, {"_id": 0}).sort("name", 1).to_list(5000)
    raise HTTPException(status_code=400, detail="Unknown category")


@api_router.get("/export/{category}.{fmt}")
async def export_category(category: str, fmt: str, current=Depends(get_current_user)):
    if category == "all":
        sections = [(c, await _records_for(c)) for c in IMPORT_CATEGORIES]
    else:
        if category not in IMPORT_CATEGORIES:
            raise HTTPException(status_code=400, detail="Unknown category")
        sections = [(category, await _records_for(category))]
    if fmt == "xlsx":
        data = to_excel(sections)
        return Response(
            content=data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="sealtech-{category}.xlsx"'},
        )
    if fmt == "pdf":
        data = to_pdf(sections)
        return Response(
            content=data,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="sealtech-{category}.pdf"'},
        )
    raise HTTPException(status_code=400, detail="Unsupported format (use xlsx or pdf)")


@api_router.get("/export/template/{category}.xlsx")
async def export_template(category: str, current=Depends(get_current_user)):
    if category not in IMPORT_CATEGORIES:
        raise HTTPException(status_code=400, detail="Unknown category")
    sections = [(category, [])]
    data = to_excel(sections)
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="sealtech-{category}-template.xlsx"'},
    )


@api_router.get("/maintenance/export.{fmt}")
async def export_maintenance(fmt: str, current=Depends(get_current_user)):
    """Export the maintenance customer list to Excel or PDF."""
    rows = await list_maintenance(current)
    headers = ["Customer", "Phone", "Property", "Property Address", "Annual Rate", "Start Date", "Last Visit", "Next Due", "Status", "Visits"]
    data_rows = []
    for r in rows:
        data_rows.append([
            r.get("contact_name", ""),
            r.get("contact_phone", ""),
            r.get("property_name", ""),
            r.get("property_address", ""),
            float(r.get("maintenance_rate", 0) or 0),
            r.get("maintenance_start_date", ""),
            r.get("last_maintenance_date", ""),
            r.get("next_maintenance_date", ""),
            r.get("status", ""),
            r.get("visit_count", 0),
        ])
    if fmt == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "Maintenance"
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill("solid", fgColor="062B67")
            cell.alignment = Alignment(horizontal="left", vertical="center")
        for row in data_rows:
            ws.append(row)
        for col_idx, hdr in enumerate(headers, 1):
            max_len = len(str(hdr))
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=1, values_only=True):
                for v in cell:
                    max_len = max(max_len, min(len(str(v) if v is not None else ""), 40))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 2
        ws.freeze_panes = "A2"
        buf = BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="sealtech-maintenance.xlsx"'},
        )
    if fmt == "pdf":
        from reportlab.lib import colors as rl_colors
        from reportlab.lib.pagesizes import letter as rl_letter, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle as RLPS
        from reportlab.lib.units import inch as rl_inch
        from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle as RLTableStyle, Paragraph as RLP, Spacer as RLSpacer
        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(rl_letter), leftMargin=0.4 * rl_inch, rightMargin=0.4 * rl_inch, topMargin=0.5 * rl_inch, bottomMargin=0.5 * rl_inch)
        styles = getSampleStyleSheet()
        title_style = RLPS("title", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=18, textColor=rl_colors.HexColor("#0A0A0A"))
        eyebrow = RLPS("eyebrow", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=8, textColor=rl_colors.HexColor("#062B67"), leading=10)
        body = RLPS("body", parent=styles["Normal"], fontName="Helvetica", fontSize=8, textColor=rl_colors.HexColor("#27272A"))
        story = [
            RLP("SEALTECH CRM EXPORT", eyebrow),
            RLP("Maintenance Customers", title_style),
            RLSpacer(1, 0.15 * rl_inch),
        ]
        table_data = [headers]
        for row in data_rows:
            display = []
            for i, v in enumerate(row):
                if i == 4:
                    display.append(RLP(f"${float(v):,.0f}" if v else "—", body))
                else:
                    display.append(RLP(str(v) if v not in (None, "") else "—", body))
            table_data.append(display)
        if len(table_data) == 1:
            story.append(RLP("No maintenance plans yet.", body))
        else:
            tbl = RLTable(table_data, repeatRows=1)
            tbl.setStyle(RLTableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#062B67")),
                ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("TOPPADDING", (0, 0), (-1, 0), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, rl_colors.HexColor("#E4E4E7")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#FAFAFA")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(tbl)
        doc.build(story)
        return Response(
            content=buf.getvalue(),
            media_type="application/pdf",
            headers={"Content-Disposition": 'attachment; filename="sealtech-maintenance.pdf"'},
        )
    raise HTTPException(status_code=400, detail="Unsupported format (use xlsx or pdf)")


# ----- Import -----
def _parse_rows(data: bytes, filename: str):
    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = data.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        return [dict(r) for r in reader]
    # default to xlsx
    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v in (None, "") for v in row):
            continue
        rec = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            v = row[i] if i < len(row) else None
            rec[h] = "" if v is None else v
        rows.append(rec)
    return rows


def _normalize_row(category: str, raw: dict) -> dict:
    """Map exported headers back to model field keys."""
    cfg = EXPORT_CATEGORIES[
        "vendors" if category == "vendors" else "subcontractors" if category == "subcontractors" else category
    ]
    header_to_key = dict(zip(cfg["headers"], cfg["keys"]))
    out = {}
    for h, v in raw.items():
        key = header_to_key.get(h.strip()) if isinstance(h, str) else None
        if not key or key.startswith("_"):
            continue
        out[key] = v if isinstance(v, str) else ("" if v is None else v)
    return out


async def _find_existing(category: str, rec: dict):
    if category == "contacts":
        q = []
        if rec.get("email"):
            q.append({"email": str(rec["email"]).lower()})
        if rec.get("company_name") and rec.get("contact_name"):
            q.append({"company_name": rec["company_name"], "contact_name": rec["contact_name"]})
        if not q:
            return None
        return await db.contacts.find_one({"$or": q})
    if category == "properties":
        if rec.get("property_address") and rec.get("property_city"):
            return await db.properties.find_one({
                "property_address": rec["property_address"],
                "property_city": rec["property_city"],
            })
        return None
    if category == "projects":
        if rec.get("title"):
            return await db.deals.find_one({"title": rec["title"]})
        return None
    if category in ("vendors", "subcontractors"):
        kind = "Vendor" if category == "vendors" else "Subcontractor"
        q = {"kind": kind, "name": rec.get("name", "")}
        if rec.get("tin_ein"):
            q["tin_ein"] = rec["tin_ein"]
        return await db.vendors.find_one(q) if rec.get("name") else None
    return None


@api_router.post("/import/{category}")
async def import_category(
    category: str,
    file: UploadFile = File(...),
    duplicate_mode: str = Form("skip"),
    current=Depends(get_current_user),
):
    if category not in IMPORT_CATEGORIES:
        raise HTTPException(status_code=400, detail="Unknown category")
    if duplicate_mode not in DUPLICATE_MODES:
        duplicate_mode = "skip"

    raw_bytes = await file.read()
    try:
        rows = _parse_rows(raw_bytes, file.filename)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not parse file: {e}")

    imported = 0
    updated = 0
    skipped = 0
    errors = []

    for idx, raw in enumerate(rows, start=2):  # row index in spreadsheet (header is row 1)
        try:
            rec = _normalize_row(category, raw)
            if not rec:
                continue
            # category-specific seeding
            if category == "vendors":
                rec["kind"] = "Vendor"
            elif category == "subcontractors":
                rec["kind"] = "Subcontractor"
            # Coerce numeric for projects
            if category == "projects":
                for nk in ["proposal_option_1", "proposal_option_2", "proposal_option_3", "proposal_option_25yr", "chosen_amount"]:
                    try:
                        rec[nk] = float(rec.get(nk) or 0)
                    except (TypeError, ValueError):
                        rec[nk] = 0.0
            # Required field check
            required = {
                "contacts": "contact_name",
                "properties": "property_name",
                "projects": "title",
                "vendors": "name",
                "subcontractors": "name",
            }[category]
            if not rec.get(required):
                skipped += 1
                errors.append({"row": idx, "error": f"Missing required field: {required}"})
                continue

            existing = await _find_existing(category, rec)

            if existing and duplicate_mode == "skip":
                skipped += 1
                continue
            if existing and duplicate_mode == "update":
                coll = "deals" if category == "projects" else ("vendors" if category in ("vendors", "subcontractors") else category)
                await db[coll].update_one({"id": existing["id"]}, {"$set": rec})
                updated += 1
                continue
            # create
            rec["id"] = str(uuid.uuid4())
            rec["created_at"] = now_iso()
            coll = "deals" if category == "projects" else ("vendors" if category in ("vendors", "subcontractors") else category)
            await db[coll].insert_one(rec.copy())
            imported += 1
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})

    return {
        "imported": imported,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
        "total": len(rows),
    }


# ----- Materials Catalog -----
@api_router.get("/materials")
async def list_materials(category: Optional[str] = None, current=Depends(get_current_user)):
    query = {"is_deleted": {"$ne": True}}
    if category and category != "All":
        query["category"] = category
    items = await db.materials.find(query, {"_id": 0}).sort("name", 1).to_list(5000)
    return items


@api_router.post("/materials", response_model=Material)
async def create_material(body: MaterialIn, current=Depends(get_current_user)):
    data = body.model_dump()
    if data.get("vendor_id") and not data.get("vendor_name"):
        v = await db.vendors.find_one({"id": data["vendor_id"]}, {"_id": 0})
        if v:
            data["vendor_name"] = v.get("name", "")
    data["id"] = str(uuid.uuid4())
    data["created_at"] = now_iso()
    data["updated_at"] = now_iso()
    data["is_deleted"] = False
    await db.materials.insert_one(data.copy())
    return strip_id(data)


@api_router.put("/materials/{material_id}", response_model=Material)
async def update_material(material_id: str, body: MaterialIn, current=Depends(get_current_user)):
    existing = await db.materials.find_one({"id": material_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Material not found")
    data = body.model_dump()
    if data.get("vendor_id") and not data.get("vendor_name"):
        v = await db.vendors.find_one({"id": data["vendor_id"]}, {"_id": 0})
        if v:
            data["vendor_name"] = v.get("name", "")
    data["id"] = existing["id"]
    data["created_at"] = existing.get("created_at")
    data["updated_at"] = now_iso()
    await db.materials.update_one({"id": material_id}, {"$set": data})
    return strip_id(data)


@api_router.delete("/materials/{material_id}")
async def delete_material(material_id: str, current=Depends(get_current_user)):
    if is_admin(current):
        await db.materials.delete_one({"id": material_id})
    else:
        await db.materials.update_one({"id": material_id}, {"$set": {"is_deleted": True, "deleted_at": now_iso(), "deleted_by": current["id"]}})
    return {"ok": True}


@api_router.post("/materials/bulk-import")
async def bulk_import_materials(file: UploadFile = File(...), current=Depends(get_current_user)):
    """Import materials from CSV or Excel. Required columns: name (rest optional)."""
    raw = await file.read()
    try:
        import pandas as pd
        from io import BytesIO as _BIO
        if (file.filename or "").lower().endswith(".csv"):
            df = pd.read_csv(_BIO(raw))
        else:
            df = pd.read_excel(_BIO(raw))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read file: {e}")

    # Normalize column names (lowercase + strip)
    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
    if "name" not in df.columns:
        raise HTTPException(status_code=400, detail="CSV must include a 'name' column")

    # Match vendors by name (case-insensitive)
    vendors = await db.vendors.find({"is_deleted": {"$ne": True}}, {"_id": 0, "id": 1, "name": 1}).to_list(500)
    vendor_lookup = {(v.get("name") or "").strip().lower(): v["id"] for v in vendors}

    created = 0
    updated = 0
    skipped = 0
    for _, row in df.iterrows():
        name = str(row.get("name", "") or "").strip()
        if not name:
            skipped += 1
            continue
        sku = str(row.get("sku", "") or "").strip()
        category = str(row.get("category", "") or "Other").strip() or "Other"
        unit = str(row.get("unit", "") or "each").strip() or "each"
        try:
            price = float(row.get("default_price") or row.get("price") or 0)
        except (TypeError, ValueError):
            price = 0.0
        try:
            shipping_pct = float(row.get("shipping_pct") or row.get("shipping%") or 0)
        except (TypeError, ValueError):
            shipping_pct = 0.0
        try:
            markup_pct = float(row.get("markup_pct") or row.get("markup%") or 0)
        except (TypeError, ValueError):
            markup_pct = 0.0
        vendor_name = str(row.get("preferred_vendor", "") or row.get("vendor", "") or "").strip()
        vendor_id = vendor_lookup.get(vendor_name.lower()) if vendor_name else None
        notes = str(row.get("notes", "") or "").strip()

        # Find existing by SKU (preferred) or name
        existing = None
        if sku:
            existing = await db.materials.find_one({"sku": sku, "is_deleted": {"$ne": True}}, {"_id": 0})
        if not existing:
            existing = await db.materials.find_one({"name": name, "is_deleted": {"$ne": True}}, {"_id": 0})

        doc = {
            "sku": sku,
            "name": name,
            "category": category,
            "unit": unit,
            "default_price": price,
            "shipping_pct": shipping_pct,
            "markup_pct": markup_pct,
            "vendor_id": vendor_id,
            "vendor_name": vendor_name,
            "notes": notes,
            "updated_at": now_iso(),
        }
        if existing:
            await db.materials.update_one({"id": existing["id"]}, {"$set": doc})
            updated += 1
        else:
            doc["id"] = str(uuid.uuid4())
            doc["created_at"] = now_iso()
            doc["is_deleted"] = False
            await db.materials.insert_one(doc)
            created += 1
    return {"created": created, "updated": updated, "skipped": skipped}


@api_router.get("/materials/export.xlsx")
async def export_materials(current=Depends(get_current_user)):
    items = await db.materials.find({"is_deleted": {"$ne": True}}, {"_id": 0}).sort("name", 1).to_list(5000)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Materials"
    headers = ["SKU", "Name", "Category", "Unit", "Default Price", "Shipping %", "Markup %", "Loaded Cost", "Preferred Vendor", "Notes", "Last Updated"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="062B67")
        cell.alignment = Alignment(horizontal="left")
    for m in items:
        price = float(m.get("default_price") or 0)
        ship_pct = float(m.get("shipping_pct") or 0)
        loaded = round(price * (1 + ship_pct / 100), 2)
        ws.append([m.get("sku", ""), m.get("name", ""), m.get("category", ""), m.get("unit", ""), price, ship_pct, float(m.get("markup_pct") or 0), loaded, m.get("vendor_name", ""), m.get("notes", ""), (m.get("updated_at") or "")[:10]])
    ws.freeze_panes = "A2"
    for i in range(1, 12):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 18
    buf = BytesIO()
    wb.save(buf)
    return Response(content=buf.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": 'attachment; filename="sealtech-materials.xlsx"'})


@api_router.get("/materials/template.xlsx")
async def materials_template(current=Depends(get_current_user)):
    """Empty CSV-ready template for users to fill out."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); ws = wb.active; ws.title = "Materials"
    headers = ["sku", "name", "category", "unit", "default_price", "shipping_pct", "markup_pct", "preferred_vendor", "notes"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="A0703A")
    ws.append(["EX-001", "Example Silicone 5gal", "Coating", "5-gal pail", 285.0, 8.0, 35.0, "ABC Supply", "Optional notes"])
    for i in range(1, 10):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 22
    buf = BytesIO()
    wb.save(buf)
    return Response(content=buf.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": 'attachment; filename="sealtech-materials-template.xlsx"'})


# ----- Dashboard -----

@api_router.get("/dashboard/materials-in-motion")
async def materials_in_motion(current=Depends(get_current_user)):
    """Projects with take-off lines that are ordered but not yet received.

    Returns a list of projects with counts and the per-vendor breakdown so the user
    can see at a glance which suppliers are still owed deliveries.
    """
    query = {"is_deleted": {"$ne": True}, "material_takeoff": {"$exists": True, "$not": {"$size": 0}}}
    # Sales reps only see their own projects
    if current.get("role") == "sales":
        query["$or"] = [
            {"assigned_to_user_id": current["id"]},
            {"created_by_user_id": current["id"]},
        ]
    cursor = db.deals.find(query, {"_id": 0, "id": 1, "title": 1, "material_takeoff": 1, "status": 1}).sort("title", 1)
    deals = await cursor.to_list(1000)

    out_projects = []
    total_ordered_lines = 0
    total_received_lines = 0
    total_pending_lines = 0
    total_open_value = 0.0
    vendor_aggregate: dict = {}

    for d in deals:
        lines = d.get("material_takeoff") or []
        if not lines:
            continue
        ordered = [ln for ln in lines if ln.get("ordered") and not ln.get("received")]
        received = [ln for ln in lines if ln.get("received")]
        pending = [ln for ln in lines if not ln.get("ordered") and not ln.get("received")]
        total_ordered_lines += len(ordered)
        total_received_lines += len(received)
        total_pending_lines += len(pending)

        if not ordered:
            continue  # only include projects with stuff actually in motion

        per_vendor: dict = {}
        open_value = 0.0
        for ln in ordered:
            v = ln.get("vendor_name") or "Unassigned"
            pv = per_vendor.setdefault(v, {"vendor_name": v, "vendor_id": ln.get("vendor_id"), "lines": 0, "value": 0.0})
            pv["lines"] += 1
            amt = float(ln.get("line_total") or 0)
            pv["value"] += amt
            open_value += amt

            va = vendor_aggregate.setdefault(v, {"vendor_name": v, "vendor_id": ln.get("vendor_id"), "lines": 0, "projects": set(), "value": 0.0})
            va["lines"] += 1
            va["projects"].add(d["id"])
            va["value"] += amt

        total_open_value += open_value
        out_projects.append({
            "id": d["id"],
            "title": d.get("title", "") or "",
            "status": d.get("status", ""),
            "lines_ordered": len(ordered),
            "lines_received": len(received),
            "lines_pending": len(pending),
            "open_value": round(open_value, 2),
            "vendors": sorted([
                {"vendor_name": v["vendor_name"], "vendor_id": v["vendor_id"], "lines": v["lines"], "value": round(v["value"], 2)}
                for v in per_vendor.values()
            ], key=lambda x: x["vendor_name"]),
        })

    # Sort projects with the most open value first
    out_projects.sort(key=lambda p: -p["open_value"])

    by_vendor = sorted([
        {
            "vendor_name": v["vendor_name"],
            "vendor_id": v["vendor_id"],
            "lines": v["lines"],
            "projects": len(v["projects"]),
            "value": round(v["value"], 2),
        }
        for v in vendor_aggregate.values()
    ], key=lambda x: -x["value"])

    return {
        "totals": {
            "projects_with_open_orders": len(out_projects),
            "lines_ordered_not_received": total_ordered_lines,
            "lines_received": total_received_lines,
            "lines_pending_to_order": total_pending_lines,
            "open_value": round(total_open_value, 2),
        },
        "projects": out_projects,
        "by_vendor": by_vendor,
    }


@api_router.get("/calendar")
async def calendar_events(start: str, end: str, current=Depends(get_current_user)):
    """Unified calendar feed for the Project Calendar page.

    Returns a flat list of color-coded events between `start` and `end` (inclusive,
    YYYY-MM-DD). Event kinds:
      - project        : scheduled_start_date → scheduled_end_date span (blue)
      - material_order : Deal.material_order_date (amber)
      - maintenance    : Deal.maintenance_visits[].visit_date OR next_maintenance_date (green)
      - coi_expiry     : Vendor GL/WC COI expiry (red)
      - invoice_due    : Invoice.due_date (purple)
    """
    def _in_range(d: str) -> bool:
        return bool(d) and start <= d <= end

    events = []

    # --- Projects (span) + Material Orders (single day) ---
    deals_cursor = db.deals.find(
        {"is_deleted": {"$ne": True}},
        {
            "_id": 0, "id": 1, "title": 1, "status": 1, "deal_type": 1,
            "scheduled_start_date": 1, "scheduled_end_date": 1, "material_order_date": 1,
            "next_maintenance_date": 1, "maintenance_visits": 1, "maintenance_plan": 1,
            "chosen_amount": 1, "property_id": 1,
        },
    )
    deals = await deals_cursor.to_list(5000)
    for d in deals:
        s = d.get("scheduled_start_date") or ""
        e = d.get("scheduled_end_date") or s
        if s and (_in_range(s) or _in_range(e) or (s <= start and (e or s) >= end)):
            events.append({
                "id": f"project-{d['id']}",
                "kind": "project",
                "title": d.get("title") or "Project",
                "start": s,
                "end": e or s,
                "color": "#1D4ED8",  # cobalt blue
                "deal_id": d["id"],
                "status": d.get("status") or "",
                "amount": float(d.get("chosen_amount") or 0),
            })
        mo = d.get("material_order_date") or ""
        if _in_range(mo):
            events.append({
                "id": f"material-{d['id']}",
                "kind": "material_order",
                "title": f"Materials: {d.get('title') or 'Project'}",
                "start": mo,
                "end": mo,
                "color": "#D97706",  # amber
                "deal_id": d["id"],
            })
        # Maintenance visits — each visit in window, plus next_maintenance_date as a tentative slot
        for v in (d.get("maintenance_visits") or []):
            vd = v.get("visit_date") or ""
            if _in_range(vd):
                events.append({
                    "id": f"maint-{d['id']}-{v.get('id','')}",
                    "kind": "maintenance",
                    "title": f"Maintenance: {d.get('title') or 'Project'}",
                    "start": vd,
                    "end": vd,
                    "color": "#16A34A",  # green
                    "deal_id": d["id"],
                    "visit_id": v.get("id", ""),
                    "amount": float(v.get("amount") or 0),
                })
        nm = d.get("next_maintenance_date") or ""
        if d.get("maintenance_plan") and _in_range(nm):
            # Avoid duplicate if there's already a visit logged that date
            already = any(ev.get("kind") == "maintenance" and ev.get("start") == nm and ev.get("deal_id") == d["id"] for ev in events)
            if not already:
                events.append({
                    "id": f"maint-next-{d['id']}",
                    "kind": "maintenance",
                    "title": f"Maintenance Due: {d.get('title') or 'Project'}",
                    "start": nm,
                    "end": nm,
                    "color": "#16A34A",
                    "deal_id": d["id"],
                    "tentative": True,
                })

    # --- COI Expirations (red) ---
    vendors_cursor = db.vendors.find(
        {"is_deleted": {"$ne": True}},
        {"_id": 0, "id": 1, "name": 1, "kind": 1, "gl_coi_expiry_date": 1, "wc_coi_expiry_date": 1, "gl_coi_on_file": 1, "wc_coi_on_file": 1},
    )
    vendors = await vendors_cursor.to_list(2000)
    for v in vendors:
        gl = v.get("gl_coi_expiry_date") or ""
        if v.get("gl_coi_on_file") and _in_range(gl):
            events.append({
                "id": f"coi-gl-{v['id']}",
                "kind": "coi_expiry",
                "title": f"GL COI Expires: {v.get('name') or 'Vendor'}",
                "start": gl,
                "end": gl,
                "color": "#B91C1C",  # red
                "vendor_id": v["id"],
                "coi_type": "GL",
            })
        wc = v.get("wc_coi_expiry_date") or ""
        if v.get("wc_coi_on_file") and _in_range(wc):
            events.append({
                "id": f"coi-wc-{v['id']}",
                "kind": "coi_expiry",
                "title": f"WC COI Expires: {v.get('name') or 'Vendor'}",
                "start": wc,
                "end": wc,
                "color": "#B91C1C",
                "vendor_id": v["id"],
                "coi_type": "WC",
            })

    # --- Invoice Due Dates (purple) — only unpaid balances ---
    inv_cursor = db.invoices.find(
        {"is_deleted": {"$ne": True}, "status": {"$nin": ["Paid", "Void"]}},
        {"_id": 0, "id": 1, "invoice_number": 1, "due_date": 1, "balance_due": 1, "bill_to_name": 1, "bill_to_company": 1, "deal_id": 1},
    )
    invs = await inv_cursor.to_list(5000)
    for inv in invs:
        dd = inv.get("due_date") or ""
        if _in_range(dd):
            who = inv.get("bill_to_company") or inv.get("bill_to_name") or ""
            events.append({
                "id": f"invoice-{inv['id']}",
                "kind": "invoice_due",
                "title": f"Invoice Due: {inv.get('invoice_number','')} — {who}",
                "start": dd,
                "end": dd,
                "color": "#7E22CE",  # purple
                "invoice_id": inv["id"],
                "deal_id": inv.get("deal_id"),
                "amount": float(inv.get("balance_due") or 0),
            })

    # --- Ad-hoc Deal Events (appointments) — teal ---
    ade_cursor = db.deal_events.find(
        {
            "is_deleted": {"$ne": True},
            "date": {"$gte": start, "$lte": end},
        },
        {"_id": 0},
    )
    ad_hoc = await ade_cursor.to_list(2000)
    EMOJI = {
        "Roof Walk": "🪜",
        "Presentation": "📊",
        "Meeting": "🤝",
        "Job Start": "🚧",
        "Other": "📅",
    }
    for ev in ad_hoc:
        emoji = EMOJI.get(ev.get("event_type") or "Other", "📅")
        time_part = f" {ev['start_time']}" if ev.get("start_time") else ""
        events.append({
            "id": f"appt-{ev['id']}",
            "kind": "appointment",
            "title": f"{emoji} {ev.get('event_type') or 'Appointment'}{time_part}: {ev.get('title') or ''}",
            "start": ev["date"],
            "end": ev["date"],
            "color": "#0F766E",  # teal
            "deal_id": ev.get("deal_id"),
            "event_type": ev.get("event_type"),
            "start_time": ev.get("start_time") or "",
            "end_time": ev.get("end_time") or "",
            "location": ev.get("location") or "",
            "notes": ev.get("notes") or "",
        })

    events.sort(key=lambda x: (x["start"], x["kind"]))
    return events


@api_router.get("/dashboard/today")
async def dashboard_today(current=Depends(get_current_user)):
    """Today + next-2-days widget for the Dashboard.

    Returns ad-hoc deal events (appointments) scheduled for today or the next
    48 hours, with the linked deal title attached so the widget can render
    a clickable card without making N+1 lookups.
    """
    today_iso = datetime.now(timezone.utc).date().isoformat()
    soon_iso = (datetime.now(timezone.utc) + timedelta(days=2)).date().isoformat()
    cur = db.deal_events.find(
        {
            "is_deleted": {"$ne": True},
            "date": {"$gte": today_iso, "$lte": soon_iso},
        },
        {"_id": 0},
    ).sort([("date", 1), ("start_time", 1)])
    events = await cur.to_list(200)
    # Attach deal title in one pass
    deal_ids = list({ev["deal_id"] for ev in events})
    deals = await db.deals.find(
        {"id": {"$in": deal_ids}}, {"_id": 0, "id": 1, "title": 1}
    ).to_list(len(deal_ids)) if deal_ids else []
    by_id = {d["id"]: d.get("title") or "" for d in deals}
    for ev in events:
        ev["deal_title"] = by_id.get(ev["deal_id"], "")
    return {"today": today_iso, "events": events}

# ---------- Soft-delete audit (last 48h) ----------
@api_router.get("/admin/recent-deletions")
async def admin_recent_deletions(hours: int = 48, current=Depends(get_current_user)):
    """Inventory of everything soft-deleted in the last `hours` hours across
    the major collections. Use after running tests, or any time you suspect
    something got removed by mistake.

    Each row includes the collection name, document id, a human-readable
    label, the timestamp it was deleted, and (for photos) the parent deal
    so you can find your way back to it.
    """
    if current.get("role") != "admin":
        raise HTTPException(403, "Admins only")
    cutoff = (datetime.now(timezone.utc) - timedelta(hours=max(1, hours))).isoformat()

    out = []
    # project_photos
    cur = db.project_photos.find(
        {"is_deleted": True, "updated_at": {"$gte": cutoff}},
        {"_id": 0, "id": 1, "deal_id": 1, "display_name": 1, "album_name": 1, "updated_at": 1, "uploader_name": 1, "size": 1},
    ).sort([("updated_at", -1)])
    rows = await cur.to_list(500)
    deal_ids = list({r["deal_id"] for r in rows})
    deals = await db.deals.find({"id": {"$in": deal_ids}}, {"_id": 0, "id": 1, "title": 1}).to_list(len(deal_ids)) if deal_ids else []
    deal_titles = {d["id"]: d.get("title") or "?" for d in deals}
    for r in rows:
        out.append({
            "kind": "photo",
            "id": r["id"],
            "label": r.get("display_name") or "(untitled)",
            "deleted_at": r.get("updated_at"),
            "deleted_by": r.get("uploader_name") or "?",
            "context": f"{deal_titles.get(r['deal_id'], '?')} · album {r.get('album_name','Default')}",
            "deal_id": r["deal_id"],
            "size_kb": (r.get("size") or 0) // 1024,
            "restorable": True,
        })

    # deals
    cur2 = db.deals.find(
        {"is_deleted": True, "updated_at": {"$gte": cutoff}},
        {"_id": 0, "id": 1, "title": 1, "status": 1, "updated_at": 1},
    ).sort([("updated_at", -1)])
    for d in await cur2.to_list(200):
        out.append({
            "kind": "deal",
            "id": d["id"],
            "label": d.get("title") or "?",
            "deleted_at": d.get("updated_at"),
            "deleted_by": "?",
            "context": f"status {d.get('status','?')}",
            "restorable": True,
        })

    # Sort newest first
    out.sort(key=lambda x: x.get("deleted_at") or "", reverse=True)
    return {
        "hours": hours,
        "cutoff": cutoff,
        "total": len(out),
        "by_kind": {k: sum(1 for x in out if x["kind"] == k) for k in {"photo", "deal"}},
        "items": out,
    }


@api_router.post("/admin/restore/{kind}/{item_id}")
async def admin_restore(kind: str, item_id: str, current=Depends(get_current_user)):
    """One-click restore for a soft-deleted item. Supports kind in {photo, deal}."""
    if current.get("role") != "admin":
        raise HTTPException(403, "Admins only")
    coll_map = {"photo": db.project_photos, "deal": db.deals}
    coll = coll_map.get(kind)
    if coll is None:
        raise HTTPException(400, f"Unsupported kind: {kind}")
    res = await coll.update_one(
        {"id": item_id, "is_deleted": True},
        {"$set": {"is_deleted": False, "updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    if res.matched_count == 0:
        raise HTTPException(404, "Not found or not deleted")
    return {"restored": True, "kind": kind, "id": item_id}





@api_router.get("/dashboard/compliance-wall")
async def dashboard_compliance_wall(current=Depends(get_current_user)):
    """Every user certification expiring within 60 days — for the dashboard
    widget. Sorted by days_until_expiration ascending (most urgent first).

    Visible to admins only — payroll/HR adjacent. Non-admin returns 403 so
    the widget hides itself.
    """
    if current.get("role") != "admin":
        raise HTTPException(403, "Admins only")

    today = datetime.now(timezone.utc).date()
    cutoff = (today + timedelta(days=60)).isoformat()

    rows = await db.user_certifications.find(
        {
            "is_deleted": {"$ne": True},
            "expiration_date": {"$ne": "", "$lte": cutoff},
        },
        {"_id": 0},
    ).to_list(500)

    # Join user info for each row (single batched query)
    user_ids = list({r["user_id"] for r in rows})
    users = await db.users.find(
        {"id": {"$in": user_ids}, "is_deleted": {"$ne": True}},
        {"_id": 0, "id": 1, "name": 1, "email": 1, "role": 1},
    ).to_list(len(user_ids)) if user_ids else []
    by_id = {u["id"]: u for u in users}

    enriched = []
    for r in rows:
        try:
            exp = datetime.fromisoformat(str(r["expiration_date"])[:10]).date()
            days = (exp - today).days
        except Exception:
            continue
        u = by_id.get(r["user_id"])
        if not u:  # user deleted but cert still there — skip silently
            continue
        enriched.append({
            "cert_id": r["id"],
            "user_id": u["id"],
            "user_name": u.get("name") or u.get("email") or "",
            "user_email": u.get("email") or "",
            "name": r["name"],
            "issuer": r.get("issuer") or "",
            "expiration_date": r["expiration_date"],
            "days_until_expiration": days,
        })
    enriched.sort(key=lambda x: x["days_until_expiration"])
    return {
        "today": today.isoformat(),
        "items": enriched,
        "expired_count":   sum(1 for x in enriched if x["days_until_expiration"] < 0),
        "due_7_count":     sum(1 for x in enriched if 0 <= x["days_until_expiration"] <= 7),
        "due_30_count":    sum(1 for x in enriched if 7 < x["days_until_expiration"] <= 30),
        "due_60_count":    sum(1 for x in enriched if 30 < x["days_until_expiration"] <= 60),
    }





@api_router.get("/dashboard/summary")
async def dashboard_summary(current=Depends(get_current_user)):
    deals = await db.deals.find({}, {"_id": 0}).to_list(5000)
    contacts_count = await db.contacts.count_documents({})
    properties_count = await db.properties.count_documents({})

    open_leads = 0
    won_deals = 0
    lost_deals = 0
    pipeline_revenue = 0.0
    won_revenue = 0.0
    total_costs = 0.0

    year_start = datetime(now_utc().year, 1, 1, tzinfo=timezone.utc).isoformat()
    profit_ytd = 0.0

    # Maintenance KPIs
    today_iso = datetime.now(timezone.utc).date().isoformat()
    soon_iso = (datetime.now(timezone.utc) + timedelta(days=30)).date().isoformat()
    maintenance_count = 0
    maintenance_due_30d = 0
    maintenance_overdue = 0
    maintenance_annual_revenue = 0.0

    for d in deals:
        status_v = d.get("status", "Lead")
        chosen = float(d.get("chosen_amount", 0) or 0)
        # Pipeline value: use chosen_amount if set, otherwise the MID proposal option (typical buy point)
        pipeline_value = chosen if chosen > 0 else proposal_mid_amount(d)
        costs = float(d.get("materials_cost", 0) or 0) + float(d.get("labor_cost", 0) or 0) + float(d.get("subcontractor_cost", 0) or 0) + float(d.get("other_expenses", 0) or 0)
        if status_v == "Won":
            won_deals += 1
            won_revenue += chosen
            total_costs += costs
            if d.get("created_at", "") >= year_start:
                profit_ytd += (chosen - costs)
        elif status_v == "Lost":
            lost_deals += 1
        else:
            open_leads += 1
            pipeline_revenue += pipeline_value

        # Maintenance roll-up
        if d.get("maintenance_plan"):
            maintenance_count += 1
            maintenance_annual_revenue += float(d.get("maintenance_rate", 0) or 0)
            nxt = d.get("next_maintenance_date") or ""
            if nxt:
                if nxt < today_iso:
                    maintenance_overdue += 1
                    maintenance_due_30d += 1
                elif nxt <= soon_iso:
                    maintenance_due_30d += 1

    return {
        "contacts_count": contacts_count,
        "properties_count": properties_count,
        "deals_count": len(deals),
        "open_leads": open_leads,
        "won_deals": won_deals,
        "lost_deals": lost_deals,
        "pipeline_revenue": round(pipeline_revenue, 2),
        "won_revenue": round(won_revenue, 2),
        "total_costs": round(total_costs, 2),
        "total_profit": round(won_revenue - total_costs, 2),
        "profit_ytd": round(profit_ytd, 2),
        "maintenance_count": maintenance_count,
        "maintenance_due_30d": maintenance_due_30d,
        "maintenance_overdue": maintenance_overdue,
        "maintenance_annual_revenue": round(maintenance_annual_revenue, 2),
        **(await _payables_summary(current)),
    }


async def _payables_summary(current) -> dict:
    """Sum unpaid bills, grouped by overdue / due-soon / total."""
    query = {"is_deleted": {"$ne": True}, "status": {"$in": ["Pending", "Approved"]}}
    if current.get("role") == "sales":
        query["created_by_user_id"] = current["id"]
    bills = await db.vendor_bills.find(query, {"_id": 0}).to_list(2000)
    today_iso = datetime.now(timezone.utc).date().isoformat()
    soon_iso = (datetime.now(timezone.utc) + timedelta(days=7)).date().isoformat()
    out_total = 0.0
    overdue_total = 0.0
    due_week_total = 0.0
    overdue_count = 0
    due_week_count = 0
    for b in bills:
        balance = float(b.get("total") or 0) - float(b.get("paid_amount") or 0)
        if balance <= 0.01:
            continue
        out_total += balance
        dd = (b.get("due_date") or "")[:10]
        if dd and dd < today_iso:
            overdue_total += balance
            overdue_count += 1
        elif dd and dd <= soon_iso:
            due_week_total += balance
            due_week_count += 1
    return {
        "payables_outstanding": round(out_total, 2),
        "payables_overdue": round(overdue_total, 2),
        "payables_due_this_week": round(due_week_total, 2),
        "payables_overdue_count": overdue_count,
        "payables_due_this_week_count": due_week_count,
    }


async def _compute_stale_deals(days: int = 14, won_grace_days: int = 30, owner_user_id: Optional[str] = None):
    """Shared engine for the Stale-Deals dashboard widget AND the weekly
    digest emailer. `owner_user_id` (optional) restricts to deals
    assigned_to / created_by that user — used both for the role=sales scope
    and for per-owner digest filtering.
    """
    days = max(1, int(days or 14))
    won_grace_days = max(1, int(won_grace_days or 30))

    query: dict = {"is_deleted": {"$ne": True}}
    if owner_user_id:
        query["$or"] = [
            {"assigned_to_user_id": owner_user_id},
            {"created_by_user_id": owner_user_id},
        ]

    deals = await db.deals.find(
        query,
        {
            "_id": 0, "id": 1, "title": 1, "status": 1, "status_history": 1,
            "created_at": 1, "chosen_amount": 1, "project_type": 1,
            "assigned_to_user_id": 1, "created_by_user_id": 1,
            "primary_contact_name": 1, "property_address": 1,
            "payment_milestones": 1,
        },
    ).to_list(5000)

    now = datetime.now(timezone.utc)
    stale_threshold = now - timedelta(days=days)
    won_threshold = now - timedelta(days=won_grace_days)

    def _parse_iso(s):
        if not s:
            return None
        try:
            return datetime.fromisoformat(str(s).replace("Z", "+00:00"))
        except Exception:
            return None

    invoice_paid_deals = set()
    async for inv in db.invoices.find(
        {"is_deleted": {"$ne": True}, "amount_paid": {"$gt": 0}},
        {"_id": 0, "deal_id": 1},
    ):
        if inv.get("deal_id"):
            invoice_paid_deals.add(inv["deal_id"])

    deposit_received_deals = set()
    for d in deals:
        for m in (d.get("payment_milestones") or []):
            if (m.get("status") == "Paid") or float(m.get("amount_received") or 0) > 0:
                deposit_received_deals.add(d["id"])
                break

    results = []
    for d in deals:
        status_v = d.get("status") or "Lead"
        history = d.get("status_history") or []
        last_change_iso = history[-1].get("at") if history else None
        last_change = _parse_iso(last_change_iso) or _parse_iso(d.get("created_at"))
        if not last_change:
            continue
        days_in_stage = int((now - last_change).total_seconds() // 86400)
        reason = None
        priority = 0
        if status_v == "Won":
            has_money = d["id"] in invoice_paid_deals or d["id"] in deposit_received_deals
            if not has_money and last_change <= won_threshold:
                reason = "no_deposit"
                priority = 100 + days_in_stage
        elif status_v in ("Lost", "Past Lead"):
            pass
        else:
            if last_change <= stale_threshold:
                reason = "stuck"
                priority = days_in_stage
        if not reason:
            continue
        results.append({
            "id": d["id"],
            "title": d.get("title") or "Untitled project",
            "status": status_v,
            "project_type": d.get("project_type") or "",
            "chosen_amount": float(d.get("chosen_amount") or 0),
            "primary_contact_name": d.get("primary_contact_name") or "",
            "property_address": d.get("property_address") or "",
            "days_in_stage": days_in_stage,
            "last_change_at": last_change.isoformat(),
            "reason": reason,
            "priority": priority,
            "owner_user_id": d.get("assigned_to_user_id") or d.get("created_by_user_id") or "",
        })

    results.sort(key=lambda r: r["priority"], reverse=True)
    return results


@api_router.get("/dashboard/stale-deals")
async def dashboard_stale_deals(
    days: int = 14,
    won_grace_days: int = 30,
    current=Depends(get_current_user),
):
    """Surface deals that haven't moved in a while.

    Two reasons can flag a deal:
      • "stuck"   — open deal (not Won/Lost/Past Lead) that hasn't changed status
                    in `days` days. The age clock starts from the last
                    status_history entry, or created_at when no history exists.
      • "no_deposit" — deal flipped to Won `won_grace_days`+ ago and still has
                       zero collected from invoices (no deposit / payment).
    """
    owner = current["id"] if current.get("role") == "sales" else None
    results = await _compute_stale_deals(days, won_grace_days, owner_user_id=owner)
    counts = {
        "stuck": sum(1 for r in results if r["reason"] == "stuck"),
        "no_deposit": sum(1 for r in results if r["reason"] == "no_deposit"),
    }
    return {
        "threshold_days": max(1, int(days or 14)),
        "won_grace_days": max(1, int(won_grace_days or 30)),
        "counts": counts,
        "deals": results[:50],
    }


async def _build_and_send_owner_digest(
    user: dict,
    deals_for_owner: list,
    days: int,
    won_grace_days: int,
    cc_email: Optional[str] = None,
    dry_run: bool = False,
) -> dict:
    """Compose + send one owner's weekly stale-deals digest. Shared by the
    on-demand endpoint and the in-process scheduler (scheduler.py).
    `user` must have at minimum {id, email, name}.
    Returns a digest summary dict for the API response.
    """
    owner_id = user.get("id", "")
    owner_email = user.get("email", "")
    owner_name = (user.get("name") or "there").split(" ")[0]
    stuck = [d for d in deals_for_owner if d["reason"] == "stuck"]
    no_deposit = [d for d in deals_for_owner if d["reason"] == "no_deposit"]

    def _row_line(d):
        amt = f"${d['chosen_amount']:,.0f}" if d.get("chosen_amount") else "—"
        return f"  • [{d['status']}] {d['title']} — {d['days_in_stage']}d at this stage · {amt}"

    text_parts = [
        f"Hi {owner_name},",
        "",
        f"Here is your weekly Stale-Deals digest from SealTech CRM ({len(deals_for_owner)} item"
        f"{'s' if len(deals_for_owner) != 1 else ''} need your attention):",
        "",
    ]
    if stuck:
        text_parts.append(f"Stuck > {days} days at the same stage ({len(stuck)}):")
        text_parts.extend(_row_line(d) for d in stuck)
        text_parts.append("")
    if no_deposit:
        text_parts.append(f"Won {won_grace_days}+ days ago with no deposit recorded ({len(no_deposit)}):")
        text_parts.extend(_row_line(d) for d in no_deposit)
        text_parts.append("")
    text_parts.extend([
        "Open the dashboard to take action: /dashboard",
        "",
        "— SealTech CRM",
    ])
    body_text = "\n".join(text_parts)

    def _html_row(d):
        amt = f"${d['chosen_amount']:,.0f}" if d.get("chosen_amount") else "&mdash;"
        return (
            f'<li><b>[{d["status"]}]</b> {d["title"]} '
            f'<span style="color:#71717A">&mdash; {d["days_in_stage"]}d at this stage &middot; {amt}</span></li>'
        )

    html_sections = []
    if stuck:
        html_sections.append(
            f"<h4 style=\"color:#B45309;margin:16px 0 6px\">Stuck &gt; {days} days at the same stage ({len(stuck)})</h4>"
            f"<ul>{''.join(_html_row(d) for d in stuck)}</ul>"
        )
    if no_deposit:
        html_sections.append(
            f"<h4 style=\"color:#BE123C;margin:16px 0 6px\">Won {won_grace_days}+ days ago with no deposit ({len(no_deposit)})</h4>"
            f"<ul>{''.join(_html_row(d) for d in no_deposit)}</ul>"
        )
    body_html = (
        f"<p>Hi {owner_name},</p>"
        f"<p>Here is your weekly Stale-Deals digest from SealTech CRM "
        f"(<b>{len(deals_for_owner)}</b> item{'s' if len(deals_for_owner) != 1 else ''} need your attention):</p>"
        + "".join(html_sections)
        + '<p><a href="/dashboard" style="background:#062B67;color:white;padding:8px 14px;text-decoration:none;'
          'font-weight:bold;letter-spacing:1px;font-size:11px;text-transform:uppercase;border-radius:2px">Open Dashboard</a></p>'
        + "<p style=\"color:#71717A;font-size:11px\">&mdash; SealTech CRM</p>"
    )

    subject = f"Stale Deals Digest — {len(deals_for_owner)} need your attention"
    entry: dict = {
        "owner_user_id": owner_id,
        "owner_email": owner_email,
        "owner_name": user.get("name", ""),
        "stuck_count": len(stuck),
        "no_deposit_count": len(no_deposit),
        "subject": subject,
    }

    if not dry_run:
        try:
            from email_sender import send_for_category
            cc = cc_email if (cc_email and cc_email != owner_email) else None
            await send_for_category(
                db, "scope",
                to=owner_email, subject=subject,
                body_text=body_text, body_html=body_html, cc=cc,
            )
            entry["sent"] = True
        except Exception as e:
            entry["sent"] = False
            entry["error"] = str(e)
    return entry


@api_router.post("/dashboard/stale-deals/digest")
async def send_stale_deals_digest(
    days: int = 14,
    won_grace_days: int = 30,
    dry_run: bool = False,
    cc_admin: bool = True,
    current=Depends(get_current_user),
):
    """Send each deal-owner a digest of their stuck deals + Won-without-deposit alerts.

    Admin-only. Intended to be fired weekly (e.g., Monday morning via cron),
    but can also be triggered on-demand from the dashboard.
    Pass `dry_run=true` to preview the recipient list + counts without sending.
    """
    if current.get("role") != "admin":
        raise HTTPException(403, "Admins only")

    rows = await _compute_stale_deals(days, won_grace_days)
    # Group by owner_user_id (drop unassigned rows for the digest — they have no recipient)
    by_owner: dict[str, list[dict]] = {}
    for r in rows:
        owner_id = r.get("owner_user_id") or ""
        if not owner_id:
            continue
        by_owner.setdefault(owner_id, []).append(r)

    # Resolve owner email/name in one round-trip
    owner_ids = list(by_owner.keys())
    users = await db.users.find(
        {"id": {"$in": owner_ids}},
        {"_id": 0, "id": 1, "email": 1, "name": 1, "role": 1},
    ).to_list(500) if owner_ids else []
    user_map = {u["id"]: u for u in users}

    admin_email = (current.get("email") or "").strip()
    digests: list[dict] = []
    sent = 0
    skipped: list[dict] = []

    for owner_id, deals_for_owner in by_owner.items():
        user = user_map.get(owner_id)
        if not user or not user.get("email"):
            skipped.append({"owner_user_id": owner_id, "reason": "no_email_on_file", "count": len(deals_for_owner)})
            continue
        cc = admin_email if (cc_admin and admin_email and admin_email != user["email"]) else None
        digest_entry = await _build_and_send_owner_digest(
            user=user,
            deals_for_owner=deals_for_owner,
            days=days,
            won_grace_days=won_grace_days,
            cc_email=cc,
            dry_run=dry_run,
        )
        if digest_entry.get("sent") is True:
            sent += 1
        digests.append(digest_entry)

    return {
        "dry_run": dry_run,
        "threshold_days": max(1, int(days or 14)),
        "won_grace_days": max(1, int(won_grace_days or 30)),
        "owners_eligible": len(by_owner),
        "sent": sent,
        "skipped": skipped,
        "digests": digests,
    }


@api_router.get("/dashboard/revenue-by-type")
async def revenue_by_type(window: str = "ytd", current=Depends(get_current_user)):
    """Booked + Received revenue grouped by project_type.
    window = 'ytd' (current calendar year) or 'all' (all-time).
    Maintenance is a special bucket fed by maintenance_visits, NOT by project_type=Maintenance deals
    (those would also be included separately if user labels a one-off project as Maintenance type)."""
    if window not in ("ytd", "all"):
        window = "ytd"
    year_start_iso = datetime(now_utc().year, 1, 1, tzinfo=timezone.utc).date().isoformat()

    def in_window(date_str: str) -> bool:
        if window == "all":
            return True
        if not date_str:
            return False
        return date_str[:10] >= year_start_iso

    # Sales filter
    query = {"is_deleted": {"$ne": True}}
    if current.get("role") == "sales":
        query["$or"] = [{"assigned_to_user_id": current["id"]}, {"created_by_user_id": current["id"]}]
    deals = await db.deals.find(query, {"_id": 0}).to_list(5000)

    types = ["Repair", "Roof Restoration", "Roof Replacement", "Maintenance", "New Construction", "Other"]
    buckets = {t: {"booked": 0.0, "received": 0.0, "count": 0} for t in types}

    for d in deals:
        # Project type breakdown from Won deals
        if d.get("status") == "Won":
            # Use chosen_date if present, else created_at
            ref = (d.get("chosen_date") or d.get("created_at") or "")[:10]
            if in_window(ref):
                ptype = d.get("project_type") or "Other"
                if ptype not in buckets:
                    ptype = "Other"
                target = buckets[ptype]
                booked = float(d.get("chosen_amount", 0) or 0)
                received = sum(float(m.get("amount", 0) or 0) for m in (d.get("payment_milestones") or []) if m.get("status") == "Paid")
                target["booked"] += booked
                target["received"] += received
                target["count"] += 1

        # Maintenance visits feed the Maintenance bucket regardless of the project's own type
        for v in (d.get("maintenance_visits") or []):
            vd = (v.get("visit_date") or "")[:10]
            if not in_window(vd):
                continue
            amt = float(v.get("amount", 0) or 0)
            buckets["Maintenance"]["booked"] += amt
            buckets["Maintenance"]["received"] += amt  # actuals from logged visits
            buckets["Maintenance"]["count"] += 1

    rows = []
    for t in types:
        b = buckets[t]
        rows.append({
            "project_type": t,
            "booked": round(b["booked"], 2),
            "received": round(b["received"], 2),
            "count": b["count"],
        })

    totals = {
        "booked": round(sum(r["booked"] for r in rows), 2),
        "received": round(sum(r["received"] for r in rows), 2),
    }
    return {"window": window, "rows": rows, "totals": totals}


# ----- Startup -----
@app.on_event("startup")
async def on_startup():
    await db.users.create_index("email", unique=True)
    await db.contacts.create_index("id", unique=True)
    await db.properties.create_index("id", unique=True)
    await db.deals.create_index("id", unique=True)
    await db.vendors.create_index("id", unique=True)
    await db.files.create_index("id", unique=True)
    await db.files.create_index([("parent_type", 1), ("parent_id", 1)])
    await db.invoices.create_index("id", unique=True)
    await db.invoices.create_index("invoice_number", unique=True, sparse=True)
    await db.invoices.create_index("deal_id")
    await db.vendor_bills.create_index("id", unique=True)
    await db.vendor_bills.create_index("vendor_id")
    await db.vendor_bills.create_index("due_date")
    await db.materials.create_index("id", unique=True)
    await db.materials.create_index("sku")
    await db.materials.create_index("name")
    await db.settings.create_index("key", unique=True)

    # Migrate deprecated status "Proposal Sent" → "Sent"
    await db.deals.update_many({"status": "Proposal Sent"}, {"$set": {"status": "Sent"}})

    # Init object storage (non-fatal)
    try:
        init_storage()
        logger.info("Object storage initialized")
    except Exception as e:
        logger.warning(f"Storage init failed (uploads will not work): {e}")

    # Seed Books module — entities + default Chart of Accounts (idempotent)
    try:
        await seed_default_entities(db)
        await gl.ensure_indexes(db)
        import period_close as pc
        await pc.ensure_indexes(db)
        import bank_rec as br
        await br.ensure_indexes(db)
        logger.info("Books entities + COA seeded; GL + period_close + bank_rec indexes ensured")
    except Exception as e:
        logger.warning(f"Books seeding failed: {e}")

    # In-process scheduler (APScheduler) — runs the daily Lead→Sent auto-flip
    # and the Monday Stale-Deals digest without needing a separate cron worker.
    try:
        import scheduler as _sched

        async def _stale_engine(days=14, won_grace_days=30):
            return await _compute_stale_deals(days=days, won_grace_days=won_grace_days)

        async def _send_one_digest(owner_id, deals_for_owner, days, won_grace_days):
            user_doc = await db.users.find_one(
                {"id": owner_id}, {"_id": 0, "id": 1, "email": 1, "name": 1}
            )
            if not user_doc or not user_doc.get("email"):
                return False
            entry = await _build_and_send_owner_digest(
                user=user_doc,
                deals_for_owner=deals_for_owner,
                days=days,
                won_grace_days=won_grace_days,
                cc_email=None,
                dry_run=False,
            )
            return entry.get("sent") is True

        await _sched.start(db=db, stale_engine=_stale_engine, send_one_digest=_send_one_digest)
    except Exception as e:
        logger.warning(f"Scheduler failed to start: {e}")

    # Kick off the COI reminder scheduler (non-blocking background task).
    # Idempotent — uses last_sent_date/next_send_date stored in coi_reminder_settings.
    try:
        await db.coi_reminder_settings.create_index("key", unique=True)
        await db.coi_reminder_history.create_index("sent_at")
        # Assessment indexes
        await db.assessments.create_index("id", unique=True)
        await db.assessments.create_index("deal_id")
        await db.assessments.create_index("status")
        await db.assessments.create_index("created_at")
        asyncio.create_task(coi_reminders.scheduler_loop(db))
        logger.info("COI reminder scheduler started")
    except Exception as e:
        logger.warning(f"COI scheduler failed to start (non-fatal): {e}")

    # Deal-events reminder: every 5 minutes, fire email reminders 1h before
    # any ad-hoc appointment (Roof Walks etc.).
    try:
        from apscheduler.schedulers.asyncio import AsyncIOScheduler
        from apscheduler.triggers.cron import CronTrigger
        await db.deal_events.create_index("deal_id")
        await db.deal_events.create_index("date")
        await db.user_notes.create_index("user_id")
        await db.user_certifications.create_index("user_id")
        await db.user_certifications.create_index("expiration_date")
        await db.user_equipment.create_index("user_id")
        if not hasattr(app.state, "_deal_events_sched"):
            sched = AsyncIOScheduler(timezone="UTC")

            async def _tick():
                try:
                    await deal_events_module.send_due_reminders(db)
                except Exception as e:
                    logger.warning(f"deal_events reminder tick failed: {e}")

            sched.add_job(_tick, CronTrigger(minute="*/5"), id="deal_event_reminders", replace_existing=True)

            # Cert-expiration reminders — daily at 13:30 UTC (≈ 7:30 AM MDT).
            async def _cert_tick():
                try:
                    import user_profile as _up
                    await _up.send_due_cert_reminders(db)
                except Exception as e:
                    logger.warning(f"cert reminder tick failed: {e}")

            sched.add_job(_cert_tick, CronTrigger(hour=13, minute=30), id="cert_expiration_reminders", replace_existing=True)
            sched.start()
            app.state._deal_events_sched = sched
            logger.info("Deal-events reminder scheduler started (every 5 min)")
            logger.info("Cert-expiration reminder scheduler started (daily 13:30 UTC)")
    except Exception as e:
        logger.warning(f"Deal-events reminder scheduler failed (non-fatal): {e}")

    admin_email = os.environ.get("ADMIN_EMAIL", "admin@roofingcrm.com").lower()
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing = await db.users.find_one({"email": admin_email})
    if existing is None:
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "email": admin_email,
            "name": "Darren Oliver",
            "role": "admin",
            "phone": "",
            "title": "Owner",
            "credentials": "CSI, IIBEC",
            "password_hash": hash_password(admin_password),
            "created_at": now_iso(),
        })
    else:
        # Make sure existing admin has role=admin (migration)
        migrate = {"role": "admin"}
        # One-time: backfill default name "Admin" → "Darren Oliver"
        if (existing.get("name") or "").strip() in ("", "Admin"):
            migrate["name"] = "Darren Oliver"
        # Note: credentials are NEVER auto-applied — every rep is responsible
        # for setting their own credentials on the Profile page.
        await db.users.update_one({"id": existing["id"]}, {"$set": migrate})

    # Start the weekly payables-email scheduler
    _start_payables_scheduler()


# ----- Friday Payables Email Scheduler -----
_payables_scheduler = None


def _start_payables_scheduler():
    """Schedule the weekly payables email to admin every Friday 7:00 AM (server timezone = UTC)."""
    global _payables_scheduler
    try:
        from apscheduler.schedulers.asyncio import AsyncIOScheduler
        from apscheduler.triggers.cron import CronTrigger
    except ImportError:
        logger.warning("APScheduler not installed — Friday payables email disabled")
        return
    if _payables_scheduler is not None:
        return

    async def send_friday_report():
        try:
            admin_email = os.environ.get("PAYABLES_REPORT_EMAIL") or os.environ.get("GMAIL_FROM_EMAIL") or os.environ.get("ADMIN_EMAIL")
            if not admin_email:
                logger.warning("No PAYABLES_REPORT_EMAIL or GMAIL_FROM_EMAIL configured")
                return
            # Find the admin user to fake a "current" context
            admin = await db.users.find_one({"role": "admin"})
            if not admin:
                logger.warning("No admin user found for scheduled payables email")
                return
            from email_sender import send_for_category
            report = await payables_report(admin)
            if report["overdue_count"] == 0 and report["due_this_week_count"] == 0:
                logger.info("Friday payables: nothing due, skipping email")
                return
            await send_for_category(
                db, "finance",
                to=admin_email,
                subject=f"SealTech Payables — Week of {report['today']} — {report['overdue_count']} overdue · {report['due_this_week_count']} due",
                body_text=_render_payables_email_text(report),
                body_html=_render_payables_email_html(report),
            )
            logger.info(f"Friday payables email sent to {admin_email}")
        except Exception as e:
            logger.error(f"Friday payables email failed: {type(e).__name__}: {e}")

    sched = AsyncIOScheduler(timezone="America/Denver")  # Aurora/Castle Rock, CO local time
    # Friday 7:00 AM Mountain Time
    sched.add_job(send_friday_report, CronTrigger(day_of_week="fri", hour=7, minute=0), id="payables_friday")
    sched.start()
    _payables_scheduler = sched
    logger.info("Payables scheduler started — Friday 7:00 AM America/Denver")


@app.on_event("shutdown")
async def shutdown_db_client():
    try:
        import scheduler as _sched
        _sched.shutdown()
    except Exception:
        pass
    client.close()


# ----- Scheduler admin endpoints -----
@api_router.get("/scheduler/jobs")
async def list_scheduler_jobs(current=Depends(get_current_user)):
    """List all scheduled background jobs + their next run time. Admin only."""
    if current.get("role") != "admin":
        raise HTTPException(403, "Admins only")
    import scheduler as _sched
    base = {"running": _sched.get_scheduler() is not None, "jobs": _sched.list_jobs()}
    # Enrich each job row with its effective trigger config so the UI editor
    # can render the right controls (time picker + day-of-week chips when supported).
    for row in base["jobs"]:
        try:
            cfg = await _sched._resolve_trigger_config(db, row["id"])
            row["supports_day_of_week"] = bool(cfg.get("supports_day_of_week"))
            row["hour"] = cfg["hour"]
            row["minute"] = cfg["minute"]
            row["day_of_week"] = cfg.get("day_of_week") or "*"
        except Exception:
            pass
    return base


@api_router.post("/scheduler/jobs/{job_id}/run")
async def run_scheduler_job(job_id: str, current=Depends(get_current_user)):
    """Trigger a scheduled job on-demand (admin only). Useful for both manual
    sanity checks and exercising the cron path from a regression test."""
    if current.get("role") != "admin":
        raise HTTPException(403, "Admins only")
    import scheduler as _sched
    try:
        result = await _sched.run_job_now(job_id)
    except KeyError as e:
        raise HTTPException(404, str(e))
    return {"ok": True, "result": result}


@api_router.put("/scheduler/jobs/{job_id}/schedule")
async def update_scheduler_job_schedule(job_id: str, body: dict = Body(...), current=Depends(get_current_user)):
    """Persist a new schedule for a job + re-register the live trigger. Admin only.

    Body shape: { hour: 0..23, minute: 0..59, day_of_week?: "mon" | "mon,fri" | "*" }
    `day_of_week` is honored only on jobs whose default `supports_day_of_week` is True.
    """
    if current.get("role") != "admin":
        raise HTTPException(403, "Admins only")
    import scheduler as _sched
    try:
        cfg = await _sched.reschedule_job(
            db,
            job_id,
            hour=body.get("hour"),
            minute=body.get("minute"),
            day_of_week=body.get("day_of_week"),
        )
    except KeyError as e:
        raise HTTPException(404, str(e))
    except (ValueError, TypeError) as e:
        raise HTTPException(400, str(e))
    return {"ok": True, "config": cfg}


# ----- Router & CORS -----
api_router.include_router(make_books_router(db, get_current_user, require_admin))
api_router.include_router(coi_reminders.create_router(db, require_admin))
api_router.include_router(project_photos.create_router(db, get_current_user))
api_router.include_router(project_photos.create_public_router(db))
api_router.include_router(trash.create_router(db, require_admin))
api_router.include_router(assessment_module.create_router(db, get_current_user))

# Google Calendar integration + Tasks (sync target)
import google_calendar as gcal_module
import tasks as tasks_module
import deal_events as deal_events_module
import email_routing as email_routing_module
import user_profile as user_profile_module
_PUBLIC_BASE_URL = os.environ.get("PUBLIC_BASE_URL") or os.environ.get("GOOGLE_OAUTH_REDIRECT_URI", "").replace("/api/oauth/calendar/callback", "")
api_router.include_router(gcal_module.make_google_calendar_router(db, get_current_user, _PUBLIC_BASE_URL))
api_router.include_router(gcal_module.make_google_oauth_callback_router(db))
api_router.include_router(tasks_module.make_tasks_router(db, get_current_user, push_task_fn=gcal_module.push_task, public_base_url=_PUBLIC_BASE_URL))
api_router.include_router(deal_events_module.make_router(db, get_current_user, public_base_url=_PUBLIC_BASE_URL))
api_router.include_router(email_routing_module.make_router(db, get_current_user))
api_router.include_router(user_profile_module.make_router(db, get_current_user, require_admin, public_base_url=_PUBLIC_BASE_URL))
import product_catalog as _product_catalog
api_router.include_router(_product_catalog.create_router(db, get_current_user))

import brochures as _brochures
api_router.include_router(_brochures.create_router(get_current_user))

import work_orders as _work_orders
# PUBLIC_BASE_URL drives the sign-link prefix in outbound emails. Falls back to
# the frontend's REACT_APP_BACKEND_URL (read from /app/frontend/.env) so the
# subcontractor can click the link from their inbox and land on the SPA.
PUBLIC_BASE_URL = os.environ.get("PUBLIC_BASE_URL", "")
if not PUBLIC_BASE_URL:
    try:
        with open("/app/frontend/.env") as _fe:
            for _ln in _fe:
                if _ln.startswith("REACT_APP_BACKEND_URL"):
                    PUBLIC_BASE_URL = _ln.split("=", 1)[1].strip()
                    break
    except OSError:
        PUBLIC_BASE_URL = ""
api_router.include_router(_work_orders.create_router(db, get_current_user, PUBLIC_BASE_URL))
api_router.include_router(_work_orders.create_public_router(db, PUBLIC_BASE_URL))


# ----- Public Proposal Signing (Sign Off link) ----------------------------
import proposal_signing as _proposal_signing


async def _compute_scope_for_signing(deal_id: str) -> dict:
    """Adapter passed into proposal_signing — returns effective scope bullets +
    client info so the public viewer can render the proposal without hitting
    any other endpoint."""
    deal = await db.deals.find_one({"id": deal_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not deal:
        return {}
    from spec_sheet import _resolve_template, _apply_scope_overrides
    base = _resolve_template(deal.get("proposed_roof_type"), deal.get("current_roof_type"))
    eff = _apply_scope_overrides(base, deal.get("scope_overrides") or {})

    # Pull client (Bill-To) name + address from the linked property + contact
    client_name = client_company = client_address = client_city = client_state = client_zip = ""
    if deal.get("property_id"):
        prop = await db.properties.find_one({"id": deal["property_id"]}, {"_id": 0})
        if prop:
            client_company = prop.get("property_name") or ""
            client_address = " ".join([prop.get("property_address", ""), prop.get("property_address_line2", "")]).strip()
            client_city = prop.get("property_city", "")
            client_state = prop.get("property_state", "")
            client_zip = prop.get("property_zip", "")
    cid = deal.get("customer_contact_id") or deal.get("contact_id")
    if cid:
        cust = await db.contacts.find_one({"id": cid}, {"_id": 0})
        if cust:
            client_name = cust.get("contact_name", "") or ""

    return {
        "client_name": client_name,
        "client_company": client_company,
        "client_address": client_address,
        "client_city": client_city,
        "client_state": client_state,
        "client_zip": client_zip,
        "scope_title": eff.get("title", ""),
        "scope_1_title": eff.get("scope_1_title", ""),
        "scope_1": eff.get("scope_1", []),
        "scope_2_title": eff.get("scope_2_title", ""),
        "scope_2": eff.get("scope_2", []),
        "key_advantages": eff.get("key_advantages", []),
    }


async def _auto_create_deposit_invoice(deal_id: str, percentage: float = 50.0) -> dict | None:
    """Auto-spawn a Draft deposit invoice when the proposal is signed.

    Returns the created invoice dict, or None when there is no sensible amount
    to invoice (project total <= 0) OR when a Deposit invoice already exists on
    this deal (idempotent — re-signing won't double up).

    Mirrors the side-effects of POST /invoices: auto-numbering, GL hooks,
    bill-to prefill from contact + property.
    """
    deal = await db.deals.find_one({"id": deal_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if deal is None:
        return None

    # Idempotency: skip if a Draft/Sent/Partial Deposit invoice already exists
    existing = await db.invoices.find_one(
        {
            "deal_id": deal_id,
            "invoice_type": "Deposit",
            "is_deleted": {"$ne": True},
            "status": {"$nin": ["Void"]},
        },
        {"_id": 0, "id": 1, "invoice_number": 1, "status": 1, "amount_paid": 1},
    )
    if existing is not None:
        return existing

    # Resolve the contract total — prefer chosen_amount, else MID proposal option.
    contract_total = float(deal.get("chosen_amount") or 0)
    if contract_total <= 0:
        contract_total = proposal_mid_amount(deal)
    if contract_total <= 0:
        return None  # nothing to invoice

    deposit_amount = round(contract_total * (percentage / 100.0), 2)
    if deposit_amount <= 0:
        return None

    invoice_today = datetime.now(timezone.utc).date().isoformat()
    bill_to: dict = {}
    customer_id = deal.get("customer_contact_id") or deal.get("contact_id")
    if customer_id:
        bill_to = await _build_bill_to_from_contact(customer_id) or {}

    # Resolve a project address from the linked property
    project_address = ""
    if deal.get("property_id"):
        prop = await db.properties.find_one({"id": deal["property_id"]}, {"_id": 0})
        if prop:
            addr1 = " ".join(
                [p for p in [prop.get("property_address", ""), prop.get("property_address_line2", "")] if p]
            ).strip()
            line2 = ", ".join(
                [p for p in [prop.get("property_city", ""), prop.get("property_state", "")] if p]
            )
            if prop.get("property_zip"):
                line2 = f"{line2} {prop.get('property_zip')}".strip()
            project_address = "  ·  ".join([p for p in [addr1, line2] if p])

    title = deal.get("title") or "Project"
    pct_label = f"{int(percentage)}%" if float(percentage).is_integer() else f"{percentage:g}%"
    description = f"{title} — {pct_label} Deposit (signed by customer)"

    data = {
        "id": str(uuid.uuid4()),
        "deal_id": deal_id,
        "customer_contact_id": customer_id or None,
        "invoice_type": "Deposit",
        "bill_to_company": bill_to.get("bill_to_company") or "",
        "bill_to_name": bill_to.get("bill_to_name") or "",
        "bill_to_address": bill_to.get("bill_to_address") or "",
        "bill_to_address_line2": bill_to.get("bill_to_address_line2") or "",
        "bill_to_city": bill_to.get("bill_to_city") or "",
        "bill_to_state": bill_to.get("bill_to_state") or "",
        "bill_to_zip": bill_to.get("bill_to_zip") or "",
        "bill_to_email": bill_to.get("bill_to_email") or "",
        "cc_email": "",
        "invoice_date": invoice_today,
        "due_date": invoice_today,
        "terms": "Due Upon Receipt",
        "project_title": title,
        "project_address": project_address,
        "project_total": contract_total,
        "notes": f"Auto-generated on proposal acceptance. Edit any field before sending.",
        "line_items": [
            {
                "description": description,
                "quantity": 1,
                "unit_price": deposit_amount,
                "amount": deposit_amount,
            }
        ],
        "status": "Draft",
        "amount_paid": 0.0,
        "payment_date": "",
        "payment_method": "",
        "payment_reference": "",
        "source_type": "proposal_signing",
        "source_id": deal_id,
        "created_at": now_iso(),
        "created_by_user_id": "public-sign",
        "is_deleted": False,
        "invoice_number": await _next_invoice_number(),
    }
    data = _recalc_invoice(data)
    await db.invoices.insert_one(data.copy())
    # Books auto-journal — best effort, never block the sign flow
    try:
        await gl.post_invoice_issue(db, data, posted_by_user_id="public-sign")
    except Exception as e:
        logger.warning(f"GL post (auto-deposit invoice) failed: {type(e).__name__}: {e}")
    return strip_id(data)


async def _build_signed_scope_pdf_public(deal: dict) -> bytes:
    """Public-side spec sheet builder for /public/proposal/{token}/pdf.
    Resolves the deal owner so the "presented by" cover info matches what
    the customer originally received over email."""
    owner_id = deal.get("assigned_to_id") or deal.get("created_by_id")
    user_doc = None
    if owner_id:
        user_doc = await db.users.find_one({"id": owner_id}, {"_id": 0})
    return await _build_spec_pdf_for_deal(deal, user_doc or {})


api_router.include_router(_proposal_signing.create_public_router(
    db,
    get_current_user,
    _compute_scope_for_signing,
    _auto_create_deposit_invoice,
    build_signed_pdf_fn=_build_signed_scope_pdf_public,
))


# ---------- User Guide PDFs ----------
import user_guide_pdf as _user_guide_pdf  # noqa: E402


@api_router.get("/docs/quick-guide.pdf")
async def docs_quick_guide_pdf(_=Depends(get_current_user)):
    """SealTech CRM Quick Reference — 2–3 page laminate-on-the-truck cheat sheet."""
    pdf = _user_guide_pdf.build_quick_guide_pdf()
    return StreamingResponse(
        io.BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="SealTech CRM - Quick Reference.pdf"'},
    )


@api_router.get("/docs/full-manual.pdf")
async def docs_full_manual_pdf(_=Depends(get_current_user)):
    """SealTech CRM Full User Manual — every screen, button, workflow."""
    pdf = _user_guide_pdf.build_full_manual_pdf()
    return StreamingResponse(
        io.BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="SealTech CRM - Full User Manual.pdf"'},
    )


# ---------- Daily Status Report (the "morning standup" PDF) ----------
import daily_status_pdf as _daily_status_pdf  # noqa: E402


async def collect_daily_status_data(_db=None) -> dict:
    """Gather everything the Daily Status PDF needs in a single pass.

    Public so the APScheduler cron in scheduler.py can call it without
    re-implementing the queries. Returns kwargs suitable for
    `daily_status_pdf.build_daily_status_pdf(**payload)`.
    """
    target_db = _db if _db is not None else db
    now = datetime.now(timezone.utc)
    today = now.date().isoformat()
    tomorrow = (now + timedelta(days=1)).date().isoformat()

    # Deals — exclude deleted/lost
    deals = await target_db.deals.find(
        {"is_deleted": {"$ne": True}}, {"_id": 0}
    ).to_list(5000)
    deal_ids = [d["id"] for d in deals]

    # Invoices grouped by deal (for "final invoice paid?" check)
    invs = await target_db.invoices.find(
        {"deal_id": {"$in": deal_ids}, "is_deleted": {"$ne": True}},
        {"_id": 0, "deal_id": 1, "status": 1, "is_final": 1, "balance_due": 1, "invoice_type": 1},
    ).to_list(5000) if deal_ids else []
    invoices_by_deal: dict = {}
    for inv in invs:
        inv["is_final"] = inv.get("is_final") or inv.get("invoice_type") == "Final"
        invoices_by_deal.setdefault(inv["deal_id"], []).append(inv)

    # Users
    users = await target_db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(500)
    users_by_id = {u["id"]: u for u in users}

    # Today's + Tomorrow's ad-hoc deal events (the ones from deal_events.py)
    today_events = await target_db.deal_events.find(
        {"is_deleted": {"$ne": True}, "date": today}, {"_id": 0},
    ).sort([("start_time", 1)]).to_list(200)
    tomorrow_events = await target_db.deal_events.find(
        {"is_deleted": {"$ne": True}, "date": tomorrow}, {"_id": 0},
    ).sort([("start_time", 1)]).to_list(200)
    # Attach deal_title for nicer rendering
    evt_deal_ids = list({ev["deal_id"] for ev in (today_events + tomorrow_events) if ev.get("deal_id")})
    if evt_deal_ids:
        evt_deals = await target_db.deals.find(
            {"id": {"$in": evt_deal_ids}}, {"_id": 0, "id": 1, "title": 1},
        ).to_list(len(evt_deal_ids))
        title_by_id = {d["id"]: d.get("title") or "" for d in evt_deals}
        for ev in today_events + tomorrow_events:
            ev["deal_title"] = title_by_id.get(ev.get("deal_id"), "")

    # Overdue tasks
    overdue_tasks = await target_db.tasks.find(
        {
            "is_deleted": {"$ne": True},
            "status": {"$ne": "Done"},
            "due_date": {"$ne": "", "$lt": today},
        },
        {"_id": 0, "title": 1, "due_date": 1, "assigned_to_user_id": 1, "deal_id": 1},
    ).sort([("due_date", 1)]).to_list(200)

    # Stale deals (>= 7 days no movement)
    stale_threshold = (now - timedelta(days=7))
    stale = []
    for d in deals:
        if d.get("status") in ("Lost", "Past Lead"):
            continue
        hist = d.get("status_history") or []
        last_iso = hist[-1].get("at") if hist else d.get("updated_at") or d.get("created_at")
        try:
            last_dt = datetime.fromisoformat(str(last_iso).replace("Z", "+00:00"))
        except Exception:
            continue
        if last_dt <= stale_threshold:
            stale.append({
                "id": d["id"],
                "title": d.get("title") or "Untitled",
                "status": d.get("status"),
                "days_in_stage": (now - last_dt).days,
                "owner_user_id": d.get("assigned_to_user_id") or d.get("created_by_user_id") or "",
            })
    stale.sort(key=lambda x: x["days_in_stage"], reverse=True)

    # COIs expiring within 30 days (vendors)
    coi_cutoff = (now + timedelta(days=30)).date().isoformat()
    coi_expiring = await target_db.vendors.find(
        {
            "is_deleted": {"$ne": True},
            "coi_expiration": {"$ne": "", "$lte": coi_cutoff, "$gte": today},
        },
        {"_id": 0, "name": 1, "vendor_name": 1, "contact_name": 1, "coi_expiration": 1},
    ).sort([("coi_expiration", 1)]).to_list(200) if "vendors" in await target_db.list_collection_names() else []

    return {
        "deals": deals,
        "invoices_by_deal": invoices_by_deal,
        "users_by_id": users_by_id,
        "today_events": today_events,
        "tomorrow_events": tomorrow_events,
        "overdue_tasks": overdue_tasks,
        "coi_expiring_soon": coi_expiring,
        "stale_deals": stale[:25],  # cap so PDF stays readable
        "now": now,
    }


@api_router.get("/reports/daily-status.pdf")
async def reports_daily_status_pdf(current=Depends(get_current_user)):
    """On-demand Daily Status PDF — the same engine the 7am cron uses.

    Any logged-in user can pull it; the contents are company-wide today
    (admin's morning standup view). Per-user filtered views are a future
    enhancement.
    """
    payload = await collect_daily_status_data(db)
    pdf = _daily_status_pdf.build_daily_status_pdf(**payload)
    filename = f"daily-status-{datetime.now(timezone.utc).date().isoformat()}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.get("/reports/daily-status/recipients")
async def reports_daily_status_recipients(current=Depends(get_current_user)):
    """Who the 7am morning email goes to — admin + every active deal owner."""
    if current.get("role") != "admin":
        raise HTTPException(403, "Admins only")
    import scheduler as _sched
    return {"recipients": await _sched._resolve_daily_status_recipients(db)}


# ---------- Final Invoice (project completion) ----------
async def _compute_final_invoice_preview(deal_id: str) -> dict:
    """Pure-read calculation of what a Final invoice would look like.

    Returns {contract_total, already_invoiced, final_amount,
             existing_final_invoice_id?, has_deal}.

    `contract_total` = chosen_amount (or MID proposal fallback) + approved
    change-orders. `already_invoiced` = sum of all non-void, non-deleted
    invoices' total_amount on this deal — including drafts, since drafts
    are queued-to-bill. Excludes any pre-existing Final invoice so the
    preview tells the user what to expect _next time they hit the button_.
    """
    deal = await db.deals.find_one(
        {"id": deal_id, "is_deleted": {"$ne": True}}, {"_id": 0},
    )
    if deal is None:
        return {"has_deal": False, "contract_total": 0.0, "already_invoiced": 0.0, "final_amount": 0.0}

    contract_total = float(deal.get("chosen_amount") or 0)
    if contract_total <= 0:
        contract_total = proposal_mid_amount(deal)
    co_total = sum(
        float(co.get("amount", 0) or 0)
        for co in (deal.get("change_orders") or [])
        if (co.get("status") or "Approved") == "Approved"
    )
    contract_total = round(contract_total + co_total, 2)

    # Existing Final invoice on this deal (non-void).
    existing_final = await db.invoices.find_one(
        {"deal_id": deal_id, "invoice_type": "Final",
         "is_deleted": {"$ne": True}, "status": {"$nin": ["Void"]}},
        {"_id": 0, "id": 1, "invoice_number": 1, "status": 1, "total": 1},
    )

    # Already-invoiced excludes the pre-existing Final (so re-running the
    # preview after creation doesn't double-count). Invoice totals live on
    # the `total` field after _recalc_invoice — fall back to `total_amount`
    # for any legacy doc that still uses the older field name.
    q = {"deal_id": deal_id, "is_deleted": {"$ne": True}, "status": {"$nin": ["Void"]}}
    if existing_final:
        q["id"] = {"$ne": existing_final["id"]}
    already_invoiced = 0.0
    async for inv in db.invoices.find(q, {"_id": 0, "total": 1, "total_amount": 1}):
        already_invoiced += float(inv.get("total") or inv.get("total_amount") or 0)
    already_invoiced = round(already_invoiced, 2)

    final_amount = max(0.0, round(contract_total - already_invoiced, 2))
    out = {
        "has_deal": True,
        "contract_total": contract_total,
        "already_invoiced": already_invoiced,
        "final_amount": final_amount,
    }
    if existing_final:
        out["existing_final_invoice_id"] = existing_final["id"]
        out["existing_final_invoice_number"] = existing_final.get("invoice_number")
    return out


async def _auto_create_final_invoice(deal_id: str, user_id: str = "manual") -> dict | None:
    """Draft a Final invoice for the project balance. Idempotent — returns the
    existing non-void Final invoice if one already exists. Returns None when
    there is no balance left to bill (final_amount <= 0)."""
    preview = await _compute_final_invoice_preview(deal_id)
    if not preview.get("has_deal"):
        return None
    if preview.get("existing_final_invoice_id"):
        # Hand back the existing one so callers can re-open it.
        existing = await db.invoices.find_one(
            {"id": preview["existing_final_invoice_id"]}, {"_id": 0},
        )
        return strip_id(existing) if existing else None

    final_amount = preview["final_amount"]
    if final_amount <= 0:
        return None

    deal = await db.deals.find_one({"id": deal_id}, {"_id": 0})
    customer_id = deal.get("customer_contact_id") or deal.get("contact_id")
    bill_to = await _build_bill_to_from_contact(customer_id) if customer_id else {}
    bill_to = bill_to or {}

    project_address = ""
    if deal.get("property_id"):
        prop = await db.properties.find_one({"id": deal["property_id"]}, {"_id": 0})
        if prop:
            addr1 = " ".join(
                [p for p in [prop.get("property_address", ""), prop.get("property_address_line2", "")] if p]
            ).strip()
            line2 = ", ".join(
                [p for p in [prop.get("property_city", ""), prop.get("property_state", "")] if p]
            )
            if prop.get("property_zip"):
                line2 = f"{line2} {prop.get('property_zip')}".strip()
            project_address = "  ·  ".join([p for p in [addr1, line2] if p])

    title = deal.get("title") or "Project"
    description = f"{title} — Final Invoice (project completion)"
    invoice_today = datetime.now(timezone.utc).date().isoformat()

    data = {
        "id": str(uuid.uuid4()),
        "deal_id": deal_id,
        "customer_contact_id": customer_id or None,
        "invoice_type": "Final",
        "bill_to_company": bill_to.get("bill_to_company") or "",
        "bill_to_name": bill_to.get("bill_to_name") or "",
        "bill_to_address": bill_to.get("bill_to_address") or "",
        "bill_to_address_line2": bill_to.get("bill_to_address_line2") or "",
        "bill_to_city": bill_to.get("bill_to_city") or "",
        "bill_to_state": bill_to.get("bill_to_state") or "",
        "bill_to_zip": bill_to.get("bill_to_zip") or "",
        "bill_to_email": bill_to.get("bill_to_email") or "",
        "cc_email": "",
        "invoice_date": invoice_today,
        "due_date": invoice_today,
        "terms": "Due Upon Receipt",
        "project_title": title,
        "project_address": project_address,
        "project_total": preview["contract_total"],
        "notes": (
            "Auto-generated on project completion. Contract total "
            f"${preview['contract_total']:,.2f} minus prior invoices "
            f"${preview['already_invoiced']:,.2f} = balance due "
            f"${final_amount:,.2f}. Edit any field before sending."
        ),
        "line_items": [
            {
                "description": description,
                "quantity": 1,
                "unit_price": final_amount,
                "amount": final_amount,
            }
        ],
        "status": "Draft",
        "amount_paid": 0.0,
        "payment_date": "",
        "payment_method": "",
        "payment_reference": "",
        "source_type": "mark_complete",
        "source_id": deal_id,
        "created_at": now_iso(),
        "created_by_user_id": user_id,
        "is_deleted": False,
        "invoice_number": await _next_invoice_number(),
    }
    data = _recalc_invoice(data)
    await db.invoices.insert_one(data.copy())
    try:
        await gl.post_invoice_issue(db, data, posted_by_user_id=user_id)
    except Exception as e:
        logger.warning(f"GL post (final invoice) failed: {type(e).__name__}: {e}")
    return strip_id(data)


@api_router.get("/deals/{deal_id}/final-invoice/preview")
async def preview_final_invoice(deal_id: str, _=Depends(get_current_user)):
    """Returns the projected Final-Invoice amount without creating anything.
    Used by the Closed-stage suggestion banner so the user sees exactly what
    they'd be billing before they click."""
    out = await _compute_final_invoice_preview(deal_id)
    if not out.get("has_deal"):
        raise HTTPException(status_code=404, detail="Deal not found")
    return out


@api_router.post("/deals/{deal_id}/final-invoice")
async def create_final_invoice(deal_id: str, current=Depends(get_current_user)):
    """Drafts the Final Invoice for project completion. Idempotent — returns
    the existing Final invoice if one already exists. 400 if there is nothing
    left to bill (e.g., the project has been fully invoiced already)."""
    deal = await db.deals.find_one({"id": deal_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if deal is None:
        raise HTTPException(status_code=404, detail="Deal not found")
    inv = await _auto_create_final_invoice(deal_id, user_id=current["id"])
    if inv is None:
        # Decide which message to show — no balance vs no contract total.
        preview = await _compute_final_invoice_preview(deal_id)
        if preview.get("contract_total", 0) <= 0:
            raise HTTPException(
                status_code=400,
                detail="No contract total on this deal yet — pick a proposal option or set chosen_amount first.",
            )
        raise HTTPException(
            status_code=400,
            detail=(
                f"Nothing left to bill — prior invoices totaling "
                f"${preview['already_invoiced']:,.2f} already cover the "
                f"${preview['contract_total']:,.2f} contract."
            ),
        )
    return inv


app.include_router(api_router)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
